"""
Excel (.xlsx) report — styled with section headers, colour-coded KPIs,
banded rows, and frozen panes. Uses openpyxl (already a dependency).
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .analyzer import AnalysisResult


# ── Palette (matches the scientific-style example) ────────────────────────
C_TITLE_BG   = "1F4E79"   # dark blue
C_TITLE_FG   = "FFFFFF"
C_SECTION_BG = "2E75B6"   # mid blue
C_SECTION_FG = "FFFFFF"
C_LABEL_BG   = "DDEBF7"   # very light blue
C_ROW_ALT    = "F2F2F2"   # banding
C_KPI_BG     = "E2EFDA"   # mint green for results
C_KPI_BORDER = "548235"
C_PURITY     = "1F77B4"
C_RECOVERY   = "ED7D31"
C_PRODUCTIV  = "70AD47"
C_BORDER     = "BFBFBF"

_BORDER = Border(*([Side(style="thin", color=C_BORDER)] * 4))


def _fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _set_cell(ws, row: int, col: int, value, *,
              bold: bool = False, fg: str | None = None,
              bg: str | None = None, align: str = "left",
              border: bool = True, font_size: int = 10,
              number_format: str | None = None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", size=font_size, bold=bold,
                  color=fg or "000000")
    if bg:
        c.fill = _fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=isinstance(value, str) and len(str(value)) > 30)
    if border:
        c.border = _BORDER
    if number_format:
        c.number_format = number_format
    return c


def export_xlsx_report(result: AnalysisResult, out_path: Path,
                       source_file: str | None = None) -> Path:
    """Write a styled .xlsx report with summary + per-cycle data."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    p = result.params
    cyc = result.cycles
    n_cyc = len(cyc)
    total_bed_mL = p.steps_per_cycle * p.bed_volume_mL
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Column widths (11 working columns: per-cycle table is the widest)
    widths = [10, 14, 12, 12, 14, 14, 14, 12, 12, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ─── TITLE BAR ─────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    _set_cell(ws, row, 1, "PSA Analyzer — Cycle Summary",
              bold=True, fg=C_TITLE_FG, bg=C_TITLE_BG, align="left",
              font_size=16, border=False)
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    _set_cell(ws, row, 1,
              f"Pressure Swing Adsorption — {p.adsorbent_name}   |   "
              f"Generated: {now}",
              fg="595959", bg="F2F2F2", align="left",
              font_size=9, border=False)
    ws.row_dimensions[row].height = 18
    row += 2

    # ─── SECTION 1: PARAMETERS ─────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    _set_cell(ws, row, 1, "1. Experimental Parameters",
              bold=True, fg=C_SECTION_FG, bg=C_SECTION_BG,
              font_size=12, align="left", border=False)
    ws.row_dimensions[row].height = 22
    row += 1

    params_rows = [
        ("Adsorbent",            p.adsorbent_name),
        ("Temperature",          f"{p.temperature_C:.1f} °C"),
        ("Bed volume (per bed)", f"{p.bed_volume_mL:.3f} mL"),
        ("Steps / cycle",        p.steps_per_cycle),
        ("Total bed volume",     f"{total_bed_mL:.3f} mL"),
        ("Smoothing",            "adaptive (window 3 for cycles 1–4, 5 thereafter)"),
        ("Final values from",    f"cycle {p.final_cycle_start} to last"),
    ]
    if source_file:
        params_rows.insert(0, ("Source file", source_file))

    for i, (k, v) in enumerate(params_rows):
        bg = C_ROW_ALT if i % 2 == 0 else "FFFFFF"
        _set_cell(ws, row, 1, k, bold=True, bg=C_LABEL_BG)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        _set_cell(ws, row, 2, v, bg=bg, align="left")
        row += 1

    row += 1

    # ─── SECTION 2: KPI CARDS ──────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    _set_cell(ws, row, 1,
              f"2. Final Results (mean of cycles "
              f"{int(p.final_cycle_start)}–{len(result.cycles)})",
              bold=True, fg=C_SECTION_FG, bg=C_SECTION_BG,
              font_size=12, align="left", border=False)
    ws.row_dimensions[row].height = 22
    row += 1

    kpis = [
        ("Final Purity",       f"{result.final_purity:.2f}",        "%",                C_PURITY),
        ("Final Recovery",     f"{result.final_recovery:.2f}",      "%",                C_RECOVERY),
        ("Final Productivity", f"{result.final_productivity:.4f}",  "t CO₂/m³·day",     C_PRODUCTIV),
        ("Elapsed Time",       f"{result.total_elapsed_hours:.2f}", "hours",            "595959"),
        ("Total Cycles",       f"{n_cyc}",                          "",                 "595959"),
    ]
    # Each KPI uses 2 cols; labels on top, values below
    label_row = row
    value_row = row + 1
    unit_row = row + 2
    for i, (label, val, unit, color) in enumerate(kpis):
        col = 1 + i * 2 - (1 if i > 0 else 0)  # space cards out a bit
        col = 1 + i * 2
        if col + 1 > 11:
            break
        ws.merge_cells(start_row=label_row, start_column=col,
                       end_row=label_row, end_column=col + 1)
        _set_cell(ws, label_row, col, label, bold=True,
                  fg="595959", bg=C_KPI_BG, align="center", font_size=9)
        ws.merge_cells(start_row=value_row, start_column=col,
                       end_row=value_row, end_column=col + 1)
        _set_cell(ws, value_row, col, val, bold=True,
                  fg=color, bg="FFFFFF", align="center", font_size=18)
        ws.merge_cells(start_row=unit_row, start_column=col,
                       end_row=unit_row, end_column=col + 1)
        _set_cell(ws, unit_row, col, unit,
                  fg="595959", bg="FFFFFF", align="center", font_size=9)
    ws.row_dimensions[label_row].height = 18
    ws.row_dimensions[value_row].height = 30
    ws.row_dimensions[unit_row].height = 16
    row = unit_row + 2

    # ─── SECTION 3: STATISTICS ─────────────────────────────────────────────
    stats = getattr(result, "final_stats", None)
    if stats:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        _set_cell(ws, row, 1,
                  "3. Statistics (over the averaging window; "
                  "CV = Std ÷ Mean × 100)",
                  bold=True, fg=C_SECTION_FG, bg=C_SECTION_BG,
                  font_size=12, align="left", border=False)
        ws.row_dimensions[row].height = 22
        row += 1

        st_headers = ["Metric", "Mean", "Std", "Min", "Max", "CV [%]"]
        for i, h in enumerate(st_headers, start=1):
            _set_cell(ws, row, i, h, bold=True, fg=C_TITLE_FG,
                      bg=C_SECTION_BG, align="center")
        row += 1
        st_rows = [("Purity [%]", "purity", "0.00", C_PURITY),
                   ("Recovery [%]", "recovery", "0.00", C_RECOVERY),
                   ("Productivity", "productivity", "0.0000", C_PRODUCTIV)]
        for j, (label, key, nf, color) in enumerate(st_rows):
            bg = C_ROW_ALT if j % 2 == 0 else "FFFFFF"
            s = stats.get(key, {})
            _set_cell(ws, row, 1, label, bold=True, fg=color, bg=bg, align="left")
            for c_i, sk, fmt in [(2, "mean", nf), (3, "std", nf),
                                 (4, "min", nf), (5, "max", nf), (6, "cv", "0.0")]:
                v = s.get(sk, float("nan"))
                if v != v:   # NaN
                    _set_cell(ws, row, c_i, "—", bg=bg, align="center")
                else:
                    _set_cell(ws, row, c_i, float(v), bg=bg,
                              align="center", number_format=fmt)
            row += 1
        row += 1
        per_cycle_no = 4
    else:
        per_cycle_no = 3

    # ─── PER-CYCLE TABLE ───────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    _set_cell(ws, row, 1, f"{per_cycle_no}. Per-Cycle Data "
                          f"(Recovery shown raw + capped at 100%)",
              bold=True, fg=C_SECTION_FG, bg=C_SECTION_BG,
              font_size=12, align="left", border=False)
    ws.row_dimensions[row].height = 22
    row += 1

    headers = ["Cycle", "Elapsed [s]", "dt [s]", "dt [min]",
               "Σ CO₂ out", "Σ N₂ out", "Σ CO₂ in",
               "Purity [%]", "Recovery [%]", "Recov cap [%]", "Productivity"]
    col_keys = ["cycle_id", "elapsed_s", "dt_s", "dt_min",
                "sum_co2_out", "sum_n2_out", "sum_co2_in",
                "purity_pct", "recovery_pct", "recovery_corrected_pct",
                "productivity_tCO2_m3_day"]
    for i, h in enumerate(headers, start=1):
        _set_cell(ws, row, i, h, bold=True, fg=C_TITLE_FG,
                  bg=C_SECTION_BG, align="center", font_size=10)
    ws.row_dimensions[row].height = 22
    header_row = row
    row += 1

    has_cap = "recovery_corrected_pct" in cyc.columns
    rec_over = "C00000"   # red for raw recovery > 100%
    for r_i, (_, cyc_row) in enumerate(cyc.iterrows()):
        bg = C_ROW_ALT if r_i % 2 == 0 else "FFFFFF"
        for c_i, key in enumerate(col_keys, start=1):
            if key == "recovery_corrected_pct" and not has_cap:
                val = min(float(cyc_row["recovery_pct"]), 100.0)
                _set_cell(ws, row, c_i, val, bg=bg,
                          align="center", number_format="0.00")
                continue
            if key not in cyc.columns:
                _set_cell(ws, row, c_i, "—", bg=bg, align="center")
                continue
            val = cyc_row[key]
            if key == "cycle_id":
                _set_cell(ws, row, c_i, int(val), bg=bg,
                          align="center", bold=True)
            elif key in ("purity_pct", "recovery_pct",
                         "recovery_corrected_pct"):
                over = key == "recovery_pct" and float(val) > 100.0
                _set_cell(ws, row, c_i, float(val), bg=bg,
                          align="center", number_format="0.00",
                          bold=over, fg=rec_over if over else None)
            elif key == "productivity_tCO2_m3_day":
                _set_cell(ws, row, c_i, float(val), bg=bg,
                          align="center", number_format="0.0000")
            elif key in ("elapsed_s", "dt_s"):
                _set_cell(ws, row, c_i, float(val), bg=bg,
                          align="center", number_format="0")
            elif key == "dt_min":
                _set_cell(ws, row, c_i, float(val), bg=bg,
                          align="center", number_format="0.00")
            else:
                _set_cell(ws, row, c_i, float(val), bg=bg,
                          align="center", number_format="0.000")
        row += 1

    # Footnote: clarify capped usage.
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    _set_cell(ws, row, 1,
              "Note: red Recovery values exceed 100% (raw). Final Recovery "
              "and the statistics use the raw (uncapped) values; the "
              "'Recov cap' column is display-only.",
              fg="595959", bg="FFFFFF", align="left", font_size=9, border=False)

    # Freeze header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    wb.save(out_path)
    return out_path
