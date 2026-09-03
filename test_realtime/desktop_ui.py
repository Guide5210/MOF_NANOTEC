#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
============================================================================
 desktop_ui.py  v3 — EC MEASUREMENT STATION : desktop companion
============================================================================
 หน้าจอ Desktop ของสถานีวัด EC — ใช้ design system ชุดเดียวกับจอสัมผัส
 ESP32-P4 (light scientific instrument)  ทุกสีและทุกคำสถานะมาจาก lab_theme.py

 v3 เปลี่ยนอะไร (เฉพาะการแสดงผล)
 -------------------------------
   - ธีมสว่างแบบเครื่องมือห้องแล็บ แทนธีมเข้มเดิม
   - คำสถานะ 8 แบบตรงกับจอสัมผัส  LIVE / CHANGING / STEADY / STALE /
     NO RESPONSE / SENSOR FAULT / DISABLED / OFFLINE
     ของเดิมมีแค่ LIVE กับ ERROR ทำให้หัววัดที่ "ตั้งใจถอดออก" ขึ้นแดงตลอดเวลา
   - รองรับ 1-4 เซนเซอร์ตาม active_mask ใน ec_ui_config.json
   - กราฟไม่กะพริบ และไม่ล้าง zoom ที่ผู้ใช้ตั้งไว้ (set_data แทน ax.clear)
   - แผงเหตุการณ์ (EVENT LOG)
   - ไม่มีเลขสีดิบในไฟล์นี้เลย ทุกสีมาจาก lab_theme.py

 v3 "ไม่" เปลี่ยนแม้แต่บรรทัดเดียว
 ---------------------------------
   list_files() _num() read_range() read_recent() downsample() make_mock_csv()
   load_sessions() save_sessions() export_csv() _stats() export_excel() export_pdf()

   ตรวจได้ด้วย:
       diff <(sed -n '51,258p' desktop_ui.py.bak) <(sed -n '51,258p' desktop_ui.py)

 ยังคงเป็น view-only เหมือนเดิม: ไม่เปิด serial ไม่แตะพอร์ตของ CONTROL
 (logger_3ec.py ยังเป็นเจ้าของพอร์ตเพียงตัวเดียว)

 ติดตั้ง:  pip install matplotlib openpyxl
 รัน:      python desktop_ui.py
============================================================================
"""

import os
import glob
import json
import random
from datetime import datetime, timedelta

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg, NavigationToolbar2Tk)
import matplotlib.dates as mdates
from matplotlib.ticker import MaxNLocator

import lab_theme as T

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "water_data")
SESSION_FILE = os.path.join(BASE_DIR, "sessions_3ec.json")   # terminal logger เขียน
REC_STATUS_FILE = os.path.join(BASE_DIR, "rec_status.json")  # สถานะบันทึกสด
CAL_STATUS_FILE = os.path.join(BASE_DIR, "cal_status.json")  # Phase 3b — ยังไม่มีก็ได้
LOGGER_SERVICE = "water-logger"

# ชื่อเดิมที่โค้ดส่วนข้อมูลอ้างถึง (export_pdf ใช้ COLORS) — ตอนนี้มาจาก lab_theme
# ต้องคงชื่อไว้ ไม่งั้นต้องไปแก้ฟังก์ชัน export ซึ่งรอบนี้ห้ามแตะ
COLORS = T.SERIES
BG    = T.BG
CARD  = T.SURFACE
TXT   = T.TEXT
MUT   = T.TEXT_DIM
GREEN = T.OK
RED   = T.ERROR

MAX_SENSORS = T.MAX_SENSORS

# จำนวนแผงกราฟสูงสุดที่ยังอ่านออกในโหมด Split
#
# ⚠️ ที่ความสูงหน้าต่างมาตรฐาน พื้นที่กราฟเหลือราว 200 px
#    หาร 4 แผงได้แผงละ 50 px ซึ่งเหลือที่ให้เส้นจริง ๆ ไม่ถึง 20 px
#    และตัวเลขบนแกน y ซ้อนกันจนอ่านไม่ออก  เกินจำนวนนี้จึงบังคับใช้ Overlay
#    (เหตุผลเดียวกับที่หน้า Measure ของจอใช้ 2 หน้า x 2 ช่อง แทน 4 ช่องเรียง)
MAX_SPLIT_PANELS = 3

# ขึ้นแถบ "SENSOR BOARD OFFLINE" เมื่อไม่มีแถวใหม่นานเท่านี้
# กว้างกว่า STALE_AFTER_S ของรายเซนเซอร์ เพราะอันนี้พูดถึงทั้งบอร์ด
STALE_BANNER_S = 20.0

# ข้อความในช่องเลือก session เมื่อยังไม่ได้เลือกอันไหน
ALL_DATA = "All data — time range"

APP_VERSION = T.UI_VERSION


def _banner():
    """พิมพ์ว่ากำลังรันไฟล์ไหนอยู่จริง

    ⚠️ มีไว้เพราะเคยเจอมาแล้ว: แก้ไฟล์ในโฟลเดอร์หนึ่ง แต่สิ่งที่ถูกสั่งรัน
       เป็นสำเนาอีกที่ (shortcut เก่า / run config ของ IDE / หน้าต่างเดิมที่
       ยังเปิดค้าง)  แล้วไล่หาสาเหตุผิดทางอยู่นาน
       บรรทัดนี้ตัดปัญหานั้นทิ้งถาวร — มันบอก path เต็มของไฟล์ที่ทำงานอยู่
    """
    import sys
    print("=" * 68)
    print(" EC MEASUREMENT STATION — desktop dashboard  [{}]".format(APP_VERSION))
    print("=" * 68)
    print("  ไฟล์ที่รันอยู่ : {}".format(os.path.abspath(__file__)))
    print("  ธีม           : {}".format(os.path.abspath(T.__file__)))
    print("  python        : {} ({})".format(
        sys.version.split()[0], os.path.basename(sys.executable)))
    print("  ค่าตั้งการแสดงผล: {}".format(T.UI_CONFIG_FILE))
    print("-" * 68)
    print("  ถ้าหน้าต่างที่ขึ้นมาเป็นธีมเข้มแบบเดิม แปลว่าคุณกำลังรัน")
    print("  ไฟล์คนละตัวกับที่เห็น path ข้างบนนี้")
    print("=" * 68)
# ============================================================================
#  อ่าน / เขียน CSV  (ฟังก์ชันล้วน ไม่พึ่ง tkinter -> ทดสอบแยกได้)
# ============================================================================
def list_files():
    return sorted(glob.glob(os.path.join(DATA_DIR, "water_log_*.csv")))


def _num(x):
    try:
        return float(x)
    except (ValueError, TypeError):
        return None


def read_range(start=None, end=None):
    """อ่านแถวในช่วง [start,end] (datetime); None = ไม่จำกัดด้านนั้น"""
    rows = []
    for path in list_files():
        try:
            with open(path, encoding="utf-8") as fh:
                fh.readline()
                for line in fh:
                    p = line.rstrip("\n").split(",")
                    if len(p) < 10:
                        continue
                    if len(p) > 10 and p[10].strip().upper() == "CAL":
                        continue          # ข้ามค่าที่เก็บตอนคาลิเบรต
                    try:
                        t = datetime.strptime(p[0], "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        continue
                    if start and t < start:
                        continue
                    if end and t > end:
                        continue
                    rows.append({
                        "t": t,
                        "ec": [_num(p[1]), _num(p[3]), _num(p[5])],
                        "tw": [_num(p[2]), _num(p[4]), _num(p[6])],
                        "ok": [p[7], p[8], p[9]],
                    })
        except Exception:
            continue
    rows.sort(key=lambda r: r["t"])
    return rows


def read_recent(minutes):
    if minutes is None:
        return read_range()
    return read_range(start=datetime.now() - timedelta(minutes=minutes))


def downsample(rows, maxpts=3000):
    step = max(1, len(rows) // maxpts)
    return rows[::step]


def make_mock_csv(minutes=120):
    """สร้างไฟล์ CSV จำลอง 3 ตัว สำหรับทดสอบ UI/Export (ไม่ต้องต่อ ESP32)"""
    os.makedirs(DATA_DIR, exist_ok=True)
    now = datetime.now()
    t0 = now - timedelta(minutes=minutes)
    # แยกตามวันเพื่อให้เหมือนของจริง
    by_day = {}
    n = minutes * 30                       # ทุก 2 วิ
    base = [1400.0, 300.0, 50.0]           # EC เริ่มต้น 3 ภาชนะต่างกัน
    for k in range(n):
        t = t0 + timedelta(seconds=k * 2)
        day = t.strftime("%Y-%m-%d")
        ec = []
        for i in range(3):
            # ภาชนะ 1 ลดลง (ล้าง), 2 ทรงตัว, 3 ค่อยขึ้น + noise
            drift = [-0.05, 0.0, 0.02][i] * k
            ec.append(max(2.0, base[i] + drift + random.gauss(0, 1.5)))
        tw = [20.5 + random.gauss(0, 0.1) for _ in range(3)]
        ok = ["1", "1", "1"]
        if 400 < k < 460:                  # จำลองตัว #2 หลุดช่วงสั้น ๆ
            ec[1] = None; tw[1] = None; ok[1] = "0"
        by_day.setdefault(day, []).append((t, ec, tw, ok))

    for day, recs in by_day.items():
        path = os.path.join(DATA_DIR, f"water_log_{day}.csv")
        with open(path, "w", newline="", encoding="utf-8") as f:
            f.write("timestamp,EC1,T1,EC2,T2,EC3,T3,ok1,ok2,ok3\n")
            for t, ec, tw, ok in recs:
                cells = []
                for i in range(3):
                    cells += ["" if ec[i] is None else f"{ec[i]:.1f}",
                              "" if tw[i] is None else f"{tw[i]:.1f}"]
                f.write(f"{t:%Y-%m-%d %H:%M:%S}," + ",".join(cells) + "," + ",".join(ok) + "\n")
    return sum(len(v) for v in by_day.values())


# ---- session (ช่วงบันทึกที่ผู้ใช้กำหนด) ----
def load_sessions():
    try:
        with open(SESSION_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def save_sessions(sessions):
    try:
        with open(SESSION_FILE, "w", encoding="utf-8") as f:
            json.dump(sessions, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


# ---- export (ฟังก์ชันล้วน) ----
def export_csv(rows, path):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        f.write("timestamp,EC1,T1,EC2,T2,EC3,T3,ok1,ok2,ok3\n")
        for r in rows:
            cells = []
            for i in range(3):
                cells += ["" if r["ec"][i] is None else f"{r['ec'][i]:.1f}",
                          "" if r["tw"][i] is None else f"{r['tw'][i]:.1f}"]
            f.write(f"{r['t']:%Y-%m-%d %H:%M:%S}," + ",".join(cells) +
                    "," + ",".join(r["ok"]) + "\n")


def _stats(rows, i):
    import statistics as st
    vals = [r["ec"][i] for r in rows if r["ec"][i] is not None]
    if not vals:
        return None
    return {
        "n": len(vals), "mean": st.mean(vals),
        "min": min(vals), "max": max(vals),
        "sd": st.pstdev(vals) if len(vals) > 1 else 0.0,
    }


def export_excel(rows, path, title="EC session", meta=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    if meta is None:
        meta = {}
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "Water Quality Report — EC (3 samples)"
    ws["A1"].font = Font(size=14, bold=True, color="1F5C99")
    ws["A3"] = "Generated"; ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if rows:
        ws["A4"] = "Span"; ws["B4"] = f"{rows[0]['t']:%Y-%m-%d %H:%M:%S} - {rows[-1]['t']:%Y-%m-%d %H:%M:%S}"
    ws["A5"] = "Points"; ws["B5"] = len(rows)
    if (meta.get("note") or "").strip():
        ws["A6"] = "Note"; ws["B6"] = meta["note"].strip()
        ws["A6"].font = Font(bold=True)
    hdr = ["Sample", "n", "Mean", "Min", "Max", "SD"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(7, c, h); cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F5C99")
    r = 8
    for i in range(3):
        s = _stats(rows, i)
        if s:
            for c, v in enumerate(["#%d" % (i+1), s["n"], round(s["mean"], 1),
                                   round(s["min"], 1), round(s["max"], 1),
                                   round(s["sd"], 2)], 1):
                ws.cell(r, c, v)
            r += 1
    for col, w in zip("ABCDEF", [10, 8, 10, 10, 10, 8]):
        ws.column_dimensions[col].width = w
    # raw
    wr = wb.create_sheet("raw_data")
    wr.append(["timestamp", "EC1", "T1", "EC2", "T2", "EC3", "T3", "ok1", "ok2", "ok3"])
    for row in rows:
        wr.append([row["t"].strftime("%Y-%m-%d %H:%M:%S"),
                   row["ec"][0], row["tw"][0], row["ec"][1], row["tw"][1],
                   row["ec"][2], row["tw"][2], *row["ok"]])
    wb.save(path)


def export_pdf(rows, path):
    from matplotlib.backends.backend_pdf import PdfPages
    ts = [r["t"] for r in rows]
    with PdfPages(path) as pdf:
        # หน้า 1: 3 กราฟแยก
        fig = Figure(figsize=(8.3, 11.7))
        for i in range(3):
            ax = fig.add_subplot(3, 1, i + 1)
            ys = [r["ec"][i] for r in rows]
            ax.plot(ts, ys, color=COLORS[i], lw=1.1)
            ax.set_title(f"Container #{i+1}", fontsize=11, fontweight="bold", loc="left")
            ax.set_ylabel("EC (uS/cm)"); ax.grid(True, alpha=0.3)
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
        fig.suptitle("EC vs Time — 3 samples", fontsize=14, fontweight="bold")
        fig.tight_layout(rect=[0, 0, 1, 0.97])
        pdf.savefig(fig)
        # หน้า 2: สถิติ
        fig2 = Figure(figsize=(8.3, 11.7)); ax2 = fig2.add_subplot(111); ax2.axis("off")
        lines = ["Water Quality Report - EC (3 samples)", "",
                 f"Generated : {datetime.now():%Y-%m-%d %H:%M:%S}"]
        if rows:
            lines.append(f"Span      : {ts[0]:%Y-%m-%d %H:%M} - {ts[-1]:%Y-%m-%d %H:%M}")
        lines += [f"Points    : {len(rows)}", "",
                  f"{'Sample':10s}{'n':>7s}{'Mean':>10s}{'Min':>10s}{'Max':>10s}{'SD':>9s}"]
        for i in range(3):
            s = _stats(rows, i)
            if s:
                lines.append(f"{'#'+str(i+1):10s}{s['n']:>7d}{s['mean']:>10.1f}"
                             f"{s['min']:>10.1f}{s['max']:>10.1f}{s['sd']:>9.2f}")
        fig2.text(0.1, 0.9, "\n".join(lines), va="top", family="monospace", fontsize=11)
        pdf.savefig(fig2)


# ============================================================================
#  ค่าตั้งของ UI และไฟล์สถานะที่ logger เขียนไว้  (อ่านอย่างเดียว)
# ============================================================================
def load_rec_status():
    """สถานะบันทึกสดจาก logger_3ec.py — คืน list ยาว MAX_SENSORS เสมอ"""
    active = []
    try:
        with open(REC_STATUS_FILE, encoding="utf-8") as fh:
            active = json.load(fh).get("active") or []
    except Exception:
        active = []
    return [bool(active[i]) if i < len(active) else False
            for i in range(MAX_SENSORS)]


def load_cal_status():
    """สถานะคาลิเบรตสด — ไฟล์นี้จะมีเมื่อทำ Phase 3b แล้ว
       ตอนนี้ยังไม่มีก็ไม่เป็นไร ฟังก์ชันคืน None แล้วแถบก็ไม่ขึ้น"""
    try:
        with open(CAL_STATUS_FILE, encoding="utf-8") as fh:
            d = json.load(fh)
        return d if isinstance(d, dict) else None
    except Exception:
        return None


# ============================================================================
#  widget พื้นฐานตามธีม
# ============================================================================
def set_text(widget, txt):
    """เขียน label เฉพาะตอนข้อความเปลี่ยนจริง

    ยืมวิธีมาจาก ui_set_text() ฝั่งจอ — ของเดิมเขียนทับทุก label ทุก 2 วินาที
    แม้ค่าไม่เปลี่ยน ซึ่งทำให้ข้อความกะพริบเบา ๆ และสิ้นเปลืองเปล่า ๆ
    """
    if getattr(widget, "_txt", None) != txt:
        widget.config(text=txt)
        widget._txt = txt


def set_fg(widget, colour):
    if getattr(widget, "_fg", None) != colour:
        widget.config(fg=colour)
        widget._fg = colour


def keep_bg(widget):
    """ทำเครื่องหมายว่า widget นี้มีสีพื้นของตัวเอง ห้าม set_bg_deep ทับ"""
    widget._keep_bg = True
    return widget


def set_bg_deep(widget, colour):
    """เปลี่ยนสีพื้นของ widget และลูกทุกตัว (ใช้ตอนการ์ดสลับเป็นโหมดปิด)

    ข้ามตัวที่ทำเครื่องหมาย keep_bg() ไว้ — เส้นคั่นกับปุ่มมีสีของตัวเองอยู่แล้ว
    ถ้าทับ เส้นคั่นจะหายไปกลืนกับพื้นการ์ด และปุ่มจะเสียสไตล์
    """
    if getattr(widget, "_keep_bg", False):
        return
    try:
        widget.config(bg=colour)
    except tk.TclError:
        pass
    for child in widget.winfo_children():
        set_bg_deep(child, colour)


def card_frame(parent, bg=None):
    """กล่องพื้นขาวขอบบาง 1 px ไม่มีเงา — เทียบเท่า ui_card() ฝั่งจอ

    ⚠️ Tk ไม่รองรับมุมโค้ง ที่ทำได้คือขอบเหลี่ยม 1 px
       สิ่งที่ทำให้อ่านเป็น 'เครื่องมือห้องแล็บ' คือพื้นสว่าง + ขอบบาง +
       ไม่มีเงา + สีมีความหมาย ไม่ใช่มุมโค้ง
    """
    return tk.Frame(parent, bg=bg or T.SURFACE, bd=0,
                    highlightthickness=T.BORDER_W,
                    highlightbackground=T.BORDER, highlightcolor=T.BORDER)


class Dot(tk.Canvas):
    """วงกลมทึบเล็ก ๆ ใช้บอกสถานะคู่กับข้อความเสมอ — ห้ามสื่อด้วยสีอย่างเดียว"""

    def __init__(self, parent, colour=T.IDLE, size=10, bg=T.SURFACE):
        tk.Canvas.__init__(self, parent, width=size, height=size, bg=bg,
                           highlightthickness=0, bd=0)
        self._oval = self.create_oval(1, 1, size - 1, size - 1,
                                      fill=colour, outline="")
        self._colour = colour

    def set(self, colour):
        if colour != self._colour:
            self.itemconfig(self._oval, fill=colour)
            self._colour = colour


_BTN = {
    "primary":   dict(bg=T.ACCENT,      fg=T.ON_ACCENT,  ab=T.ACCENT_DEEP,
                      af=T.ON_ACCENT,   bd=T.ACCENT,      bold=True),
    "secondary": dict(bg=T.SURFACE,     fg=T.TEXT,    ab=T.SURFACE_ALT,
                      af=T.TEXT,        bd=T.BORDER,      bold=False),
    "quiet":     dict(bg=T.SURFACE_ALT, fg=T.TEXT,    ab=T.BORDER,
                      af=T.TEXT,        bd=T.SURFACE_ALT, bold=False),
    "danger":    dict(bg=T.SURFACE,     fg=T.ERROR,   ab=T.ERROR_SOFT,
                      af=T.ERROR,       bd=T.ERROR,       bold=False),
}


def button(parent, text, kind="secondary", command=None, pady=6):
    s = _BTN[kind]
    return tk.Button(parent, text=text, command=command, bd=0, relief="flat",
                     bg=s["bg"], fg=s["fg"],
                     activebackground=s["ab"], activeforeground=s["af"],
                     disabledforeground=T.IDLE,
                     font=T.f(T.FONT_BODY, "bold" if s["bold"] else "normal"),
                     padx=14, pady=pady, cursor="hand2",
                     highlightthickness=1, highlightbackground=s["bd"],
                     highlightcolor=s["bd"])


class Segmented(tk.Frame):
    """แถบตัวเลือกติดกัน — แทนปุ่มเดี่ยว ๆ ของเดิม

    พื้นของเฟรมเป็นสีขอบ แล้วเว้น 1 px ระหว่างปุ่ม เส้นคั่นจึงเกิดเองโดยไม่ต้อง
    วาดเพิ่ม (Tk ไม่มี separator ในแนวนี้ให้ใช้)
    """

    def __init__(self, parent, options, command):
        tk.Frame.__init__(self, parent, bg=T.BORDER, bd=0,
                          highlightthickness=T.BORDER_W,
                          highlightbackground=T.BORDER)
        self._cmd = command
        self._buttons = []
        for i, (label, value) in enumerate(options):
            b = tk.Button(self, text=label, bd=0, relief="flat",
                          font=T.f(T.FONT_LABEL + 1), padx=13, pady=5,
                          cursor="hand2", highlightthickness=0,
                          command=lambda v=value: self._cmd(v))
            b.pack(side="left", padx=(0 if i == 0 else 1, 0))
            self._buttons.append((b, value))
        self.value = object()

    def set_enabled(self, value, on):
        for b, v in self._buttons:
            if v == value:
                b.config(state=("normal" if on else "disabled"),
                         disabledforeground=T.IDLE,
                         cursor="hand2" if on else "")

    def set(self, value):
        if value == self.value:
            return
        self.value = value
        for b, v in self._buttons:
            on = (v == value)
            b.config(bg=T.ACCENT if on else T.SURFACE,
                     fg=T.ON_ACCENT if on else T.TEXT,
                     activebackground=T.ACCENT_DEEP if on else T.SURFACE_ALT,
                     activeforeground=T.ON_ACCENT if on else T.TEXT,
                     font=T.f(T.FONT_LABEL + 1, "bold" if on else "normal"))


# ============================================================================
#  การ์ดของ sensor หนึ่งตัว
# ============================================================================
class SensorCard(object):
    """โครงเดียวกับ sensor_card.c ฝั่งจอ: ชื่อ / จุด+สถานะ / ค่า / หน่วย+อุณหภูมิ
       / เส้นคั่น / บรรทัดความสดหรือเหตุผลที่อ่านไม่ได้"""

    def __init__(self, parent, index, name, value_px, wide=False, compact=False):
        self.index = index
        self.wide = wide
        self._bg = T.SURFACE
        self.frame = card_frame(parent)
        pad = T.SP_2
        # เรียงสองแถวเมื่อไร ความสูงของการ์ดคูณสอง ต้องกระชับลง
        # ไม่งั้นแผงกราฟจะไม่เหลือพื้นที่เลย
        vpad = T.SP_1 if compact else T.SP_1 + 4

        body = tk.Frame(self.frame, bg=T.SURFACE)
        body.pack(fill="both", expand=True, padx=pad, pady=(vpad, 0))

        left = tk.Frame(body, bg=T.SURFACE)
        left.pack(side="left", fill="both", expand=True)

        self.name = tk.Label(left, text=name, bg=T.SURFACE, fg=T.TEXT_DIM,
                             font=T.f(T.FONT_LABEL, "bold"), anchor="w")
        self.name.pack(fill="x")

        strow = tk.Frame(left, bg=T.SURFACE)
        strow.pack(fill="x", pady=(6, 0))
        self.dot = Dot(strow, T.IDLE, 10, T.SURFACE)
        self.dot.pack(side="left", pady=4)
        self.state = tk.Label(strow, text=T.OFFLINE, bg=T.SURFACE, fg=T.IDLE,
                              font=T.f(T.FONT_STATE, "bold"), anchor="w")
        self.state.pack(side="left", padx=(T.SP_1, 0))

        right = tk.Frame(body, bg=T.SURFACE) if wide else left
        if wide:
            right.pack(side="right", anchor="se")

        anchor = "e" if wide else "w"
        self.value = tk.Label(right, text=T.NO_VALUE, bg=T.SURFACE, fg=T.TEXT,
                              font=T.fm(value_px, "bold"), anchor=anchor)
        self.value.pack(fill="x", pady=(0 if wide else (2 if compact else 6), 0))

        meta = tk.Frame(right, bg=T.SURFACE)
        meta.pack(fill="x", anchor=anchor)
        side = "right" if wide else "left"
        self.temp = tk.Label(meta, text="", bg=T.SURFACE, fg=T.TEXT_DIM,
                             font=T.f(T.FONT_LABEL))
        self.unit = tk.Label(meta, text="uS/cm", bg=T.SURFACE, fg=T.TEXT_DIM,
                             font=T.f(T.FONT_LABEL))
        if wide:
            self.temp.pack(side="right", padx=(T.SP_2, 0))
            self.unit.pack(side="right")
        else:
            self.unit.pack(side="left")
            self.temp.pack(side="left", padx=(T.SP_2, 0))

        keep_bg(tk.Frame(self.frame, bg=T.BORDER, height=1)).pack(
            fill="x", padx=pad, pady=(4 if compact else T.SP_1, 0))
        self.foot = tk.Label(self.frame, text="No data yet", bg=T.SURFACE,
                             fg=T.TEXT_DIM, font=T.f(T.FONT_LABEL),
                             anchor="w", justify="left")
        self.foot.pack(fill="x", padx=pad, pady=(4, vpad))

    # ------------------------------------------------------------------
    def update(self, state, ec, temp, age_s, fails=0, decimals=1):
        st = T.sensor_card_style(state)

        if st["bg"] != self._bg:
            set_bg_deep(self.frame, st["bg"])
            self._bg = st["bg"]
            # set_bg_deep ทับสีพื้นของ canvas จุดสถานะด้วย ตั้งกลับให้ตรงกัน
            self.dot.config(bg=st["bg"])

        set_text(self.state, state)
        set_fg(self.state, st["state_fg"])
        self.dot.set(st["state_fg"])

        txt = T.format_ec(ec, decimals) if st["show_value"] else T.NO_VALUE
        set_text(self.value, txt)
        set_fg(self.value, st["value_fg"])

        set_text(self.temp, T.format_temp(temp) if st["show_value"] else "")

        hint = T.state_hint(state, fails)
        set_text(self.foot, hint or T.format_freshness(age_s))
        set_fg(self.foot, T.ERROR if state == T.SENSOR_FAULT else T.TEXT_DIM)

    def destroy(self):
        self.frame.destroy()


# ============================================================================
#  แผงกราฟ
# ============================================================================
class LabToolbar(NavigationToolbar2Tk):
    """แถบเครื่องมือกราฟที่เหลือเฉพาะปุ่มที่ใช้จริงในงานแล็บ

    ตัด Back / Forward / Subplots ออก ด้วยเหตุผลสองข้อ:
      1. ผู้ใช้ไม่เคยต้องใช้ประวัติการซูมย้อนหลังบนหน้าจอเฝ้าดูค่าสด
         มีปุ่ม Home กับ "Return to live" ก็พอแล้ว
      2. Tk วาดไอคอนของปุ่มที่ถูก disable ด้วยลาย stipple โปร่ง ๆ ซึ่งบนพื้นขาว
         ของธีมนี้กลายเป็นลายตารางที่ดูเหมือนภาพเสีย — สองปุ่มนั้น disable
         อยู่แทบตลอดเวลาเพราะยังไม่มีประวัติการซูม
    """

    toolitems = [t for t in NavigationToolbar2Tk.toolitems
                 if t[0] in ("Home", "Pan", "Zoom", "Save")]



class ChartPanel(object):
    """สร้าง Line2D ครั้งเดียวแล้วเปลี่ยนแค่ข้อมูล

    ⚠️ ของเดิมเรียก ax.clear() ทั้งสามแกนแล้ว canvas.draw() ทุกรอบ ซึ่ง
       (1) ทำลาย artist ทิ้งแล้วสร้างใหม่ — กราฟกะพริบทั้งแผ่น
       (2) ล้าง xlim/ylim ที่ผู้ใช้เพิ่งซูมไว้ทุก 10 วินาที
       ตรงนี้จึงใช้ set_data() + draw_idle() แทน และ autoscale เฉพาะตอนที่
       ผู้ใช้ยังตามค่าสดอยู่
    """

    def __init__(self, parent, names, decimals=1):
        self.names = names
        self.decimals = decimals
        self.frame = card_frame(parent)

        head = tk.Frame(self.frame, bg=T.SURFACE)
        head.pack(fill="x", padx=T.SP_2, pady=(10, 0))
        self.cap = tk.Label(head, text="EC vs TIME", bg=T.SURFACE,
                            fg=T.TEXT_DIM, font=T.f(T.FONT_LABEL, "bold"))
        self.cap.pack(side="left")
        self.pts = tk.Label(head, text="", bg=T.SURFACE, fg=T.TEXT_DIM,
                            font=T.f(T.FONT_LABEL))
        self.pts.pack(side="right")

        self.fig = Figure(figsize=(9, 3.4), dpi=100)
        T.mpl_style_figure(self.fig)
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.frame)
        # ขอความสูงขั้นต่ำไว้ ไม่งั้นเมื่อพื้นที่ไม่พอ Tk จะบีบ canvas จน
        # แถบเครื่องมือ zoom/pan หลุดออกนอกหน้าต่างไปเลย
        self.canvas.get_tk_widget().config(bg=T.SURFACE, highlightthickness=0,
                                           height=T.CHART_MIN_H)
        # ⚠️ ต้อง pack toolbar ก่อน canvas
        #    canvas ใช้ expand=True ถ้า pack ก่อน มันจะกินพื้นที่จนหมด
        #    แล้วแถบ zoom/pan โดนตัดหลุดออกนอกการ์ดไปเลยเมื่อหน้าต่างเตี้ย
        self.toolbar = LabToolbar(self.canvas, self.frame,
                                  pack_toolbar=False)
        self.toolbar.update()
        self._theme_toolbar()
        self.toolbar.pack(side="bottom", fill="x", padx=T.SP_1, pady=(0, 2))

        self.canvas.get_tk_widget().pack(fill="both", expand=True,
                                         padx=T.SP_1, pady=(4, 0))

        self._sig = None
        self.axes = []
        self.lines = {}
        self.ax_of = {}
        self.notes = {}

    # ------------------------------------------------------------------
    def _theme_toolbar(self):
        """NavigationToolbar2Tk เป็น tk.Frame ธรรมดา ตั้งสีลูก ๆ ได้ตรง ๆ
           ปุ่มบางตัวเป็น Checkbutton จึงต้องลอง config ทีละ option"""
        for w in [self.toolbar] + list(self.toolbar.winfo_children()):
            for opts in ({"bg": T.SURFACE},
                         {"fg": T.TEXT_DIM},
                         {"activebackground": T.SURFACE_ALT},
                         {"selectcolor": T.SURFACE},
                         {"font": T.f(T.FONT_LABEL)},
                         {"relief": "flat", "bd": 0, "highlightthickness": 0}):
                try:
                    w.config(**opts)
                except tk.TclError:
                    pass

        # ⚠️ matplotlib สร้างไอคอนของปุ่มโดยผสมสีกับพื้นหลัง "ตอนที่สร้าง"
        #    เราเพิ่งเปลี่ยนพื้นหลังทีหลัง ไอคอนของปุ่มที่ถูก disable
        #    (back/forward) จึงกลายเป็นลายตาราง ต้องสั่งสร้างไอคอนใหม่
        try:
            for b in self.toolbar._buttons.values():
                self.toolbar._set_image_for_button(b)
        except Exception:
            pass

    def _following_paused(self):
        return bool(str(getattr(self.toolbar, "mode", "") or ""))

    def bind_release(self, cb):
        self.canvas.mpl_connect("button_release_event",
                                lambda e: cb(self._following_paused()))

    # ------------------------------------------------------------------
    def rebuild(self, active, mode):
        """สร้างแกนใหม่เฉพาะตอน mask หรือโหมดกราฟเปลี่ยน ไม่ใช่ทุกรอบ refresh"""
        sig = (tuple(active), mode)
        if sig == self._sig:
            return
        self._sig = sig
        self.fig.clear()
        self.axes = []
        self.lines = {}
        self.ax_of = {}
        self.notes = {}

        forced = (mode == "overlay" and len(active) > MAX_SPLIT_PANELS)
        set_text(self.cap, "EC vs TIME — {}   ·   uS/cm{}".format(
            mode.upper(),
            "   ·   split needs more height than {} panels allow".format(
                len(active)) if forced else ""))

        if not active:
            ax = self.fig.add_subplot(111)
            T.mpl_style_axes(ax)
            ax.set_xticks([])
            ax.set_yticks([])
            ax.grid(False)
            ax.text(0.5, 0.5, "No sensor enabled", ha="center", va="center",
                    color=T.IDLE, transform=ax.transAxes)
            self.axes = [ax]
            self.canvas.draw_idle()
            return

        if mode == "overlay":
            ax = self.fig.add_subplot(111)
            T.mpl_style_axes(ax)
            ax.xaxis_date()
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
            ax.set_ylabel("EC  uS/cm", color=T.TEXT_DIM, fontsize=9)
            for i in active:
                line, = ax.plot([], [], color=T.SERIES[i], lw=1.5,
                                label=self.names[i])
                self.lines[i] = line
            # วาง legend "นอกกรอบ" ด้านบน — ถ้าวางในกรอบมันจะทับเส้นข้อมูล
            # ซึ่งเป็นสิ่งเดียวที่ผู้ใช้ต้องการดู
            leg = ax.legend(loc="lower left", bbox_to_anchor=(0.0, 1.005),
                            ncol=len(active), frameon=False, fontsize=9,
                            handlelength=1.8, columnspacing=1.6,
                            borderaxespad=0.0)
            for txt in leg.get_texts():
                txt.set_color(T.TEXT_DIM)
            self.axes = [ax]
        else:
            n = len(active)
            base = None
            for k, i in enumerate(active):
                ax = self.fig.add_subplot(n, 1, k + 1, sharex=base)
                base = base or ax
                last = (k == n - 1)
                T.mpl_style_axes(ax, labelbottom=last)
                ax.xaxis_date()
                ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
                # แผงเตี้ยลงเมื่อมีหลายตัว ถ้าปล่อยให้ matplotlib เลือกจำนวนขีดเอง
                # ตัวเลขบนแกน y จะซ้อนทับกันจนอ่านไม่ออกในโหมด 4 เซนเซอร์
                ax.yaxis.set_major_locator(MaxNLocator(nbins=3))
                ax.set_ylabel("#{}".format(i + 1), color=T.SERIES[i],
                              fontsize=11, fontweight="bold", labelpad=8)
                line, = ax.plot([], [], color=T.SERIES[i], lw=1.5)
                self.lines[i] = line
                self.ax_of[i] = ax
                self.notes[i] = ax.text(0.5, 0.5, "", ha="center", va="center",
                                        color=T.IDLE, fontsize=10,
                                        transform=ax.transAxes, visible=False)
                self.axes.append(ax)

        # overlay ต้องเว้นที่ด้านบนให้ legend ที่อยู่นอกกรอบ
        self.fig.subplots_adjust(left=0.062, right=0.988,
                                 top=0.82 if mode == "overlay" else 0.97,
                                 bottom=0.15, hspace=0.16)
        self.canvas.draw_idle()

    # ------------------------------------------------------------------
    def update(self, rows, follow, downsampled):
        if not self.lines:
            self.canvas.draw_idle()
            return

        xs = mdates.date2num([r["t"] for r in rows]) if rows else []
        n_slots = len(rows[-1]["ec"]) if rows else 0

        for i, line in self.lines.items():
            has_data = False
            if not rows or i >= n_slots:
                line.set_data([], [])
            else:
                ys = [(r["ec"][i] if r["ec"][i] is not None else float("nan"))
                      for r in rows]
                line.set_data(xs, ys)
                has_data = any(v == v for v in ys)   # v == v เป็น False เมื่อเป็น nan

            # แผงของเซนเซอร์ที่ไม่มีข้อมูลต้องบอกตรง ๆ ว่าไม่มีข้อมูล
            # ไม่ใช่ปล่อยให้ขึ้นแกน -0.05..0.05 ซึ่งดูเหมือนค่าจริงที่เป็นศูนย์
            ax = self.ax_of.get(i)
            note = self.notes.get(i)
            if ax is not None and note is not None:
                note.set_visible(not has_data)
                if not has_data:
                    note.set_text("no data")
                    ax.set_ylim(0.0, 1.0)
                    ax.tick_params(labelleft=False)
                else:
                    ax.tick_params(labelleft=True)

        if rows and follow:
            for ax in self.axes:
                skip = any(ax is self.ax_of.get(i) and self.notes[i].get_visible()
                           for i in self.notes)
                if skip:
                    continue
                ax.relim()
                ax.autoscale_view()

        set_text(self.pts, "{:,} points ({})".format(
            len(rows), "downsampled" if downsampled else "raw"))
        self.canvas.draw_idle()


# ============================================================================
#  แผงเหตุการณ์
# ============================================================================
class EventLog(object):
    """บันทึกเหตุการณ์แบบเครื่องมือวัด ไม่ใช่ terminal

    จำนวนบรรทัดคงที่ เพื่อไม่ให้ความสูงของแผงขยับแล้วดัน layout ทั้งหน้า
    """

    ROWS = 3

    def __init__(self, parent):
        self.frame = card_frame(parent)
        self.frame.pack_propagate(False)
        tk.Label(self.frame, text="EVENT LOG", bg=T.SURFACE, fg=T.TEXT_DIM,
                 font=T.f(T.FONT_LABEL, "bold"), anchor="w").pack(
            fill="x", padx=T.SP_2, pady=(T.SP_1, 4))
        self.rows = []
        self.holders = []
        self.visible = self.ROWS
        for k in range(self.ROWS):
            sep = keep_bg(tk.Frame(self.frame, bg=T.BORDER, height=1))
            sep.pack(fill="x")
            r = tk.Frame(self.frame, bg=T.SURFACE)
            r.pack(fill="x", padx=T.SP_2, pady=3)
            self.holders.append((sep, r))
            t = tk.Label(r, text="", bg=T.SURFACE, fg=T.TEXT_DIM,
                         font=T.fm(T.FONT_LABEL), width=9, anchor="w")
            t.pack(side="left")
            d = Dot(r, T.IDLE, 8, T.SURFACE)
            d.pack(side="left", padx=(4, T.SP_1))
            m = tk.Label(r, text="", bg=T.SURFACE, fg=T.TEXT,
                         font=T.f(T.FONT_LABEL + 1), anchor="w")
            m.pack(side="left", fill="x", expand=True)
            self.rows.append((t, d, m))
        self.set_rows(self.ROWS)

    def set_rows(self, n):
        """ตัดบรรทัดเมื่อจอเตี้ย — ต้องยอมตัด event log ก่อนตัดกราฟ

        n = 0 คือซ่อนแผงทั้งอัน (จอเตี้ยมากจริง ๆ)
        """
        n = max(0, min(self.ROWS, int(n)))
        self.visible = n
        for k, (sep, row) in enumerate(self.holders):
            if k < n:
                sep.pack(fill="x", before=row)
                row.pack(fill="x", padx=T.SP_2, pady=3)
            else:
                sep.pack_forget()
                row.pack_forget()
        self.frame.config(height=max(1, T.EVENT_HEAD_H + n * T.EVENT_ROW_H))

    def render(self, events):
        for k, (t, d, m) in enumerate(self.rows):
            if k >= self.visible:
                continue
            if k < len(events):
                when, colour, kind, text = events[k]
                set_text(t, when.strftime("%H:%M:%S"))
                d.set(colour)
                set_text(m, "{} · {}".format(kind, text) if text else kind)
                set_fg(m, T.TEXT)
            else:
                set_text(t, "")
                d.set(T.SURFACE)
                set_text(m, "")


# ============================================================================
#  กล่องตั้งค่าการแสดงผล
# ----------------------------------------------------------------------------
#  ⚠️ ค่าที่ตั้งในนี้เป็น "การแสดงผลของ PC" ล้วน ๆ ไม่ได้ไปสั่งอะไรกับบอร์ด
#     CONTROL ยังคง poll ครบทุกตัวเหมือนเดิม และ CSV ยังเก็บครบทุกช่อง
# ============================================================================
class DisplaySetup(tk.Toplevel):
    def __init__(self, parent, app):
        tk.Toplevel.__init__(self, parent)
        self.app = app
        self.title("Display setup")
        self.configure(bg=T.BG)
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        # วางกลางหน้าต่างหลัก ไม่ใช่มุมซ้ายบนของจอ
        self.update_idletasks()
        try:
            x = parent.winfo_rootx() + (parent.winfo_width() - self.winfo_width()) // 2
            y = parent.winfo_rooty() + 90
            self.geometry("+{}+{}".format(max(x, 0), max(y, 0)))
        except Exception:
            pass

        wrap = tk.Frame(self, bg=T.BG)
        wrap.pack(fill="both", expand=True, padx=T.SP_3, pady=T.SP_3)

        tk.Label(wrap, text="SENSORS SHOWN ON THIS DASHBOARD", bg=T.BG,
                 fg=T.TEXT_DIM, font=T.f(T.FONT_LABEL, "bold"),
                 anchor="w").pack(fill="x")
        tk.Label(wrap, text="ตัวที่ปิดจะขึ้น DISABLED และไม่ถูกวาดบนกราฟ\n"
                            "การบันทึกข้อมูลดิบและการทำงานของบอร์ดไม่เปลี่ยน",
                 bg=T.BG, fg=T.TEXT_DIM, font=T.f(T.FONT_LABEL),
                 anchor="w", justify="left").pack(fill="x", pady=(2, T.SP_1))

        box = card_frame(wrap)
        box.pack(fill="x", pady=(0, T.SP_2))
        self.vars = []
        for i in range(MAX_SENSORS):
            v = tk.IntVar(value=1 if (app.mask >> i) & 1 else 0)
            self.vars.append(v)
            cb = tk.Checkbutton(box, text="  " + app.names[i], variable=v,
                                bg=T.SURFACE, fg=T.TEXT, selectcolor=T.SURFACE,
                                activebackground=T.SURFACE,
                                activeforeground=T.TEXT, bd=0,
                                highlightthickness=0, anchor="w",
                                font=T.f(T.FONT_BODY), cursor="hand2")
            cb.pack(fill="x", padx=T.SP_2, pady=(T.SP_1 if i == 0 else 2,
                                                 T.SP_1 if i == MAX_SENSORS - 1 else 2))

        tk.Label(wrap, text="TREND LAYOUT", bg=T.BG, fg=T.TEXT_DIM,
                 font=T.f(T.FONT_LABEL, "bold"), anchor="w").pack(fill="x")
        self.mode = tk.StringVar(value=app.chart_mode)
        mrow = tk.Frame(wrap, bg=T.BG)
        mrow.pack(fill="x", pady=(T.SP_1, T.SP_3))
        for label, val in (("Split", "split"), ("Overlay", "overlay")):
            tk.Radiobutton(mrow, text="  " + label, variable=self.mode,
                           value=val, bg=T.BG, fg=T.TEXT, selectcolor=T.SURFACE,
                           activebackground=T.BG, activeforeground=T.TEXT,
                           bd=0, highlightthickness=0, font=T.f(T.FONT_BODY),
                           cursor="hand2").pack(side="left", padx=(0, T.SP_3))

        act = tk.Frame(wrap, bg=T.BG)
        act.pack(fill="x")
        button(act, "Cancel", "secondary", self.destroy).pack(side="right")
        button(act, "Apply", "primary", self._apply).pack(
            side="right", padx=(0, T.SP_1))

    def _apply(self):
        mask = 0
        for i, v in enumerate(self.vars):
            if v.get():
                mask |= (1 << i)
        if not mask:
            messagebox.showwarning("Display setup",
                                   "ต้องเปิดอย่างน้อยหนึ่งตัว", parent=self)
            return
        self.app.apply_display_config(mask, self.mode.get())
        self.destroy()


# ============================================================================
#  หน้าต่างหลัก
# ============================================================================
class App(object):
    RANGES = [("10 min", 10), ("1 hr", 60), ("6 hr", 360),
              ("24 hr", 1440), ("All", None)]

    def __init__(self, root):
        self.root = root
        self.cfg = T.load_ui_config()
        self.mask = self.cfg["active_mask"]
        self.names = self.cfg["sensor_names"]
        self.decimals = int(self.cfg.get("ec_decimals", 1))
        self.chart_mode = self.cfg["chart_mode"]

        self.minutes = 60
        self.mock = False
        self.follow_live = True
        self.cards = []
        self.events = []
        self.sessions = load_sessions()
        self._rec_prev = None
        self._board_offline = None
        self._layout_sig = None

        # ⚠️ ต้องคำนวณสเกลก่อนสร้าง widget ตัวแรก เพราะ T.f() อ่านค่าจากตรงนี้
        T.init_scaling(root)
        T.apply_ttk_theme(root)

        root.title("EC Measurement Station  ·  {}".format(APP_VERSION))
        root.configure(bg=T.BG)

        # ขนาดหน้าต่าง: คำนวณจากจอจริง ไม่ใช่เลขตายตัว
        # เครื่องที่จอเตี้ยกว่าเลขที่ฮาร์ดโค้ดไว้จะโดนบีบทันทีโดยไม่มีใครรู้ตัว
        win = self.cfg.get("window")
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        if isinstance(win, (list, tuple)) and len(win) == 2:
            w, h = int(win[0]), int(win[1])
        else:
            w, h = min(1280, sw - 80), min(980, sh - 90)
        w, h = max(w, 940), max(h, 620)
        root.geometry("{}x{}+{}+{}".format(w, h, max((sw - w) // 2, 0), 24))
        root.minsize(940, 620)
        self._density = None
        self._dens_job = None

        self.sess_var = tk.StringVar()

        self._build_header()
        self._build_banner()
        self._build_cards_area()
        self._build_controls()
        self._build_recbar()
        self._build_events()
        self._build_chart()

        self._log("SESSION", T.ACCENT, "Dashboard started")
        self._apply_density(h)
        self.rebuild_layout()
        self.root.bind("<Configure>", self._on_configure)
        self.refresh_now()
        self.refresh_charts()
        self._ticks = 0
        self._schedule()

    # ---------------------------------------------------------------- utils
    def active(self):
        return T.mask_to_list(self.mask, MAX_SENSORS)

    def _log(self, kind, colour, text=""):
        self.events.insert(0, (datetime.now(), colour, kind, text))
        del self.events[40:]
        if hasattr(self, "evlog"):
            self.evlog.render(self.events)

    # ---------------------------------------------------------------- build
    def _build_header(self):
        bar = tk.Frame(self.root, bg=T.SURFACE)
        bar.pack(fill="x")
        self.hdr_sep = keep_bg(tk.Frame(self.root, bg=T.BORDER, height=1))
        self.hdr_sep.pack(fill="x")

        inner = tk.Frame(bar, bg=T.SURFACE)
        inner.pack(fill="x", padx=T.SP_3, pady=T.SP_2)

        # ⚠️ ต้อง pack ฝั่งขวาก่อนฝั่งซ้ายที่ยืดได้
        #    ไม่งั้นฝั่งซ้ายจะกินพื้นที่จนป้ายเวลาโดนตัดหายไปครึ่งคำ
        button(inner, "Display setup", "secondary",
               lambda: DisplaySetup(self.root, self)).pack(side="right")
        self.fresh = tk.Label(inner, text="", bg=T.SURFACE, fg=T.TEXT_DIM,
                              font=T.fm(T.FONT_LABEL), anchor="e")
        self.fresh.pack(side="right", padx=T.SP_2)

        # ชื่อกับสรุปแยกสองบรรทัด เหมือน screen_overview ของ P4
        # สรุปยาวได้โดยไม่ไปเบียดอะไร
        left = tk.Frame(inner, bg=T.SURFACE)
        left.pack(side="left", fill="x", expand=True)

        trow = tk.Frame(left, bg=T.SURFACE)
        trow.pack(fill="x")
        self.hdot = Dot(trow, T.IDLE, 12, T.SURFACE)
        self.hdot.pack(side="left", pady=4)
        tk.Label(trow, text="EC MEASUREMENT STATION", bg=T.SURFACE, fg=T.TEXT,
                 font=T.f(T.FONT_TITLE, "bold")).pack(side="left",
                                                      padx=(T.SP_1 + 2, 0))

        self.summary = tk.Label(left, text="", bg=T.SURFACE, fg=T.TEXT_DIM,
                                font=T.f(T.FONT_LABEL + 1), anchor="w")
        self.summary.pack(fill="x", padx=(22, 0))

    def _build_banner(self):
        self.banner = tk.Frame(self.root, bg=T.WARN_SOFT, bd=0,
                               highlightthickness=1,
                               highlightbackground=T.WARN_LINE)
        inner = tk.Frame(self.banner, bg=T.WARN_SOFT)
        inner.pack(fill="x", padx=T.SP_2, pady=T.SP_1)
        self.bdot = Dot(inner, T.WARN, 10, T.WARN_SOFT)
        self.bdot.pack(side="left", pady=3)
        self.btitle = tk.Label(inner, text="", bg=T.WARN_SOFT, fg=T.TEXT,
                               font=T.f(T.FONT_BODY, "bold"))
        self.btitle.pack(side="left", padx=(T.SP_1, T.SP_2))
        self.bdetail = tk.Label(inner, text="", bg=T.WARN_SOFT, fg=T.TEXT_DIM,
                                font=T.f(T.FONT_LABEL + 1))
        self.bdetail.pack(side="left")
        self.bbtn = keep_bg(button(inner, "Return to live", "secondary",
                                   self.return_to_live, pady=3))
        self._banner_inner = inner
        self._banner_kind = None
        self._banner_visible = False

    def _build_cards_area(self):
        self.cards_area = tk.Frame(self.root, bg=T.BG)
        # ⚠️ ความสูงตายตัว ไม่ปล่อยให้โตตามเนื้อหา
        #    ถ้าปล่อย พอพื้นที่ไม่พอ Tk จะไปบีบแผงกราฟแทน ซึ่งกลับหัวกลับหาง
        # ⚠️ การ์ดวางด้วย grid() จึงต้องใช้ grid_propagate ไม่ใช่ pack_propagate
        #    (เรียกผิดตัว = ไม่มีผลอะไรเลย และไม่มี error ให้เห็นด้วย)
        self.cards_area.grid_propagate(False)
        self.cards_area.pack_propagate(False)
        self.cards_area.pack(fill="x", padx=T.SP_3, pady=(T.SP_2, 0))

    def _build_controls(self):
        bar = tk.Frame(self.root, bg=T.BG)
        bar.pack(fill="x", padx=T.SP_3, pady=(T.SP_2, 0))

        tk.Label(bar, text="Range", bg=T.BG, fg=T.TEXT_DIM,
                 font=T.f(T.FONT_LABEL + 1)).pack(side="left", padx=(0, T.SP_1))
        self.seg_range = Segmented(bar, self.RANGES, self.set_range)
        self.seg_range.pack(side="left")
        self.seg_range.set(self.minutes)

        self.seg_mode = Segmented(bar, [("Split", "split"),
                                        ("Overlay", "overlay")],
                                  self.set_chart_mode)
        self.seg_mode.pack(side="left", padx=(T.SP_2, 0))
        self.seg_mode.set(self.chart_mode)

        button(bar, "Export PDF", "primary",
               lambda: self.do_export("pdf")).pack(side="right")
        button(bar, "Export Excel", "secondary",
               lambda: self.do_export("xlsx")).pack(side="right", padx=T.SP_1)
        button(bar, "Export CSV", "secondary",
               lambda: self.do_export("csv")).pack(side="right",
                                                   padx=(T.SP_3, 0))

    def _build_recbar(self):
        outer = card_frame(self.root)
        outer.pack(fill="x", padx=T.SP_3, pady=(T.SP_2, 0))
        bar = tk.Frame(outer, bg=T.SURFACE)
        bar.pack(fill="x", padx=T.SP_2, pady=T.SP_1)

        # ⚠️ pack ฝั่งขวาก่อน ไม่งั้นพอมี 4 เซนเซอร์ จุด REC จะดันจนปุ่ม Mock
        #    โดนตัดหายไปครึ่งปุ่ม
        self.mock_btn = button(bar, "Mock data", "secondary", self.toggle_mock,
                               pady=3)
        self.mock_btn.pack(side="right")

        tk.Label(bar, text="Recording", bg=T.SURFACE, fg=T.TEXT_DIM,
                 font=T.f(T.FONT_LABEL + 1)).pack(side="left", padx=(0, T.SP_2))
        # กล่องของตัวเอง — เพราะจุด REC ถูก pack ทีหลังตอน rebuild_layout()
        # ถ้า pack ลง bar ตรง ๆ มันจะไปโผล่ต่อท้าย widget ที่ pack ไปแล้ว
        self.rec_box = tk.Frame(bar, bg=T.SURFACE)
        self.rec_box.pack(side="left")
        self.rec_dots = []
        self.rec_lbls = []
        for i in range(MAX_SENSORS):
            d = Dot(self.rec_box, T.IDLE, 9, T.SURFACE)
            l = tk.Label(self.rec_box, text="", bg=T.SURFACE, fg=T.IDLE,
                         font=T.f(T.FONT_LABEL + 1, "bold"))
            self.rec_dots.append(d)
            self.rec_lbls.append(l)

        tk.Label(bar, text="Session", bg=T.SURFACE, fg=T.TEXT_DIM,
                 font=T.f(T.FONT_LABEL + 1)).pack(side="left", padx=(T.SP_3, T.SP_1))
        self.sess_combo = ttk.Combobox(bar, textvariable=self.sess_var,
                                       width=28, state="readonly",
                                       style="Lab.TCombobox")
        self.sess_combo.pack(side="left")
        self._reload_sessions()

    def _build_events(self):
        self.evlog = EventLog(self.root)
        self.evlog.frame.pack(side="bottom", fill="x",
                              padx=T.SP_3, pady=(T.SP_2, T.SP_2))

    def _build_chart(self):
        self.chart = ChartPanel(self.root, self.names, self.decimals)
        self.chart.frame.pack(fill="both", expand=True,
                              padx=T.SP_3, pady=(T.SP_2, 0))
        self.chart.bind_release(self._on_chart_release)

    # ---------------------------------------------------------------- layout
    def rebuild_layout(self):
        act = self.active()
        sig = (tuple(act), self.chart_mode, self._density)
        if sig == self._layout_sig:
            return
        self._layout_sig = sig

        for c in self.cards:
            c.destroy()
        self.cards = []

        rows, cols = T.grid_for(len(act))
        for k in range(cols):
            self.cards_area.grid_columnconfigure(k, weight=1, uniform="card")
        for k in range(cols, MAX_SENSORS):
            self.cards_area.grid_columnconfigure(k, weight=0, uniform="")

        for k in range(2):
            self.cards_area.grid_rowconfigure(
                k, weight=1 if k < rows else 0,
                uniform="cardrow" if k < rows else "")

        dens = self._density or "roomy"
        px = T.value_font_size(len(act), dens)
        wide = (len(act) == 1)
        # กระชับเมื่อเรียงสองแถว หรือเมื่อหน้าต่างเตี้ย
        compact = (rows > 1) or dens in ("tight", "min")
        for k, i in enumerate(act):
            card = SensorCard(self.cards_area, i, self.names[i], px,
                              wide=wide, compact=compact)
            card.frame.grid(row=k // cols, column=k % cols, sticky="nsew",
                            padx=(0 if k % cols == 0 else T.SP_1, 0),
                            pady=(0 if k // cols == 0 else T.SP_1, 0))
            self.cards.append(card)

        # ⚠️ ความสูงของแถบการ์ดต้อง "วัดจากของจริง" ไม่ใช่ใส่เลขคาดเดา
        #    ใส่เลขตายตัวแล้วเมื่อไรที่ฟอนต์ใหญ่ขึ้น (จอ DPI สูง / ฟอนต์ระบบ
        #    คนละตัว) บรรทัดล่างของการ์ดจะโดนตัดหายไปเงียบ ๆ
        #    วัดแล้วตรึงไว้ เพื่อไม่ให้มันไปแย่งพื้นที่กราฟตอนเนื้อหาเปลี่ยน
        self.cards_area.update_idletasks()
        nat = max([c.frame.winfo_reqheight() for c in self.cards] or [1])
        nat = max(nat, T.card_height_for(len(act), rows, dens))
        self.cards_area.config(height=rows * nat + (rows - 1) * T.SP_1)

        for i in range(MAX_SENSORS):
            self.rec_dots[i].pack_forget()
            self.rec_lbls[i].pack_forget()
        for i in act:
            self.rec_dots[i].pack(side="left", padx=(0, 4))
            self.rec_lbls[i].pack(side="left", padx=(0, T.SP_2))

        self.seg_mode.set_enabled("split", len(act) <= MAX_SPLIT_PANELS)
        self.chart.rebuild(act, self.effective_mode())

    def effective_mode(self):
        """โหมดกราฟที่ใช้จริง — บังคับ overlay เมื่อแผงเยอะเกินจะอ่านออก"""
        if len(self.active()) > MAX_SPLIT_PANELS:
            return "overlay"
        return self.chart_mode

    def apply_display_config(self, mask, mode):
        changed = (mask != self.mask) or (mode != self.chart_mode)
        self.mask = mask
        self.chart_mode = mode
        self.cfg["active_mask"] = mask
        self.cfg["chart_mode"] = mode
        T.save_ui_config(self.cfg)
        self.seg_mode.set(mode)
        self.rebuild_layout()
        self.refresh_now()
        self.refresh_charts()
        if changed:
            self._log("DISPLAY", T.ACCENT, "Showing {} · trend {}".format(
                ", ".join(self.names[i] for i in self.active()) or "none", mode))

    # ------------------------------------------------------------- density
    def _on_configure(self, ev):
        """ปรับความหนาแน่นเมื่อย่อ/ขยายหน้าต่าง (หน่วงไว้กัน event ถี่)"""
        if ev.widget is not self.root:
            return
        if self._dens_job:
            self.root.after_cancel(self._dens_job)
        self._dens_job = self.root.after(
            200, lambda: self._apply_density(self.root.winfo_height()))

    def _apply_density(self, h):
        self._dens_job = None
        d = T.density_for(h)
        if d == self._density:
            return
        self._density = d
        n = T.EVENT_ROWS_BY_DENSITY[d]
        # ⚠️ ห้ามเช็ค winfo_ismapped() ตรงนี้ ตอนบูตหน้าต่างยังไม่ถูก realize
        #    มันจะคืน 0 ทั้งที่ pack ไปแล้ว ทำให้การซ่อนไม่เกิดขึ้น
        self.evlog.frame.pack_forget()
        if n:
            self.evlog.set_rows(n)
            self.evlog.render(self.events)
            self.evlog.frame.pack(side="bottom", fill="x",
                                  padx=T.SP_3, pady=(T.SP_2, T.SP_2))
        self.rebuild_layout()

    # ---------------------------------------------------------------- range
    def set_range(self, m):
        self.minutes = m
        self.sess_var.set(ALL_DATA)
        self.follow_live = True
        self.seg_range.set(m)
        self.refresh_charts()

    def set_chart_mode(self, mode):
        self.apply_display_config(self.mask, mode)

    def return_to_live(self):
        self.follow_live = True
        self.sess_var.set(ALL_DATA)
        try:
            self.chart.toolbar.home()
        except Exception:
            pass
        self.refresh_charts()

    def _on_chart_release(self, paused):
        if paused and self.follow_live:
            self.follow_live = False
            self.refresh_now()

    # ---------------------------------------------------------------- data
    def _current_rows(self):
        """คืนข้อมูลตามที่เลือก: session (ถ้าเลือก) หรือช่วงเวลา
           — ตรรกะเดิม ไม่เปลี่ยน"""
        sel = self.sess_var.get()
        if sel in getattr(self, "_sess_map", {}):
            st, en = self._sess_map[sel]
            return read_range(st, en)
        return read_recent(self.minutes)

    # ---------------------------------------------------------------- refresh
    def refresh_now(self):
        # ⚠️ หน้าต่างนี้กว้างกว่าที่ต้องใช้คำนวณสถานะ (ต้องการแค่ 6 ค่าล่าสุด)
        #    โดยตั้งใจ — ถ้าแคบแค่ 5 นาทีแบบเดิม พอบอร์ดเงียบเกิน 5 นาที
        #    read_recent() จะคืน list ว่าง แล้วทุกการ์ดเด้งไป OFFLINE ทั้งที่
        #    ความจริงคือ STALE (เคยมีข้อมูล แต่ขาดไป) ซึ่งคนละเรื่องกัน
        #    ไม่มีต้นทุนเพิ่ม เพราะ read_range() อ่านทุกไฟล์ทุกบรรทัดอยู่แล้ว
        rows = read_recent(30)
        now = datetime.now()
        last = rows[-1] if rows else None
        age = (now - last["t"]).total_seconds() if last else None
        n_slots = len(last["ec"]) if last else 0

        win = rows[-T.STABLE_WINDOW_SAMPLES:] if rows else []
        counts = {"total": len(self.active())}
        act = self.active()

        for card in self.cards:
            i = card.index
            if i >= n_slots:
                state, ec, tw, fails = T.OFFLINE, None, None, 0
            else:
                vals = [r["ec"][i] for r in win]
                oks = [r["ok"][i] == "1" for r in win]
                state = T.monitor_state(vals, oks, age, enabled=True)
                fails = T.consecutive_fails(oks)
                ec = last["ec"][i]
                tw = last["tw"][i]
            card.update(state, ec, tw, age, fails, self.decimals)

            key = {T.STEADY: "steady", T.LIVE: "live", T.CHANGING: "live",
                   T.STALE: "stale", T.NO_RESPONSE: "fault",
                   T.SENSOR_FAULT: "fault", T.OFFLINE: "offline"}.get(state)
            if key:
                counts[key] = counts.get(key, 0) + 1

        # นับเฉพาะช่องที่ "มีอยู่จริง" — ไม่นับ slot ที่ยังไม่มีหัววัดเลย
        # ไม่งั้น mask 0b0011 บนข้อมูล 3 ช่อง จะรายงานว่ามี 4 SENSORS
        n_disabled = sum(1 for i in range(max(n_slots, max(act) + 1 if act else 0))
                         if i not in act)
        if n_disabled:
            counts["disabled"] = n_disabled
            counts["total"] = counts["total"] + n_disabled

        set_text(self.summary, T.summary_text(counts))
        set_fg(self.summary, T.ERROR if counts.get("fault") else T.TEXT_DIM)

        board_live = age is not None and age < STALE_BANNER_S
        self.hdot.set(T.OK if board_live else (T.WARN if age is not None
                                               else T.IDLE))
        set_text(self.fresh, T.format_freshness(age))

        is_offline = not board_live
        if self._board_offline is None:
            self._board_offline = is_offline
        elif is_offline != self._board_offline:
            self._board_offline = is_offline
            if is_offline:
                self._log("BOARD", T.ERROR, "No new rows from the sensor board")
            else:
                self._log("BOARD", T.OK, "Data stream recovered")

        self._update_rec()
        self._update_banner(age)

    def _update_rec(self):
        active = load_rec_status()
        act = self.active()
        for i in range(MAX_SENSORS):
            if i not in act:
                continue
            if active[i]:
                self.rec_dots[i].set(T.REC)
                set_text(self.rec_lbls[i], "#{} REC".format(i + 1))
                set_fg(self.rec_lbls[i], T.REC)
            else:
                self.rec_dots[i].set(T.IDLE)
                set_text(self.rec_lbls[i], "#{} idle".format(i + 1))
                set_fg(self.rec_lbls[i], T.IDLE)

        if self._rec_prev is not None:
            for i in range(MAX_SENSORS):
                if active[i] != self._rec_prev[i]:
                    self._log("SESSION",
                              T.ACCENT if active[i] else T.TEXT_DIM,
                              "{} {}".format(self.names[i],
                                             "started" if active[i] else "stopped"))
            if active != self._rec_prev:
                self._reload_sessions()
        self._rec_prev = active

    def _update_banner(self, age):
        cal = load_cal_status()
        sel = self.sess_var.get()
        viewing_history = (sel in getattr(self, "_sess_map", {})) or not self.follow_live

        if age is None or age > STALE_BANNER_S:
            self._show_banner("error", "SENSOR BOARD OFFLINE",
                              "ไม่มีข้อมูลใหม่ — ตรวจว่า logger_3ec.py "
                              "ยังทำงานและสาย USB ยังเสียบอยู่", False)
        elif cal and cal.get("busy"):
            n = cal.get("sensor", "?")
            std = cal.get("standard", "?")
            self._show_banner("info",
                              "CALIBRATING SENSOR {:02d} @ {} uS/cm".format(
                                  int(n) if str(n).isdigit() else 0, std),
                              cal.get("phase", "in progress"), False)
        elif self.mock:
            self._show_banner("warn", "MOCK DATA",
                              "กำลังแสดงข้อมูลจำลอง ไม่ใช่ค่าจากหัววัดจริง", False)
        elif viewing_history:
            self._show_banner("warn", "HISTORY VIEW — LIVE FOLLOW PAUSED",
                              "ค่าบนกราฟไม่ใช่ค่าปัจจุบัน", True)
        else:
            self._hide_banner()

    def _show_banner(self, kind, title, detail, show_button):
        palette = {"warn":  (T.WARN_SOFT, T.WARN_LINE, T.WARN),
                   "info":  (T.ACCENT_SOFT, T.ACCENT_LINE, T.ACCENT),
                   "error": (T.ERROR_SOFT, T.ERROR_LINE, T.ERROR)}[kind]
        bg, line, dot = palette
        if self._banner_kind != kind:
            self._banner_kind = kind
            set_bg_deep(self.banner, bg)
            self.banner.config(highlightbackground=line, highlightcolor=line)
            self.bdot.config(bg=bg)
            set_fg(self.btitle, T.TEXT)
            set_fg(self.bdetail, T.TEXT_DIM)
        self.bdot.set(dot)
        set_text(self.btitle, title)
        set_text(self.bdetail, detail)

        if show_button:
            if not self.bbtn.winfo_ismapped():
                self.bbtn.pack(side="right")
        elif self.bbtn.winfo_ismapped():
            self.bbtn.pack_forget()

        if not self._banner_visible:
            self.banner.pack(fill="x", padx=T.SP_3, pady=(T.SP_2, 0),
                             after=self.hdr_sep)
            self._banner_visible = True

    def _hide_banner(self):
        if self._banner_visible:
            self.banner.pack_forget()
            self._banner_visible = False
            self._banner_kind = None

    def refresh_charts(self):
        rows = self._current_rows()
        shown = downsample(rows)
        self.chart.rebuild(self.active(), self.effective_mode())
        self.chart.update(shown, self.follow_live, len(shown) < len(rows))

    # ---------------------------------------------------------------- session
    def _reload_sessions(self):
        self.sessions = load_sessions()
        labels = []
        self._sess_map = {}
        for s in self.sessions:
            try:
                st = datetime.fromisoformat(s["start"])
                en = datetime.fromisoformat(s["end"])
            except Exception:
                continue
            lbl = "SENSOR {:02d}   {:%d %b %H:%M} → {:%H:%M}".format(
                int(s.get("sensor", 0)), st, en)
            labels.append(lbl)
            self._sess_map[lbl] = (st, en)
        labels.reverse()
        self.sess_combo["values"] = [ALL_DATA] + labels
        if self.sess_var.get() not in self._sess_map:
            self.sess_var.set(ALL_DATA)
        self.sess_combo.bind("<<ComboboxSelected>>", self._on_sess)

    def _on_sess(self, _e):
        if self.sess_var.get() == ALL_DATA:
            self.follow_live = True
        else:
            self.follow_live = False
        self.refresh_charts()
        self.refresh_now()

    # ---------------------------------------------------------------- mock
    def toggle_mock(self):
        if not self.mock:
            if not messagebox.askyesno(
                    "Mock data",
                    "สร้างข้อมูลจำลอง 2 ชั่วโมงลงโฟลเดอร์ water_data/\n"
                    "(ถ้ามีไฟล์ของวันนี้อยู่ จะถูกเขียนทับ)\n\n"
                    "แนะนำให้หยุด logger_3ec.py ก่อน เพื่อไม่ให้เขียนชนกัน\n"
                    "ดำเนินการ?"):
                return
            try:
                n = make_mock_csv(120)
            except Exception as e:
                messagebox.showerror("Mock data", str(e))
                return
            self.mock = True
            self.mock_btn.config(text="Stop mock")
            self._log("MOCK", T.WARN, "Generated {:,} simulated rows".format(n))
        else:
            self.mock = False
            self.mock_btn.config(text="Mock data")
            self._log("MOCK", T.TEXT_DIM,
                      "Mock mode off (files left in water_data/)")
        self.refresh_now()
        self.refresh_charts()

    # ---------------------------------------------------------------- export
    def do_export(self, kind):
        rows = self._current_rows()
        if not rows:
            messagebox.showinfo("Export", "ไม่มีข้อมูลในช่วงนี้")
            return
        note = ""
        if kind in ("pdf", "xlsx"):
            note = simpledialog.askstring(
                "Note", "เพิ่ม Note ลงรายงาน (ภาษาอังกฤษสั้น ๆ, เว้นว่างได้):",
                parent=self.root) or ""
        sel = self.sess_var.get()
        tag = (sel if sel in self._sess_map else "{}min".format(self.minutes))
        tag = tag.replace(" ", "_").replace("→", "-").replace(":", "")
        default = "ec3_{}_{:%Y%m%d_%H%M}.{}".format(tag, datetime.now(), kind)
        ext = {"csv": ".csv", "xlsx": ".xlsx", "pdf": ".pdf"}[kind]
        path = filedialog.asksaveasfilename(
            defaultextension=ext, initialfile=default,
            filetypes=[(kind.upper(), "*" + ext)])
        if not path:
            return
        meta = {"sample": sel if sel in self._sess_map else "-", "note": note}
        try:
            if kind == "csv":
                export_csv(rows, path)
            elif kind == "xlsx":
                export_excel(downsample(rows, 100000), path, meta=meta)
            else:
                # ใช้ scientific report ถ้ามี report_3ec.py
                try:
                    from report_3ec import generate_pdf_3ec
                    generate_pdf_3ec(downsample(rows, 5000), path, meta=meta)
                except Exception:
                    export_pdf(downsample(rows, 5000), path)
            self._log("EXPORT", T.OK, "{} · {}".format(
                kind.upper(), os.path.basename(path)))
            messagebox.showinfo("Export", "บันทึกแล้ว:\n{}".format(path))
        except Exception as e:
            self._log("EXPORT", T.ERROR, "{} failed: {}".format(kind.upper(), e))
            messagebox.showerror("Export", "ผิดพลาด: {}".format(e))

    # ---------------------------------------------------------------- loop
    def _schedule(self):
        self.refresh_now()
        self._ticks += 1
        if self._ticks % 2 == 0:
            self.refresh_charts()
        self.root.after(2000, self._schedule)


if __name__ == "__main__":
    _banner()
    # ⚠️ ต้องอยู่ก่อน tk.Tk() เท่านั้น — เรียกทีหลังไม่มีผล
    T.enable_dpi_awareness()
    root = tk.Tk()
    App(root)
    root.mainloop()
