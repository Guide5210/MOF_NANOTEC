#!/usr/bin/env python3
"""
============================================================================
 report_3ec.py — Scientific PDF report สำหรับระบบ EC 3 ตัว
============================================================================
 สร้าง PDF สไตล์เดียวกับ report.py เดิม (header น้ำเงิน, ตารางสถิติแถบสี,
 กราฟต่อ sensor + กล่องสถิติ, แกนเวลาสัมพัทธ์, footer KMITL)
 แต่รองรับ 3 ภาชนะ (EC1/EC2/EC3)

 เรียกจาก desktop_ui.py:
   from report_3ec import generate_pdf_3ec
   generate_pdf_3ec(rows, path, meta={"sample": "..."})
 โดย rows = list ของ {"t": datetime, "ec":[e1,e2,e3], "tw":[t1,t2,t3], "ok":[..]}
============================================================================
"""

from datetime import datetime

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

try:
    from scipy import stats as scipy_stats
    HAVE_SCIPY = True
except Exception:
    HAVE_SCIPY = False

plt.rcParams.update({
    "font.family": "DejaVu Sans", "font.size": 10,
    "axes.linewidth": 0.8, "axes.edgecolor": "#333333",
    "axes.labelcolor": "#222222", "axes.titlesize": 12, "axes.titleweight": "bold",
    "xtick.color": "#333333", "ytick.color": "#333333",
    "xtick.labelsize": 9, "ytick.labelsize": 9, "figure.dpi": 100,
})

C_PRIMARY = "#1F5C99"
C_GREY = "#666666"
C_LINE = "#CCCCCC"
COLORS = ["#1565C0", "#2E7D32", "#EF6C00"]     # 3 ภาชนะ


# ---------------------------------------------------------------- helpers
def pick_time_unit(sec):
    if sec < 120:      return 1.0, "s"
    if sec < 7200:     return 60.0, "min"
    if sec < 172800:   return 3600.0, "h"
    return 86400.0, "day"


def compute_stats(vals):
    x = np.array([v for v in vals if v is not None], dtype=float)
    n = len(x)
    if n == 0:
        return None
    mean = float(np.mean(x))
    sd = float(np.std(x, ddof=1)) if n > 1 else 0.0
    se = sd / np.sqrt(n) if n > 1 else 0.0
    rsd = (sd / mean * 100) if mean else 0.0
    if n > 1:
        tcrit = scipy_stats.t.ppf(0.975, df=n - 1) if HAVE_SCIPY else 1.96
        ci = tcrit * se
    else:
        ci = 0.0
    return {"n": n, "mean": mean, "sd": sd, "se": se, "rsd": rsd,
            "min": float(x.min()), "max": float(x.max()),
            "range": float(x.max() - x.min()), "ci_lo": mean - ci, "ci_hi": mean + ci}


def _header(fig, title, subtitle):
    fig.text(0.08, 0.945, title, fontsize=17, fontweight="bold", color="#222")
    if subtitle:
        fig.text(0.08, 0.917, subtitle, fontsize=9.5, color=C_GREY)
    fig.add_artist(plt.Line2D([0.08, 0.92], [0.905, 0.905],
                   color=C_PRIMARY, lw=2.2, transform=fig.transFigure))


def _footer(fig, page, total):
    fig.text(0.08, 0.035, "ESP32 Water Monitor (EC x3) — KMITL", fontsize=8, color="#999")
    fig.text(0.92, 0.035, f"{page} / {total}", fontsize=8, color="#999", ha="right")


def _section(fig, y, text):
    fig.text(0.08, y, text, fontsize=12.5, fontweight="bold", color=C_PRIMARY)


def _fmt(v, p=1):
    return f"{v:.{p}f}"


def _rel(ts, t0, div):
    return [(t - t0).total_seconds() / div for t in ts]


# ---------------------------------------------------------------- calibration
#  หน้าสุดท้ายของทุกรายงาน = บันทึกการคาลิเบรตล่าสุด
#
#  รายงานการวัดที่ไม่บอกว่าเครื่องมือถูกสอบเทียบเมื่อไหร่และผลเป็นอย่างไร
#  พิสูจน์อะไรไม่ได้เลย  หน้านี้จึงเป็นส่วนหนึ่งของรายงาน ไม่ใช่ของแถม
#  ข้อมูลทั้งกราฟและตารางมาจาก trace ไฟล์เดียวกับที่ Excel ใช้ ตัวเลขจึงตรงกัน
C_CAL_PRE = "#B0663C"
C_CAL_POST = "#2E7D32"
C_STD = "#C62828"

_RESULT_TEXT = {
    "ok": ("PASS", "#2E7D32", "#EAF5EA"),
    "out_of_tolerance": ("OUT OF TOLERANCE", "#C62828", "#FDEDED"),
    "fail": ("FAILED (sensor did not accept command)", "#C62828", "#FDEDED"),
    "unstable": ("ABORTED (reading never settled)", "#B0663C", "#FFF6E8"),
    "unverified": ("UNVERIFIED (could not re-read after cal)", "#B0663C", "#FFF6E8"),
}


def _cal_module():
    try:
        import calibration
        return calibration
    except Exception:
        return None


def _cal_record(idx):
    """(entry, trace) ของการคาลิเบรตล่าสุดของ sensor idx — (None, []) ถ้าไม่มี"""
    cal = _cal_module()
    if cal is None:
        return None, []
    try:
        e = cal.latest_record(idx)
    except Exception:
        return None, []
    if not e:
        return None, []
    return e, cal.load_trace(e.get("trace_file"))


def _cal_x(trace, entry, idx_phase=None):
    """แกน x = วินาทีเทียบกับ 'ตอนยิงคำสั่ง cal' (0 = จุดที่คาลิเบรต)"""
    off = float(entry.get("cal_offset_s") or 0.0)
    return [t["elapsed"] - off for t in trace]


def _no_cal_box(fig, msg):
    fig.patches.append(plt.Rectangle((0.08, 0.60), 0.84, 0.14,
                       transform=fig.transFigure, facecolor="#FFF6E8",
                       edgecolor="#E0C08A", lw=0.9, zorder=0))
    fig.text(0.5, 0.67, msg, fontsize=12, color="#8A5A20",
             ha="center", va="center", zorder=1)


def _cal_result_badge(fig, y, entry):
    label, col, bg = _RESULT_TEXT.get(entry.get("result", ""),
                                      (str(entry.get("result", "-")), "#666", "#F2F2F2"))
    fig.patches.append(plt.Rectangle((0.08, y - 0.006), 0.84, 0.026,
                       transform=fig.transFigure, facecolor=bg,
                       edgecolor=col, lw=0.9, zorder=0))
    fig.text(0.10, y + 0.003, f"Result: {label}", fontsize=10.5,
             fontweight="bold", color=col, zorder=1)
    if entry.get("after") is not None and entry.get("standard"):
        fig.text(0.90, y + 0.003,
                 f"{entry['after']:.1f} vs {entry['standard']:g} uS/cm  "
                 f"({entry.get('error_us', 0):+.2f})",
                 fontsize=9.5, color=col, ha="right", zorder=1)


def _cal_stats_table(fig, y, rowdefs, first_col_w="Phase"):
    """rowdefs = list ของ (label, stats-dict)  คืน y ล่างสุดหลังวาดเสร็จ"""
    headers = [first_col_w, "n", "Mean", "SD", "SE", "RSD%", "Min", "Max", "95% CI"]
    xcols = [0.09, 0.295, 0.365, 0.455, 0.535, 0.615, 0.685, 0.755, 0.830]
    fig.patches.append(plt.Rectangle((0.08, y - 0.004), 0.84, 0.022,
                       transform=fig.transFigure, facecolor=C_PRIMARY,
                       edgecolor="none", zorder=0))
    for h, xc in zip(headers, xcols):
        fig.text(xc, y + 0.003, h, fontsize=8.5, fontweight="bold",
                 color="white", zorder=1)
    y -= 0.026
    for k, (label, st) in enumerate(rowdefs):
        if k % 2 == 0:
            fig.patches.append(plt.Rectangle((0.08, y - 0.004), 0.84, 0.020,
                               transform=fig.transFigure, facecolor="#F2F6FA",
                               edgecolor="none", zorder=0))
        if not st:
            fig.text(0.09, y + 0.002, f"{label}   (no data)", fontsize=8,
                     color="#999", zorder=1)
            y -= 0.021
            continue
        cells = [label, str(st["n"]), _fmt(st["mean"], 2), _fmt(st["sd"], 3),
                 _fmt(st["se"], 3), _fmt(st["rsd"], 2), _fmt(st["min"]),
                 _fmt(st["max"]), f"[{_fmt(st['ci_lo'], 2)}, {_fmt(st['ci_hi'], 2)}]"]
        for c, xc in zip(cells, xcols):
            fig.text(xc, y + 0.002, c, fontsize=(8 if xc < 0.830 else 7),
                     color="#222", zorder=1)
        y -= 0.021
    return y


def _draw_cal_axes(ax, trace, entry, idx, color, label=None, show_used=True):
    """วาด trace ของ sensor idx ลงแกนที่ให้มา"""
    x = _cal_x(trace, entry)
    pre = [(xx, t["ec"][idx]) for xx, t in zip(x, trace)
           if t["phase"] == "pre" and t["ec"][idx] is not None]
    post = [(xx, t["ec"][idx]) for xx, t in zip(x, trace)
            if t["phase"] == "post" and t["ec"][idx] is not None]
    if pre:
        ax.plot([a for a, _ in pre], [b for _, b in pre], "-o", ms=2.6, lw=1.0,
                color=color, alpha=0.55,
                label=(label + " before" if label else "before cal"))
    if post:
        ax.plot([a for a, _ in post], [b for _, b in post], "-o", ms=2.6, lw=1.3,
                color=color, label=(label + " after" if label else "after cal"))
    if show_used:
        used = [(xx, t["ec"][idx]) for xx, t in zip(x, trace)
                if t["used"] and t["ec"][idx] is not None]
        if used:
            ax.plot([a for a, _ in used], [b for _, b in used], "o", ms=5.5,
                    mfc="none", mec=color, mew=1.1)


def _page_calibration_one(pdf, idx, page_no, total):
    """หน้าคาลิเบรตของ sensor ตัวเดียว"""
    fig = plt.figure(figsize=(8.27, 11.69))
    _header(fig, f"Calibration Record — Container #{idx+1}",
            "latest in-sensor single-point calibration (SEN0706 register 0110H)")

    entry, trace = _cal_record(idx)
    if not entry:
        _no_cal_box(fig, f"No calibration record for Container #{idx+1}\n"
                         f"run the calibration routine before the next experiment")
        _footer(fig, page_no, total)
        pdf.savefig(fig); plt.close(fig)
        return

    _section(fig, 0.875, "1. Calibration Details")
    y = 0.845
    dur_t = entry.get("dur_total_s")
    info = [
        ("Calibrated at", (entry.get("cal_time") or entry.get("time", "-")).replace("T", " ")),
        ("Standard solution", f"{entry.get('standard', '-'):g} uS/cm"),
        ("Reading before cal", f"{entry.get('before', '-')} uS/cm"),
        ("Reading after cal", f"{entry.get('after', '-')} uS/cm"),
        ("Deviation from standard", f"{entry.get('error_us', '-')} uS/cm "
                                    f"({entry.get('error_pct', '-')} %)"),
        ("Stability criterion", f"spread <= {entry.get('stable_spread_limit', '-')} uS/cm "
                                f"over {entry.get('stable_window', '-')} readings"),
        ("Pass criterion", f"|deviation| <= {entry.get('verify_tolerance_pct', 5):.0f} % "
                           f"of standard"),
        ("Time — stabilize", f"{entry.get('dur_pre_s', '-')} s"),
        ("Time — command + settle", f"{entry.get('dur_cal_s', '-')} s"),
        ("Time — verify", f"{entry.get('dur_post_s', '-')} s"),
        ("Time — total", f"{dur_t} s" if dur_t is not None else "-"),
    ]
    for k, v in info:
        fig.text(0.09, y, k, fontsize=9.5, color=C_GREY)
        fig.text(0.42, y, str(v), fontsize=9.5, color="#111")
        y -= 0.021

    # ตำแหน่งจากนี้กำหนดตายตัว — เนื้อหาหน้านี้ยาวคงที่ การไล่ y ต่อกันไป
    # เรื่อย ๆ ทำให้ป้ายแกน x ไปชนหัวข้อถัดไปเมื่อจำนวนบรรทัดเปลี่ยน
    _cal_result_badge(fig, 0.596, entry)

    # ---- กราฟ ----
    _section(fig, 0.548, "2. Calibration Trace")
    ax = fig.add_axes([0.12, 0.315, 0.80, 0.205])
    std = float(entry.get("standard") or 0)
    if trace:
        _draw_cal_axes(ax, trace, entry, idx, COLORS[idx])
        ax.axhline(std, color=C_STD, lw=1.2, ls="--",
                   label=f"standard {std:g}")
        ax.axvline(0, color="#555", lw=1.0, ls=":", label="cal command")
        ax.set_xlabel("Time relative to calibration command (s)")
        ax.set_ylabel("EC (uS/cm)")
        ax.grid(True, alpha=0.25, lw=0.6)
        ax.margins(x=0.02)
        ax.legend(fontsize=7.5, loc="best", framealpha=0.9)
    else:
        ax.text(0.5, 0.5, "no trace data recorded for this event",
                ha="center", va="center", fontsize=10, color="#999")
        ax.set_xticks([]); ax.set_yticks([])

    # ---- ตารางสถิติ ----
    y = 0.252
    _section(fig, y, "3. Statistics of readings used for the decision")
    y -= 0.035
    cal = _cal_module()
    rows = []
    if cal and trace:
        for phase, label in (("pre", "Before cal"), ("post", "After cal")):
            rows.append((label, cal.cal_stats(
                [t["ec"][idx] for t in trace if t["phase"] == phase and t["used"]])))
    else:
        rows = [("Before cal", entry.get("stats_pre")),
                ("After cal", entry.get("stats_post"))]
    y = _cal_stats_table(fig, y, rows)

    y -= 0.018
    fig.text(0.08, y,
             "Circled points are the readings that met the stability criterion and were\n"
             "used to compute the statistics above. Same data set as the matching\n"
             "workbook in calibration_data/.",
             fontsize=8, color=C_GREY, va="top")

    _footer(fig, page_no, total)
    pdf.savefig(fig); plt.close(fig)


def _page_calibration_all(pdf, page_no, total):
    """หน้าคาลิเบรตรวมทั้ง 3 ตัว"""
    fig = plt.figure(figsize=(8.27, 11.69))
    _header(fig, "Calibration Record — all three sensors",
            "latest in-sensor single-point calibration of each sensor")

    recs = [_cal_record(i) for i in range(3)]
    if not any(e for e, _ in recs):
        _no_cal_box(fig, "No calibration records found\n"
                         "run the calibration routine before the next experiment")
        _footer(fig, page_no, total)
        pdf.savefig(fig); plt.close(fig)
        return

    # ---- ตารางสรุปครั้งล่าสุดของแต่ละตัว ----
    _section(fig, 0.875, "1. Latest Calibration of Each Sensor")
    y = 0.842
    headers = ["Sensor", "Calibrated at", "Std", "Before", "After", "Dev.", "Result"]
    xcols = [0.09, 0.20, 0.395, 0.465, 0.545, 0.635, 0.725]
    fig.patches.append(plt.Rectangle((0.08, y - 0.004), 0.84, 0.022,
                       transform=fig.transFigure, facecolor=C_PRIMARY,
                       edgecolor="none", zorder=0))
    for h, xc in zip(headers, xcols):
        fig.text(xc, y + 0.003, h, fontsize=8.5, fontweight="bold",
                 color="white", zorder=1)
    y -= 0.026
    for i, (e, _tr) in enumerate(recs):
        if i % 2 == 0:
            fig.patches.append(plt.Rectangle((0.08, y - 0.004), 0.84, 0.020,
                               transform=fig.transFigure, facecolor="#F2F6FA",
                               edgecolor="none", zorder=0))
        if not e:
            fig.text(0.09, y + 0.002, f"#{i+1}", fontsize=8, color="#222", zorder=1)
            fig.text(0.20, y + 0.002, "never calibrated", fontsize=8,
                     color="#C62828", zorder=1)
            y -= 0.021
            continue
        label, col, _bg = _RESULT_TEXT.get(e.get("result", ""),
                                           (str(e.get("result", "-")), "#666", ""))
        cells = [f"#{i+1}",
                 (e.get("cal_time") or e.get("time", "-")).replace("T", " ")[:19],
                 f"{e.get('standard', 0):g}",
                 f"{e.get('before', '-')}", f"{e.get('after', '-')}",
                 f"{e.get('error_us', 0):+.2f}" if e.get("error_us") is not None else "-"]
        for c, xc in zip(cells, xcols):
            fig.text(xc, y + 0.002, c, fontsize=8, color="#222", zorder=1)
        fig.text(xcols[6], y + 0.002, label, fontsize=8, color=col,
                 fontweight="bold", zorder=1)
        y -= 0.021

    # ---- ความต่างระหว่าง 3 ตัว ----
    y -= 0.022
    afters = [e.get("after") for e, _ in recs if e and e.get("after") is not None]
    stds = {e.get("standard") for e, _ in recs if e}
    if len(afters) == 3 and len(stds) == 1:
        spread = max(afters) - min(afters)
        std = float(list(stds)[0])
        limit = std * 0.03
        okc = spread <= limit
        col = "#2E7D32" if okc else "#C62828"
        bg = "#EAF5EA" if okc else "#FDEDED"
        fig.patches.append(plt.Rectangle((0.08, y - 0.030), 0.84, 0.050,
                           transform=fig.transFigure, facecolor=bg,
                           edgecolor=col, lw=0.9, zorder=0))
        fig.text(0.10, y + 0.006, "Agreement between sensors after calibration",
                 fontsize=10, fontweight="bold", color=col, zorder=1)
        fig.text(0.10, y - 0.018,
                 f"max spread = {spread:.2f} uS/cm   "
                 f"({'within' if okc else 'EXCEEDS'} 3% of standard = {limit:.2f} uS/cm)",
                 fontsize=9.5, color=col, zorder=1)

    # ---- กราฟรวม (ตำแหน่งตายตัว กันป้ายแกนชนหัวข้อ) ----
    _section(fig, 0.665, "2. Calibration Traces (aligned at the calibration command)")
    ax = fig.add_axes([0.12, 0.395, 0.80, 0.245])
    drew = False
    for i, (e, tr) in enumerate(recs):
        if not e or not tr:
            continue
        _draw_cal_axes(ax, tr, e, i, COLORS[i], label=f"#{i+1}", show_used=False)
        drew = True
    if drew:
        for st in sorted(stds):
            if st:
                ax.axhline(float(st), color=C_STD, lw=1.1, ls="--")
        ax.axvline(0, color="#555", lw=1.0, ls=":")
        ax.set_xlabel("Time relative to each sensor's calibration command (s)")
        ax.set_ylabel("EC (uS/cm)")
        ax.grid(True, alpha=0.25, lw=0.6)
        ax.margins(x=0.02)
        ax.legend(fontsize=7, ncol=3, loc="best", framealpha=0.9)
    else:
        ax.text(0.5, 0.5, "no trace data recorded", ha="center", va="center",
                fontsize=10, color="#999")
        ax.set_xticks([]); ax.set_yticks([])

    # ---- ตารางสถิติ 3 ตัว ----
    y = 0.330
    _section(fig, y, "3. Statistics After Calibration (readings used for the decision)")
    y -= 0.035
    cal = _cal_module()
    rows = []
    for i, (e, tr) in enumerate(recs):
        st = None
        if cal and tr:
            st = cal.cal_stats([t["ec"][i] for t in tr
                                if t["phase"] == "post" and t["used"]])
        elif e:
            st = e.get("stats_post")
        rows.append((f"Container #{i+1}", st))
    y = _cal_stats_table(fig, y, rows, first_col_w="Sensor")

    y -= 0.018
    fig.text(0.08, y,
             "Each sensor is calibrated in its own event; traces are time-aligned at\n"
             "t = 0 (the moment its calibration command was sent). SD = sample standard\n"
             "deviation (n-1). Same data set as the workbooks in calibration_data/.",
             fontsize=8, color=C_GREY, va="top")

    _footer(fig, page_no, total)
    pdf.savefig(fig); plt.close(fig)


# ---------------------------------------------------------------- pages
def _page_cover(pdf, rows, stats, meta, total):
    fig = plt.figure(figsize=(8.27, 11.69))
    _header(fig, "Water Quality Report — EC (3 samples)",
            "ESP32 3x RS485 EC sensors — Automated Analysis")

    _section(fig, 0.875, "1. Measurement Parameters")
    dur = rows[-1]["t"] - rows[0]["t"]
    params = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Sample / ID", meta.get("sample", "-")),
        ("Instrument", "ESP32 CONTROL V1.2 + 3x SEN0706 (RS485)"),
        ("Data span", f"{rows[0]['t']:%Y-%m-%d %H:%M} - {rows[-1]['t']:%Y-%m-%d %H:%M}"),
        ("Duration", str(dur)),
        ("Sampling", "~2.5 s / point"),
        ("Points", str(len(rows))),
        ("EC calibration", "in-sensor single-point (register 0110H)"),
    ]
    yy = 0.845
    for k, v in params:
        fig.text(0.10, yy, k, fontsize=9.5, color="#333")
        fig.text(0.42, yy, str(v), fontsize=9.5, color="#111")
        yy -= 0.023

    yy -= 0.02
    _section(fig, yy, "2. Summary Statistics")
    yy -= 0.04
    headers = ["Sample", "n", "Mean", "SD", "SE", "RSD%", "Min", "Max", "95% CI"]
    xcols = [0.09, 0.27, 0.35, 0.45, 0.54, 0.62, 0.70, 0.78, 0.865]
    fig.patches.append(plt.Rectangle((0.08, yy - 0.004), 0.84, 0.022,
                       transform=fig.transFigure, facecolor=C_PRIMARY, edgecolor="none", zorder=0))
    for h, xc in zip(headers, xcols):
        fig.text(xc, yy + 0.003, h, fontsize=8.5, fontweight="bold", color="white", zorder=1)
    yy -= 0.026
    for i in range(3):
        s = stats[i]
        if not s:
            continue
        if i % 2 == 0:
            fig.patches.append(plt.Rectangle((0.08, yy - 0.004), 0.84, 0.020,
                               transform=fig.transFigure, facecolor="#F2F6FA", edgecolor="none", zorder=0))
        cells = [f"#{i+1} (uS/cm)", str(s["n"]), _fmt(s["mean"]), _fmt(s["sd"]),
                 _fmt(s["se"]), _fmt(s["rsd"], 2), _fmt(s["min"]), _fmt(s["max"]),
                 f"[{_fmt(s['ci_lo'])}, {_fmt(s['ci_hi'])}]"]
        for c, xc in zip(cells, xcols):
            fig.text(xc, yy + 0.002, c, fontsize=(8 if xc < 0.865 else 7), color="#222", zorder=1)
        yy -= 0.021

    yy -= 0.015
    note = ("SD = sample standard deviation (n-1).  SE = SD/sqrt(n).  RSD% = 100*SD/Mean.\n"
            "95% CI: Student's t-distribution" + ("" if HAVE_SCIPY else " (normal approx.)") +
            ".  EC temperature-compensated in-sensor.")
    fig.text(0.08, yy, note, fontsize=8, color=C_GREY, va="top")

    # ---- Note ของผู้ใช้ (ถ้ามี) ----
    user_note = (meta.get("note") or "").strip()
    if user_note:
        import textwrap
        ny = yy - 0.06
        _section(fig, ny, "3. Note")
        ny -= 0.03
        wrapped = []
        for para in user_note.split("\n"):
            wrapped += textwrap.wrap(para, width=95) or [""]
        box_h = 0.018 * len(wrapped) + 0.02
        fig.patches.append(plt.Rectangle((0.08, ny - box_h + 0.012), 0.84, box_h,
                           transform=fig.transFigure, facecolor="#FFFDF5",
                           edgecolor="#E0D8B0", lw=0.8, zorder=0))
        fig.text(0.10, ny, "\n".join(wrapped), fontsize=9.5, color="#333", va="top", zorder=1)

    _footer(fig, 1, total)
    pdf.savefig(fig); plt.close(fig)


def _page_sensor(pdf, rows, stats, idx, page_no, total, t0, div, uname):
    fig = plt.figure(figsize=(8.27, 11.69))
    _header(fig, f"EC ภาชนะ #{idx+1} vs Time".replace("ภาชนะ ", "Container "), "raw data")

    ts = [r["t"] for r in rows]
    x = _rel(ts, t0, div)
    y = [r["ec"][idx] for r in rows]
    ax = fig.add_axes([0.12, 0.46, 0.80, 0.40])
    ax.plot(x, y, color=COLORS[idx], lw=1.1)
    ax.set_xlabel(f"Time ({uname})"); ax.set_ylabel("EC (uS/cm)")
    ax.grid(True, alpha=0.25, lw=0.6); ax.margins(x=0.01)

    s = stats[idx]
    if s:
        _section(fig, 0.37, "Descriptive Statistics")
        fig.patches.append(plt.Rectangle((0.08, 0.135), 0.84, 0.205,
                           transform=fig.transFigure, facecolor="#F7FAFC",
                           edgecolor=C_LINE, lw=0.8, zorder=0))
        lines = [
            ("n", f"{s['n']}"),
            ("Mean +/- SD", f"{_fmt(s['mean'])} +/- {_fmt(s['sd'])} uS/cm"),
            ("Standard Error", f"{_fmt(s['se'])} uS/cm"),
            ("RSD%", f"{_fmt(s['rsd'],2)} %"),
            ("Min / Max", f"{_fmt(s['min'])} / {_fmt(s['max'])} uS/cm"),
            ("Range", f"{_fmt(s['range'])} uS/cm"),
            ("95% CI", f"[{_fmt(s['ci_lo'])}, {_fmt(s['ci_hi'])}] uS/cm"),
        ]
        yy = 0.315
        for k, v in lines:
            fig.text(0.11, yy, k, fontsize=9.5, color="#333", fontweight="bold")
            fig.text(0.40, yy, v, fontsize=9.5, color="#111")
            yy -= 0.026

    _footer(fig, page_no, total)
    pdf.savefig(fig); plt.close(fig)


def _page_combined(pdf, rows, page_no, total, t0, div, uname):
    fig = plt.figure(figsize=(8.27, 11.69))
    _header(fig, "EC — 3 samples overlay", "raw data, comparison")

    ts = [r["t"] for r in rows]
    x = _rel(ts, t0, div)
    ax = fig.add_axes([0.12, 0.46, 0.80, 0.40])
    for i in range(3):
        ax.plot(x, [r["ec"][i] for r in rows], color=COLORS[i], lw=1.1, label=f"#{i+1}")
    ax.set_xlabel(f"Time ({uname})"); ax.set_ylabel("EC (uS/cm)")
    ax.grid(True, alpha=0.25, lw=0.6); ax.legend(loc="best", framealpha=0.9)
    ax.margins(x=0.01)

    _footer(fig, page_no, total)
    pdf.savefig(fig); plt.close(fig)


def _page_ec_temp(pdf, rows, idx, page_no, total):
    """EC vs Water Temperature (scatter สีตามเวลา + Pearson r) ต่อ sensor"""
    fig = plt.figure(figsize=(8.27, 11.69))
    _header(fig, f"EC vs Water Temperature — Container #{idx+1}",
            "scatter colored by time")

    pairs = [(r["tw"][idx], r["ec"][idx], k)
             for k, r in enumerate(rows)
             if r["tw"][idx] is not None and r["ec"][idx] is not None]
    ax = fig.add_axes([0.12, 0.40, 0.78, 0.46])
    if len(pairs) >= 2:
        import numpy as np
        tw = np.array([p[0] for p in pairs])
        ec = np.array([p[1] for p in pairs])
        ck = np.array([p[2] for p in pairs], dtype=float)
        sc = ax.scatter(tw, ec, c=ck, cmap="viridis", s=10, alpha=0.8)
        cb = fig.colorbar(sc, ax=ax, fraction=0.045, pad=0.02)
        cb.set_label("point index (time order)", fontsize=8)
        # Pearson r
        if tw.std() > 0 and ec.std() > 0:
            r = float(np.corrcoef(tw, ec)[0, 1])
            if HAVE_SCIPY:
                _, p = scipy_stats.pearsonr(tw, ec)
                rtxt = f"Pearson r = {r:.3f}  (p = {p:.2e}, n = {len(pairs)})"
            else:
                rtxt = f"Pearson r = {r:.3f}  (n = {len(pairs)})"
        else:
            rtxt = f"n = {len(pairs)} (no variance)"
        fig.text(0.12, 0.885, rtxt, fontsize=10, fontweight="bold", color="#222")
        fig.text(0.12, 0.865,
                 "Note: EC already temperature-compensated in-sensor; "
                 "residual r reflects sample change over time, not raw T-dependence.",
                 fontsize=7.5, color=C_GREY)
        ax.set_xlabel("Water Temperature (C)"); ax.set_ylabel("EC (uS/cm)")
        ax.grid(True, alpha=0.25, lw=0.6)
    else:
        ax.text(0.5, 0.5, "no data", ha="center", va="center",
                color=C_GREY, transform=ax.transAxes)

    _footer(fig, page_no, total)
    pdf.savefig(fig); plt.close(fig)


# ---------------------------------------------------------------- entry
def generate_pdf_3ec(rows, path, meta=None):
    """รายงานรวม 3 ตัว: cover+สถิติ, EC-time ต่อตัว(3), overlay, EC-vs-temp ต่อตัว(3)"""
    if not rows:
        raise ValueError("ไม่มีข้อมูล")
    if meta is None:
        meta = {}

    stats = [compute_stats([r["ec"][i] for r in rows]) for i in range(3)]
    t0 = rows[0]["t"]
    span = (rows[-1]["t"] - rows[0]["t"]).total_seconds()
    div, uname = pick_time_unit(span)

    # cover + 3 EC-time + overlay + 3 EC-temp + calibration (หน้าสุดท้าย)
    total = 1 + 3 + 1 + 3 + 1
    with PdfPages(path) as pdf:
        _page_cover(pdf, rows, stats, meta, total)          # หน้า 1
        for i in range(3):
            _page_sensor(pdf, rows, stats, i, 2 + i, total, t0, div, uname)   # 2,3,4
        _page_combined(pdf, rows, 5, total, t0, div, uname)  # หน้า 5 overlay
        for i in range(3):
            _page_ec_temp(pdf, rows, i, 6 + i, total)        # 6,7,8 EC-vs-temp
        _page_calibration_all(pdf, 9, total)                 # หน้า 9 calibration
    return path


# ---------------------------------------------------------------- CSV + CLI
def _num(x):
    try:
        return float(x)
    except (ValueError, TypeError):
        return None


def read_csv_rows(data_dir="water_data", since=None, until=None):
    """อ่าน CSV 3-EC ทุกไฟล์ในโฟลเดอร์ -> rows (กรองช่วงเวลาได้)"""
    import glob
    import os
    rows = []
    for path in sorted(glob.glob(os.path.join(data_dir, "water_log_*.csv"))):
        try:
            with open(path, encoding="utf-8") as fh:
                fh.readline()
                for line in fh:
                    p = line.rstrip("\n").split(",")
                    if len(p) < 10:
                        continue
                    # คอลัมน์ที่ 11 = flag  ("CAL" = เก็บตอนหัววัดอยู่ในน้ำยามาตรฐาน)
                    # ข้อมูลดิบยังอยู่ในไฟล์ครบ แต่ไม่ให้ปนกับข้อมูลการทดลอง
                    if len(p) > 10 and p[10].strip().upper() == "CAL":
                        continue
                    try:
                        t = datetime.strptime(p[0], "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        continue
                    if since and t < since:
                        continue
                    if until and t > until:
                        continue
                    rows.append({"t": t,
                                 "ec": [_num(p[1]), _num(p[3]), _num(p[5])],
                                 "tw": [_num(p[2]), _num(p[4]), _num(p[6])],
                                 "ok": [p[7], p[8], p[9]]})
        except Exception:
            continue
    rows.sort(key=lambda r: r["t"])
    return rows


# ---------------------------------------------------------------- single sensor
def generate_pdf_one(rows, idx, path, meta=None):
    """PDF ของ sensor ตัวเดียว (idx=0/1/2): cover สถิติ + กราฟ + สถิติใต้กราฟ"""
    if not rows:
        raise ValueError("ไม่มีข้อมูล")
    if meta is None:
        meta = {}
    s = compute_stats([r["ec"][idx] for r in rows])
    if not s:
        raise ValueError(f"sensor #{idx+1} ไม่มีข้อมูลที่ใช้ได้")
    t0 = rows[0]["t"]
    span = (rows[-1]["t"] - rows[0]["t"]).total_seconds()
    div, uname = pick_time_unit(span)
    ts = [r["t"] for r in rows]

    with PdfPages(path) as pdf:
        # ---- หน้า 1: cover + parameters + สถิติ ----
        fig = plt.figure(figsize=(8.27, 11.69))
        _header(fig, f"Water Quality Report — EC Container #{idx+1}",
                "ESP32 RS485 EC sensor — single-sample report")
        _section(fig, 0.875, "1. Measurement Parameters")
        dur = rows[-1]["t"] - rows[0]["t"]
        params = [
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Sample / ID", meta.get("sample", "-")),
            ("Sensor", f"Container #{idx+1} (SEN0706 addr {idx+1})"),
            ("Data span", f"{rows[0]['t']:%Y-%m-%d %H:%M} - {rows[-1]['t']:%Y-%m-%d %H:%M}"),
            ("Duration", str(dur)),
            ("Points", str(s["n"])),
            ("EC calibration", "in-sensor single-point (register 0110H)"),
        ]
        yy = 0.845
        for k, v in params:
            fig.text(0.10, yy, k, fontsize=9.5, color="#333")
            fig.text(0.42, yy, str(v), fontsize=9.5, color="#111")
            yy -= 0.023

        yy -= 0.02
        _section(fig, yy, "2. Descriptive Statistics")
        yy -= 0.04
        fig.patches.append(plt.Rectangle((0.08, 0.40), 0.84, yy - 0.40 + 0.02,
                           transform=fig.transFigure, facecolor="#F7FAFC",
                           edgecolor=C_LINE, lw=0.8, zorder=0))
        lines = [
            ("n", f"{s['n']}"),
            ("Mean +/- SD", f"{_fmt(s['mean'])} +/- {_fmt(s['sd'])} uS/cm"),
            ("Standard Error", f"{_fmt(s['se'])} uS/cm"),
            ("RSD%", f"{_fmt(s['rsd'],2)} %"),
            ("Min / Max", f"{_fmt(s['min'])} / {_fmt(s['max'])} uS/cm"),
            ("Range", f"{_fmt(s['range'])} uS/cm"),
            ("95% CI", f"[{_fmt(s['ci_lo'])}, {_fmt(s['ci_hi'])}] uS/cm"),
        ]
        for k, v in lines:
            fig.text(0.11, yy, k, fontsize=10, color="#333", fontweight="bold")
            fig.text(0.40, yy, v, fontsize=10, color="#111")
            yy -= 0.028
        note = ("SD = sample standard deviation (n-1).  SE = SD/sqrt(n).\n"
                "95% CI: Student's t-distribution" + ("" if HAVE_SCIPY else " (normal approx.)") +
                ".  EC temperature-compensated in-sensor.")
        fig.text(0.08, yy - 0.01, note, fontsize=8, color=C_GREY, va="top")

        # ---- Note ของผู้ใช้ (ถ้ามี) ----
        user_note = (meta.get("note") or "").strip()
        if user_note:
            ny = yy - 0.06
            _section(fig, ny, "3. Note")
            ny -= 0.03
            # ตัดบรรทัดยาวให้พอดีหน้า (~90 ตัว/บรรทัด)
            import textwrap
            wrapped = []
            for para in user_note.split("\n"):
                wrapped += textwrap.wrap(para, width=95) or [""]
            box_h = 0.018 * len(wrapped) + 0.02
            fig.patches.append(plt.Rectangle((0.08, ny - box_h + 0.012), 0.84, box_h,
                               transform=fig.transFigure, facecolor="#FFFDF5",
                               edgecolor="#E0D8B0", lw=0.8, zorder=0))
            fig.text(0.10, ny, "\n".join(wrapped), fontsize=9.5, color="#333",
                     va="top", zorder=1)

        _footer(fig, 1, 3)
        pdf.savefig(fig); plt.close(fig)

        # ---- หน้า 2: กราฟ ----
        fig = plt.figure(figsize=(8.27, 11.69))
        _header(fig, f"EC Container #{idx+1} vs Time", "raw data")
        x = _rel(ts, t0, div)
        ax = fig.add_axes([0.12, 0.30, 0.80, 0.56])
        ax.plot(x, [r["ec"][idx] for r in rows], color=COLORS[idx], lw=1.1)
        ax.set_xlabel(f"Time ({uname})"); ax.set_ylabel("EC (uS/cm)")
        ax.grid(True, alpha=0.25, lw=0.6); ax.margins(x=0.01)
        _footer(fig, 2, 3)
        pdf.savefig(fig); plt.close(fig)

        # ---- หน้า 3: บันทึกการคาลิเบรตของ sensor ตัวนี้ ----
        _page_calibration_one(pdf, idx, 3, 3)
    return path


def generate_excel_one(rows, idx, path, meta=None):
    """Excel ของ sensor ตัวเดียว: Summary + raw"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    if meta is None:
        meta = {}
    s = compute_stats([r["ec"][idx] for r in rows])
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = f"EC Container #{idx+1} — Report"
    ws["A1"].font = Font(size=14, bold=True, color="1F5C99")
    info = [("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Sample", meta.get("sample", "-")),
            ("Sensor", f"Container #{idx+1} (addr {idx+1})")]
    if rows:
        info.append(("Span", f"{rows[0]['t']:%Y-%m-%d %H:%M:%S} - {rows[-1]['t']:%Y-%m-%d %H:%M:%S}"))
    if (meta.get("note") or "").strip():
        info.append(("Note", meta["note"].strip()))
    r = 3
    for k, v in info:
        ws.cell(r, 1, k).font = Font(bold=True); ws.cell(r, 2, v); r += 1
    r += 1
    if s:
        hdr = ["n", "Mean", "SD", "SE", "RSD%", "Min", "Max", "Range"]
        for c, h in enumerate(hdr, 1):
            cell = ws.cell(r, c, h); cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F5C99")
        r += 1
        vals = [s["n"], round(s["mean"], 1), round(s["sd"], 2), round(s["se"], 2),
                round(s["rsd"], 2), round(s["min"], 1), round(s["max"], 1), round(s["range"], 1)]
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, v)
    for col, w in zip("ABCDEFGH", [12, 10, 10, 8, 8, 10, 10, 10]):
        ws.column_dimensions[col].width = w

    wr = wb.create_sheet("raw_data")
    wr.append(["timestamp", f"EC{idx+1}", f"T{idx+1}", "ok"])
    for row in rows:
        wr.append([row["t"].strftime("%Y-%m-%d %H:%M:%S"),
                   row["ec"][idx], row["tw"][idx], row["ok"][idx]])
    wb.save(path)
    return path


def export_sensor_session(idx, since, until=None, data_dir="water_data",
                          out_dir=None, sample="-", auto_open=False, note=""):
    """สร้าง PDF+Excel ของ sensor ตัวเดียว ลง folder ของตัวเอง"""
    import os
    if out_dir is None:
        base = os.path.dirname(os.path.abspath(data_dir)) if os.path.isabs(data_dir) \
            else os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(base, f"sensor_{idx+1}")
    os.makedirs(out_dir, exist_ok=True)

    rows = read_csv_rows(data_dir, since=since, until=until)
    # กรองเฉพาะแถวที่ sensor นี้มีค่า (ok)
    rows = [r for r in rows if r["ec"][idx] is not None]
    if not rows:
        print(f"[report3] sensor #{idx+1}: no data in the session window")
        return None

    meta = {"sample": sample, "note": note}
    stamp = since.strftime("%Y%m%d_%H%M%S")
    pdf_path = os.path.join(out_dir, f"EC{idx+1}_session_{stamp}.pdf")
    xlsx_path = os.path.join(out_dir, f"EC{idx+1}_session_{stamp}.xlsx")
    generate_pdf_one(rows, idx, pdf_path, meta=meta)
    generate_excel_one(rows, idx, xlsx_path, meta=meta)
    print(f"[report3] EC#{idx+1} -> {pdf_path}")
    print(f"[report3] EC#{idx+1} -> {xlsx_path}")
    if auto_open:
        open_file(pdf_path)
    return pdf_path, xlsx_path


def generate_excel_3ec(rows, path, meta=None):
    """Excel รวม 3 ตัว: Summary (สถิติ 3 ตัว) + raw ทั้ง 3 คอลัมน์"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    if meta is None:
        meta = {}
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "Water Quality Report — EC (3 samples)"
    ws["A1"].font = Font(size=14, bold=True, color="1F5C99")
    r = 3
    info = [("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Sample", meta.get("sample", "-"))]
    if rows:
        info.append(("Span", f"{rows[0]['t']:%Y-%m-%d %H:%M:%S} - {rows[-1]['t']:%Y-%m-%d %H:%M:%S}"))
    info.append(("Points", len(rows)))
    if (meta.get("note") or "").strip():
        info.append(("Note", meta["note"].strip()))
    for k, v in info:
        ws.cell(r, 1, k).font = Font(bold=True); ws.cell(r, 2, v); r += 1
    r += 1
    hdr = ["Sample", "n", "Mean", "SD", "SE", "RSD%", "Min", "Max"]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(r, c, h); cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F5C99")
    r += 1
    for i in range(3):
        s = compute_stats([row["ec"][i] for row in rows])
        if s:
            for c, v in enumerate([f"#{i+1}", s["n"], round(s["mean"], 1),
                                   round(s["sd"], 2), round(s["se"], 2),
                                   round(s["rsd"], 2), round(s["min"], 1),
                                   round(s["max"], 1)], 1):
                ws.cell(r, c, v)
            r += 1
    for col, w in zip("ABCDEFGH", [12, 8, 10, 10, 8, 8, 10, 10]):
        ws.column_dimensions[col].width = w

    wr = wb.create_sheet("raw_data")
    wr.append(["timestamp", "EC1", "T1", "EC2", "T2", "EC3", "T3", "ok1", "ok2", "ok3"])
    for row in rows:
        wr.append([row["t"].strftime("%Y-%m-%d %H:%M:%S"),
                   row["ec"][0], row["tw"][0], row["ec"][1], row["tw"][1],
                   row["ec"][2], row["tw"][2], *row["ok"]])
    wb.save(path)
    return path


def export_session_report(owner_idx, since, until=None, data_dir="water_data",
                          out_dir=None, sample="-", auto_open=False, note=""):
    """รายงานรวม 3 ตัว (PDF 8 หน้า + Excel) ครอบช่วง session ของ sensor ตัวที่ owner_idx
       เก็บใน folder sensor_{owner+1}/ (เจ้าของ session ที่กดหยุด)"""
    import os
    if out_dir is None:
        base = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(base, f"sensor_{owner_idx+1}")
    os.makedirs(out_dir, exist_ok=True)

    rows = read_csv_rows(data_dir, since=since, until=until)
    if not rows:
        print(f"[report3] session EC#{owner_idx+1}: no data in the window")
        return None

    meta = {"sample": sample, "note": note}
    stamp = since.strftime("%Y%m%d_%H%M%S")
    pdf_path = os.path.join(out_dir, f"EC{owner_idx+1}_session_{stamp}.pdf")
    xlsx_path = os.path.join(out_dir, f"EC{owner_idx+1}_session_{stamp}.xlsx")
    generate_pdf_3ec(rows, pdf_path, meta=meta)          # รายงานรวม 3 ตัว
    generate_excel_3ec(rows, xlsx_path, meta=meta)
    print(f"[report3] session EC#{owner_idx+1} (all 3 sensors) -> {pdf_path}")
    print(f"[report3] session EC#{owner_idx+1} (all 3 sensors) -> {xlsx_path}")
    if auto_open:
        open_file(pdf_path)
    return pdf_path, xlsx_path


def export_combined_report(since, until=None, data_dir="water_data",
                           out_dir=None, sample="-", note="", auto_open=False):
    """รายงานรวม 3 ตัว (PDF 8 หน้า + Excel) ของทั้ง run — เก็บใน reports/
       ใช้ตอนปิด terminal (q / Ctrl+C)"""
    import os
    if out_dir is None:
        base = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(base, "reports")
    os.makedirs(out_dir, exist_ok=True)

    rows = read_csv_rows(data_dir, since=since, until=until)
    if not rows:
        print("[report3] run: no data in the window - skipping combined report")
        return None

    meta = {"sample": sample, "note": note}
    stamp = (since or datetime.now()).strftime("%Y%m%d_%H%M%S")
    pdf_path = os.path.join(out_dir, f"run_3ec_{stamp}.pdf")
    xlsx_path = os.path.join(out_dir, f"run_3ec_{stamp}.xlsx")
    generate_pdf_3ec(rows, pdf_path, meta=meta)
    generate_excel_3ec(rows, xlsx_path, meta=meta)
    print(f"[report3] combined run report (3 sensors) -> {pdf_path}")
    print(f"[report3] combined run report (3 sensors) -> {xlsx_path}")
    if auto_open:
        open_file(pdf_path)
    return pdf_path, xlsx_path


def open_file(path):
    """เปิดไฟล์ด้วยโปรแกรมเริ่มต้นของ OS"""
    import sys
    import os
    import subprocess
    try:
        if sys.platform.startswith("win"):
            os.startfile(os.path.abspath(path))
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
        print(f"[report3] opened: {path}")
    except Exception as e:
        print(f"[report3] cannot open automatically ({e}) - open it at {os.path.abspath(path)}")


def generate_from_csv(data_dir="water_data", output=None, since=None, until=None,
                      auto_open=False, meta=None):
    """อ่าน CSV -> สร้าง PDF (ใช้จาก logger ตอน Ctrl+C หรือสั่ง CLI)"""
    import os
    rows = read_csv_rows(data_dir, since=since, until=until)
    if not rows:
        print("[report3] no data in the requested window")
        return None
    if output is None:
        output = f"report_3ec_{datetime.now():%Y%m%d_%H%M%S}"
    if not output.endswith(".pdf"):
        output += ".pdf"
    generate_pdf_3ec(rows, output, meta=meta or {})
    print(f"[report3] PDF written: {output}  ({len(rows)} rows)")
    if auto_open:
        open_file(output)
    return output


if __name__ == "__main__":
    import argparse
    import os
    ap = argparse.ArgumentParser(description="Build the 3-sensor EC PDF report from the CSV files")
    ap.add_argument("--data-dir", default="water_data", help="folder containing the CSV files")
    ap.add_argument("--output", default=None, help="output PDF name (omit = auto)")
    ap.add_argument("--sample", default="-", help="sample name / ID")
    ap.add_argument("--open", action="store_true", help="open the PDF when done")
    ap.add_argument("--selftest", action="store_true", help="self-test with simulated data")
    args = ap.parse_args()

    if args.selftest:
        import random
        from datetime import timedelta
        now = datetime.now()
        rows = []
        for k in range(2000):
            t = now - timedelta(seconds=(2000 - k) * 2)
            rows.append({"t": t,
                         "ec": [1400 - k*0.1 + random.gauss(0, 1),
                                300 + random.gauss(0, 1),
                                50 + k*0.02 + random.gauss(0, 1)],
                         "tw": [20.5]*3, "ok": ["1", "1", "1"]})
        generate_pdf_3ec(rows, "test_3ec.pdf", {"sample": "CALF-20 test"})
        print("wrote test_3ec.pdf")
    else:
        res = generate_from_csv(data_dir=args.data_dir, output=args.output,
                                auto_open=args.open, meta={"sample": args.sample})
        if res is None:
            raise SystemExit(1)

