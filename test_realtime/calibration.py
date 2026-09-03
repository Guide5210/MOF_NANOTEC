#!/usr/bin/env python3
"""
============================================================================
 calibration.py — คาลิเบรตเซนเซอร์ EC (DFRobot SEN0706) แบบมีขั้นตอนนำ
============================================================================
 ผู้ใช้ต้อง cal ทุกวัน จึงทำให้เป็นขั้นตอนสั้น ๆ ที่กดผ่านคีย์เดียวได้
 และบันทึกประวัติทุกครั้งไว้เป็นหลักฐานย้อนกลับ (สำคัญสำหรับงานแล็บ)

 หลักการ
 -------
 SEN0706 คาลิเบรตที่ตัวเซนเซอร์เองผ่าน Modbus ไม่ใช่การชดเชยด้วยซอฟต์แวร์
 ค่าที่คาลิเบรตแล้วจึงติดอยู่กับเซนเซอร์ ใช้ได้กับทั้งไฟล์ CSV และจอ ESP32-P4
 พร้อมกัน โดยไม่ต้องแก้อะไรสองที่

   เขียน register 0x0110 = 0x0004        (รหัสสั่งคาลิเบรต EC)
   เขียน register 0x0111 = ค่ามาตรฐาน x 10   (1413 uS/cm -> 14130 = 0x3732)
   ด้วย function 0x10 (write multiple registers) ทีเดียวสองตัว

 แต่ Python ต่อกับบอร์ด ESP32 ผ่าน USB ไม่ได้ต่อกับบัส RS485 โดยตรง
 จึงส่งคำสั่งผ่านบอร์ด ESP32 ให้มันเป็นคนยิง Modbus ให้:

   ส่งไปที่บอร์ด:   C<n>,<uS>      เช่น  C1,1413
   บอร์ดตอบกลับ:    [cal] ok n=1 ...   หรือ  [cal] fail n=1 rc=...

 ขั้นตอนที่ผู้ใช้เจอ
 -------------------
   1. ล้างหัววัดด้วยน้ำ DI แล้วซับให้หมาด
   2. จุ่มลงน้ำยามาตรฐาน 1413 uS/cm คนเบา ๆ แล้วปล่อยนิ่ง
   3. โปรแกรมเฝ้าดูจนค่านิ่ง (ไม่แกว่งเกินเกณฑ์ติดกันหลายวินาที)
   4. ยิงคำสั่งคาลิเบรต แล้ววัดซ้ำเพื่อยืนยันว่าเข้าใกล้ค่ามาตรฐานจริง
============================================================================
"""

import json
import os
import sys
import time
from datetime import datetime

# console_utf8 ทำหน้าที่แค่ให้คอนโซล Windows แสดงภาษาไทยได้ — เป็นเรื่องการแสดงผล
# ล้วน ๆ ไม่เกี่ยวกับการเก็บข้อมูล  ถ้าไฟล์นี้ไม่อยู่ในโฟลเดอร์ โปรแกรมต้องยัง
# ทำงานได้ตามปกติ ไม่ใช่ตายทั้งตัว  (เคยเกิดจริงตอนคัดลอกโปรเจกต์ไปอีกเครื่อง
# แล้วลืมไฟล์นี้ — logger ล้มทั้งระบบเพราะ helper ที่ไม่สำคัญเลย)
try:
    import console_utf8
    console_utf8.enable()
except ImportError:
    def _enable_utf8_console():
        """สำเนาย่อของ console_utf8.enable() — ใช้เมื่อหาไฟล์นั้นไม่เจอ"""
        if sys.platform == "win32":
            try:
                import ctypes
                ctypes.windll.kernel32.SetConsoleOutputCP(65001)
                ctypes.windll.kernel32.SetConsoleCP(65001)
            except Exception:
                pass
        for _st in (sys.stdout, sys.stderr):
            try:
                _st.reconfigure(encoding="utf-8", errors="replace")
            except (AttributeError, ValueError):
                pass
    _enable_utf8_console()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CAL_LOG = os.path.join(BASE_DIR, "calibration_log.json")

# โฟลเดอร์เก็บข้อมูลดิบ + Excel ของการคาลิเบรตแต่ละครั้ง
CAL_DATA_DIR = os.path.join(BASE_DIR, "calibration_data")
CAL_MASTER_XLSX = os.path.join(CAL_DATA_DIR, "calibration_log.xlsx")

# น้ำยามาตรฐานที่ใช้บ่อย (uS/cm) — SEN0706 K=1 วัดได้ 1~2000 uS/cm
STANDARDS = [84, 1413]
DEFAULT_STANDARD = 84      # จุดใช้งานจริง: wash endpoint อยู่ช่วง 57-296 uS/cm

# เกณฑ์ว่า "ค่านิ่งพอที่จะคาลิเบรต"
#
# เดิมเกณฑ์เป็นตัวเลขตายตัว 15 uS/cm ซึ่งผูกกับการ cal ที่ 1413 (=1%)
# พอย้ายมา cal ที่ 84 uS/cm เกณฑ์เดิมกลายเป็น ±18% คือหลวมจนไม่ได้กรองอะไรเลย
# จึงเปลี่ยนเป็นสัดส่วนของค่ามาตรฐาน -> ใช้ได้ทุกจุดโดยไม่ต้องแก้เลข
STABLE_WINDOW = 8          # ต้องนิ่งติดกันกี่ค่า
STABLE_SPREAD_PCT = 0.02   # ช่วงกว้างในหน้าต่างต้องไม่เกิน 2% ของค่ามาตรฐาน
STABLE_SPREAD_MIN = 1.0    # แต่ไม่ต่ำกว่านี้ (ความละเอียดหัววัด 0.1 uS/cm)
STABLE_TIMEOUT = 120       # ถ้าไม่นิ่งภายในกี่วินาที ให้เลิก
VERIFY_TOLERANCE = 0.05    # หลัง cal ต้องเพี้ยนไม่เกิน 5% ของค่ามาตรฐาน


def stable_spread(standard_us):
    """เกณฑ์ช่วงกว้างที่ยอมรับว่า 'นิ่ง' ที่ค่ามาตรฐานนั้น (uS/cm)
       84 uS/cm -> 1.7  |  1413 uS/cm -> 28.3"""
    return max(STABLE_SPREAD_MIN, float(standard_us) * STABLE_SPREAD_PCT)


# ชื่อเดิม เผื่อมีโค้ดอื่นอ้างถึง — ค่าที่จุดมาตรฐานเริ่มต้น
STABLE_SPREAD = stable_spread(DEFAULT_STANDARD)


def log_event(entry):
    """บันทึกประวัติการคาลิเบรตต่อท้ายไฟล์ JSON"""
    try:
        data = []
        if os.path.exists(CAL_LOG):
            with open(CAL_LOG, encoding="utf-8") as f:
                data = json.load(f)
        data.append(entry)
        with open(CAL_LOG, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"  !! cannot write calibration history: {e}")


def last_calibration(sensor):
    """คืนรายการคาลิเบรตล่าสุดของเซนเซอร์ตัวนั้น (None ถ้ายังไม่เคย)"""
    try:
        if not os.path.exists(CAL_LOG):
            return None
        with open(CAL_LOG, encoding="utf-8") as f:
            data = json.load(f)
        hits = [e for e in data
                if e.get("sensor") == sensor and e.get("result") == "ok"]
        return hits[-1] if hits else None
    except Exception:
        return None


def days_since_calibration(sensor):
    """คาลิเบรตครั้งล่าสุดผ่านมากี่วัน (None ถ้ายังไม่เคย)"""
    e = last_calibration(sensor)
    if not e:
        return None
    try:
        when = datetime.fromisoformat(e["time"])
        return (datetime.now() - when).total_seconds() / 86400.0
    except Exception:
        return None


# ============================================================================
#  สถิติ + ข้อมูลดิบของการคาลิเบรต
# ----------------------------------------------------------------------------
#  ทุกครั้งที่คาลิเบรต จะเก็บ "ค่าทุก sample ของทั้ง 3 ตัว" ตลอดช่วง
#  ก่อน-หลังยิงคำสั่ง ลงไฟล์ trace แยกต่างหาก  รายงาน PDF และ Excel
#  จึงคำนวณจากชุดข้อมูลเดียวกันเป๊ะ ไม่มีทางไม่ตรงกัน
# ============================================================================
import csv
import math
import statistics

TRACE_HEADER = ["timestamp", "elapsed_s", "phase", "used", "EC1", "EC2", "EC3"]


def cal_stats(vals):
    """สถิติชุดเดียวที่ทั้ง PDF และ Excel ใช้ร่วมกัน (ddof=1 ตามมาตรฐานงานแล็บ)"""
    x = [float(v) for v in vals if v is not None]
    n = len(x)
    if n == 0:
        return None
    mean = statistics.fmean(x)
    sd = statistics.stdev(x) if n > 1 else 0.0
    se = sd / math.sqrt(n) if n > 1 else 0.0
    rsd = (sd / mean * 100.0) if mean else 0.0
    if n > 1:
        try:
            from scipy import stats as _sp
            tcrit = float(_sp.t.ppf(0.975, df=n - 1))
        except Exception:
            tcrit = 1.96
        ci = tcrit * se
    else:
        ci = 0.0
    return {"n": n, "mean": mean, "sd": sd, "se": se, "rsd": rsd,
            "min": min(x), "max": max(x), "range": max(x) - min(x),
            "ci_lo": mean - ci, "ci_hi": mean + ci}


def _round_stats(st, p=3):
    if not st:
        return None
    return {k: (round(v, p) if isinstance(v, float) else v) for k, v in st.items()}


def write_trace_csv(path, trace):
    """trace = list ของ dict {t, elapsed, phase, used, ec:[3]}"""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(TRACE_HEADER)
        for r in trace:
            w.writerow([r["t"].strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
                        f"{r['elapsed']:.2f}", r["phase"], 1 if r["used"] else 0]
                       + ["" if v is None else f"{v:.1f}" for v in r["ec"]])
        f.flush()
        os.fsync(f.fileno())
    return path


def load_trace(path):
    """อ่าน trace กลับมา -> list ของ dict (ใช้โดย report_3ec)"""
    rows = []
    if not path or not os.path.exists(path):
        return rows
    try:
        with open(path, encoding="utf-8") as f:
            for r in csv.DictReader(f):
                def num(k):
                    v = (r.get(k) or "").strip()
                    try:
                        return float(v)
                    except ValueError:
                        return None
                rows.append({
                    "t": r.get("timestamp", ""),
                    "elapsed": num("elapsed_s") or 0.0,
                    "phase": (r.get("phase") or "").strip(),
                    "used": (r.get("used") or "0").strip() == "1",
                    "ec": [num("EC1"), num("EC2"), num("EC3")],
                })
    except Exception:
        pass
    return rows


def all_events():
    """ทุกรายการใน calibration_log.json (เรียงตามเวลา)"""
    try:
        if not os.path.exists(CAL_LOG):
            return []
        with open(CAL_LOG, encoding="utf-8") as f:
            data = json.load(f)
        return sorted([e for e in data if isinstance(e, dict)],
                      key=lambda e: e.get("time", ""))
    except Exception:
        return []


def latest_record(sensor, require_trace=True):
    """
    รายการคาลิเบรตล่าสุดของเซนเซอร์ตัวนั้น (sensor นับจาก 0)

    ใช้ "ครั้งล่าสุดที่ลงมือทำจริง" ไม่ว่าผลจะผ่านหรือไม่ผ่าน แล้วให้รายงาน
    แสดงสถานะตรง ๆ  ถ้าเลือกเฉพาะครั้งที่ผ่าน รายงานจะโชว์ผลเก่าที่ดูสวย
    ทั้งที่ครั้งล่าสุดไม่ผ่าน ซึ่งอันตรายกว่าการเห็นความจริง
    """
    hits = [e for e in all_events() if e.get("sensor") == sensor + 1]
    if require_trace:
        with_trace = [e for e in hits if e.get("trace_file")]
        if with_trace:
            return with_trace[-1]
    return hits[-1] if hits else None


# ============================================================================
#  Excel — ไฟล์ต่อครั้ง + สมุดรวม
# ============================================================================
_H_FILL = "1F5C99"


def _style_header(ws, row, ncol):
    from openpyxl.styles import Font, PatternFill
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_H_FILL)


def _autosize(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, wd in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wd


def write_event_xlsx(path, entry, trace):
    """
    Excel ของการคาลิเบรตครั้งเดียว
      Summary — ข้อมูลครั้งนั้น + เวลาที่ใช้แต่ละขั้น + สถิติ pre/post ทั้ง 3 ตัว
      Trace   — ค่าดิบทุก sample (ชุดเดียวกับที่ PDF เอาไปวาดกราฟ)
    """
    import openpyxl
    from openpyxl.styles import Font

    os.makedirs(os.path.dirname(path), exist_ok=True)
    n = entry.get("sensor", 0)
    std = entry.get("standard", 0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Calibration Record — EC#{n}"
    ws["A1"].font = Font(size=14, bold=True, color=_H_FILL)

    info = [
        ("Time (start)", entry.get("time", "-")),
        ("Time (cal command)", entry.get("cal_time", "-")),
        ("Sensor", f"EC#{n}  (SEN0706 addr {n})"),
        ("Standard solution", f"{std} uS/cm"),
        ("Result", entry.get("result", "-")),
        ("Before calibration", entry.get("before", "-")),
        ("After calibration", entry.get("after", "-")),
        ("Error vs standard", f"{entry.get('error_pct', '-')} %"),
        ("Error (uS/cm)", entry.get("error_us", "-")),
        ("", ""),
        ("Duration — stabilize before", f"{entry.get('dur_pre_s', '-')} s"),
        ("Duration — send + settle", f"{entry.get('dur_cal_s', '-')} s"),
        ("Duration — verify after", f"{entry.get('dur_post_s', '-')} s"),
        ("Duration — total", f"{entry.get('dur_total_s', '-')} s"),
        ("", ""),
        ("Stability criterion", f"spread <= {entry.get('stable_spread_limit', '-')} uS/cm "
                                f"in {STABLE_WINDOW} consecutive readings"),
        ("Pass criterion", f"|after - standard| <= {VERIFY_TOLERANCE * 100:.0f}% "
                           f"= {round(std * VERIFY_TOLERANCE, 2)} uS/cm"),
        ("Trace file", os.path.basename(entry.get("trace_file", "-") or "-")),
    ]
    r = 3
    for k, v in info:
        ws.cell(row=r, column=1, value=k).font = Font(bold=bool(k))
        ws.cell(row=r, column=2, value=v)
        r += 1

    # ---- ตารางสถิติ ----
    r += 1
    ws.cell(row=r, column=1, value="Statistics (readings used for the decision)").font = \
        Font(size=12, bold=True, color=_H_FILL)
    r += 1
    hdr = ["Phase", "Sensor", "n", "Mean", "SD", "SE", "RSD %",
           "Min", "Max", "Range", "95% CI lo", "95% CI hi"]
    for c, h in enumerate(hdr, start=1):
        ws.cell(row=r, column=c, value=h)
    _style_header(ws, r, len(hdr))
    r += 1

    for phase, label in (("pre", "Before cal"), ("post", "After cal")):
        for i in range(3):
            # ใช้ _round_stats ตัวเดียวกับที่เขียนลง JSON -> ตัวเลขตรงกันเป๊ะ
            st = _round_stats(cal_stats([t["ec"][i] for t in trace
                                         if t["phase"] == phase and t["used"]]))
            if not st:
                continue
            vals = [label, f"EC#{i+1}" + ("  <-- calibrated" if i + 1 == n else ""),
                    st["n"], st["mean"], st["sd"], st["se"], st["rsd"],
                    st["min"], st["max"], st["range"], st["ci_lo"], st["ci_hi"]]
            for c, v in enumerate(vals, start=1):
                ws.cell(row=r, column=c, value=v)
            r += 1

    _autosize(ws, [30, 26, 6, 10, 9, 9, 8, 9, 9, 9, 11, 11])

    # ---- แผ่นข้อมูลดิบ ----
    ws2 = wb.create_sheet("Trace")
    for c, h in enumerate(TRACE_HEADER, start=1):
        ws2.cell(row=1, column=c, value=h)
    _style_header(ws2, 1, len(TRACE_HEADER))
    for i, t in enumerate(trace, start=2):
        ws2.cell(row=i, column=1, value=t["t"].strftime("%Y-%m-%d %H:%M:%S.%f")[:-3])
        ws2.cell(row=i, column=2, value=round(t["elapsed"], 2))
        ws2.cell(row=i, column=3, value=t["phase"])
        ws2.cell(row=i, column=4, value=1 if t["used"] else 0)
        for k in range(3):
            v = t["ec"][k]
            ws2.cell(row=i, column=5 + k, value=None if v is None else round(v, 1))
    _autosize(ws2, [23, 11, 8, 7, 10, 10, 10])

    wb.save(path)
    return path


def rebuild_master_xlsx(path=None):
    """
    สร้างสมุดรวมใหม่ทั้งไฟล์จาก calibration_log.json ทุกครั้ง

    ทำไมสร้างใหม่แทนการต่อท้าย: ต่อท้ายแล้วถ้าไฟล์เสียหรือมีคนเปิดค้างไว้
    ข้อมูลจะหายไปเงียบ ๆ  แต่ JSON คือแหล่งความจริงอยู่แล้ว การสร้างใหม่
    จึงซ่อมตัวเองได้เสมอและไม่มีทางซ้ำแถว
    """
    import openpyxl
    from openpyxl.styles import Font

    path = path or CAL_MASTER_XLSX
    os.makedirs(os.path.dirname(path), exist_ok=True)
    events = all_events()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calibration log"
    ws["A1"] = "Calibration log — all events"
    ws["A1"].font = Font(size=14, bold=True, color=_H_FILL)
    ws["A2"] = f"generated {datetime.now():%Y-%m-%d %H:%M:%S}  |  {len(events)} events"

    hdr = ["Time", "Sensor", "Standard uS/cm", "Result", "Before", "After",
           "Error uS/cm", "Error %", "n (pre)", "SD (pre)", "n (post)", "SD (post)",
           "RSD% (post)", "Stabilize s", "Cal s", "Verify s", "Total s", "Trace file"]
    for c, h in enumerate(hdr, start=1):
        ws.cell(row=4, column=c, value=h)
    _style_header(ws, 4, len(hdr))

    r = 5
    for e in events:
        pre = e.get("stats_pre") or {}
        post = e.get("stats_post") or {}
        vals = [e.get("time", ""), e.get("sensor", ""), e.get("standard", ""),
                e.get("result", ""), e.get("before", ""), e.get("after", ""),
                e.get("error_us", ""), e.get("error_pct", ""),
                pre.get("n", ""), pre.get("sd", ""),
                post.get("n", ""), post.get("sd", ""), post.get("rsd", ""),
                e.get("dur_pre_s", ""), e.get("dur_cal_s", ""),
                e.get("dur_post_s", ""), e.get("dur_total_s", ""),
                os.path.basename(e.get("trace_file", "") or "")]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=v)
        r += 1

    ws.freeze_panes = "A5"
    _autosize(ws, [21, 7, 15, 17, 9, 9, 12, 9, 9, 10, 10, 10, 12, 12, 8, 10, 9, 30])

    try:
        wb.save(path)
    except PermissionError:
        print(f"  !! cannot write {os.path.basename(path)} (file open elsewhere?) - skipped")
        return None
    return path


class Calibrator:
    """
    คุมขั้นตอนคาลิเบรตผ่านบอร์ด ESP32

    read_ec ต้องเป็นฟังก์ชันที่คืนค่า EC ล่าสุดของเซนเซอร์ตัวนั้น (float
    หรือ None ถ้าอ่านไม่ได้)  แยกออกมาแบบนี้เพื่อให้ใช้ได้ทั้งตอนที่ logger
    กำลังรันอยู่ (อ่านจากสตรีมที่ logger รับอยู่แล้ว) และตอนรันเดี่ยว ๆ
    """

    def __init__(self, ser, read_ec, pump=None, say=print, read_all=None):
        self.ser = ser
        self.read_ec = read_ec
        self.pump = pump or (lambda: None)   # ให้ผู้เรียกดูดข้อมูล serial ต่อ
        self.say = say
        # อ่านทั้ง 3 ตัวพร้อมกัน — ค่ามาในบรรทัด DATA เดียวกันอยู่แล้ว จึงเก็บฟรี
        # ทำให้รายงานเทียบ 3 ตัวในน้ำยาเดียวกันได้ (ตัวชี้วัดว่า 3 ตัวตรงกันจริง)
        self.read_all = read_all or (lambda: [self.read_ec(i) for i in range(3)])
        self._trace = []          # ค่าดิบทุก sample ของครั้งนี้
        self._t0 = None           # เวลาเริ่มจับ trace
        self.last_entry = None    # รายการที่บันทึกล่าสุด (ผู้เรียกเอาไปใช้ต่อได้)

    # ---- ชั้นสื่อสารกับบอร์ด ----

    def _send_cal(self, sensor, standard_us):
        """ยิงคำสั่งคาลิเบรตแล้วรอคำตอบจากบอร์ด คืน (ok, ข้อความ)"""
        cmd = f"C{sensor + 1},{int(standard_us)}\n"
        try:
            self.ser.write(cmd.encode("ascii"))
            self.ser.flush()
        except Exception as e:
            return False, f"ส่งคำสั่งไม่ได้: {e}"

        # บอร์ดจะตอบ [cal] ... ภายในไม่กี่วินาที (ต้องรอคิว Modbus ของมันด้วย)
        deadline = time.time() + 8
        while time.time() < deadline:
            line = self.pump()
            if line and "[cal]" in line:
                return ("ok" in line.split("[cal]")[1][:12]), line.strip()
            if not line:
                time.sleep(0.02)
        return False, "บอร์ดไม่ตอบภายใน 8 วินาที"

    # ---- ขั้นตอนย่อย ----

    def _record(self, phase):
        """
        เก็บค่าปัจจุบันของทั้ง 3 ตัวลง trace

        ปัดเป็น 1 ตำแหน่งตั้งแต่ตอนบันทึก ไม่ใช่ตอนเขียนไฟล์ — เพราะ
        register ของ SEN0706 เก็บ EC x10 เป็นจำนวนเต็ม ความละเอียดจริง
        คือ 0.1 uS/cm อยู่แล้ว เลขหลังจากนั้นเป็นความแม่นยำปลอม

        ที่สำคัญกว่า: ถ้าปัดตอนเขียนไฟล์ สถิติใน JSON (คิดจากค่าในหน่วยความจำ)
        กับสถิติที่รายงานคำนวณใหม่ (คิดจากไฟล์) จะไม่ตรงกัน — กลายเป็นมี
        "ค่าจริง" สองเวอร์ชันอีก  ปัดที่จุดเดียวตั้งแต่ต้นจึงตรงกันทุกที่
        """
        vals = [None if v is None else round(float(v), 1) for v in self.read_all()]
        now = datetime.now()
        if self._t0 is None:
            self._t0 = now
        self._trace.append({"t": now,
                            "elapsed": (now - self._t0).total_seconds(),
                            "phase": phase, "used": False,
                            "ec": vals})
        return vals

    def wait_stable(self, sensor, label, standard_us=None, phase="pre"):
        """
        รอจนค่านิ่ง คืน (ค่าเฉลี่ยตอนนิ่ง, ข้อความ) หรือ (None, เหตุผล)

        ดูจากช่วงกว้างของค่าล่าสุดหลายค่า ไม่ใช่ดูค่าติดกันสองค่า เพราะ
        สัญญาณอาจบังเอิญเท่ากันสองครั้งทั้งที่ยังไต่ขึ้นอยู่

        ทุก sample ที่เห็นระหว่างรอจะถูกเก็บลง trace (ทั้ง 3 ตัว) เพื่อเอาไป
        วาดกราฟในรายงาน  ส่วนสถิติจะคิดจากเฉพาะหน้าต่างสุดท้ายที่ผ่านเกณฑ์
        ซึ่งทำเครื่องหมาย used=1 ไว้ — คือค่าที่ใช้ตัดสินใจจริง
        """
        limit = stable_spread(standard_us if standard_us else DEFAULT_STANDARD)
        self.say(f"  waiting for a stable reading ({label})  criterion: spread <= "
                 f"{limit:.1f} uS/cm over {STABLE_WINDOW} readings ... Ctrl+C to abort")
        buf = []
        idxbuf = []                        # ตำแหน่งใน trace ของแต่ละค่าใน buf
        deadline = time.time() + STABLE_TIMEOUT
        last_show = 0.0

        while time.time() < deadline:
            self.pump()
            v = self.read_ec(sensor)
            if v is None:
                time.sleep(0.1)
                continue

            self._record(phase)
            buf.append(v)
            idxbuf.append(len(self._trace) - 1)
            if len(buf) > STABLE_WINDOW:
                buf.pop(0)
                idxbuf.pop(0)

            now = time.time()
            if now - last_show >= 1.0:
                last_show = now
                spread = (max(buf) - min(buf)) if len(buf) > 1 else 999
                bar = "stable" if len(buf) == STABLE_WINDOW and spread <= limit else "drifting"
                self.say(f"    EC = {v:7.1f} uS/cm   spread {spread:5.1f}   {bar}")

            if len(buf) == STABLE_WINDOW:
                spread = max(buf) - min(buf)
                if spread <= limit:
                    for k in idxbuf:              # ค่าชุดนี้คือค่าที่ใช้ตัดสิน
                        self._trace[k]["used"] = True
                    avg = sum(buf) / len(buf)
                    return avg, f"นิ่งที่ {avg:.1f} uS/cm (แกว่ง {spread:.1f})"
            time.sleep(0.1)

        return None, f"ค่าไม่นิ่งภายใน {STABLE_TIMEOUT} วินาที"

    # ---- ขั้นตอนเต็ม ----

    def _finish(self, entry):
        """เขียนไฟล์ทั้งหมดของการคาลิเบรตครั้งนี้ แล้วบันทึกลงประวัติ"""
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = f"cal_{stamp}_EC{entry['sensor']}"
        try:
            if self._trace:
                entry["trace_file"] = write_trace_csv(
                    os.path.join(CAL_DATA_DIR, base + ".csv"), self._trace)
                entry["xlsx_file"] = write_event_xlsx(
                    os.path.join(CAL_DATA_DIR, base + ".xlsx"), entry, self._trace)
        except Exception as e:
            self.say(f"  !! cannot write into calibration_data: {e}")
            self.say("     the record in calibration_log.json is still complete")

        log_event(entry)
        self.last_entry = entry

        try:
            rebuild_master_xlsx()
        except Exception as e:
            self.say(f"  !! cannot rebuild the master workbook: {e}")

        if entry.get("trace_file"):
            self.say(f"  calibration data -> calibration_data/{base}.csv / .xlsx")

    def run(self, sensor, standard_us=DEFAULT_STANDARD, ask=None):
        """
        คาลิเบรตเซนเซอร์หนึ่งตัวจนจบ คืน True ถ้าสำเร็จ

        ask(ข้อความ) ต้องคืน True เมื่อผู้ใช้พร้อมไปต่อ (None = ไม่ถาม)

        ทุกเส้นทางที่จบ (สำเร็จ/ไม่นิ่ง/บอร์ดปฏิเสธ/ตรวจซ้ำไม่ได้) จะบันทึก
        ประวัติเสมอ พร้อมข้อมูลดิบเท่าที่เก็บได้ — ครั้งที่ล้มเหลวคือข้อมูล
        ที่มีค่าที่สุดตอนไล่หาว่าหัววัดตัวไหนเริ่มเสื่อม
        """
        n = sensor + 1
        started = datetime.now()
        t_begin = time.time()
        self._trace = []
        self._t0 = None
        limit = stable_spread(standard_us)

        self.say("")
        self.say(f"  ===== CALIBRATE EC#{n} with {standard_us:g} uS/cm standard =====")

        prev = days_since_calibration(sensor)
        if prev is not None:
            self.say(f"  last calibrated {prev:.1f} days ago")

        def base_entry(result, **kw):
            e = {"time": started.isoformat(timespec="seconds"),
                 "sensor": n, "standard": standard_us, "result": result,
                 "stable_spread_limit": round(limit, 2),
                 "stable_window": STABLE_WINDOW,
                 "verify_tolerance_pct": VERIFY_TOLERANCE * 100,
                 "dur_total_s": round(time.time() - t_begin, 1)}
            e.update(kw)
            e["stats_pre"] = _round_stats(cal_stats(
                [t["ec"][sensor] for t in self._trace
                 if t["phase"] == "pre" and t["used"]]))
            e["stats_post"] = _round_stats(cal_stats(
                [t["ec"][sensor] for t in self._trace
                 if t["phase"] == "post" and t["used"]]))
            # ค่าของอีก 2 ตัวในน้ำยาเดียวกัน ณ ตอนนั้น -> ใช้ดูว่า 3 ตัวตรงกันไหม
            e["stats_post_all"] = [
                _round_stats(cal_stats([t["ec"][k] for t in self._trace
                                        if t["phase"] == "post" and t["used"]]))
                for k in range(3)]
            return e

        if ask:
            self.say("  1) rinse the probe with DI water, then blot it dry")
            self.say(f"  2) immerse in the {standard_us:g} uS/cm standard, stir gently, let it settle")
            self.say("     (put all three probes in the SAME beaker - the other two are")
            self.say("      recorded as well, to show how closely they now agree)")
            if not ask("  press Enter when ready (type x then Enter to cancel): "):
                self.say("  cancelled")
                return False

        # ---- ขั้นที่ 1: รอค่านิ่งก่อนคาลิเบรต ----
        before, msg = self.wait_stable(sensor, "ก่อนคาลิเบรต", standard_us, "pre")
        dur_pre = round(time.time() - t_begin, 1)
        if before is None:
            self.say(f"  !! {msg}")
            self.say("     stir the solution, let the temperature equalise, then try again")
            self._finish(base_entry("unstable", detail=msg, dur_pre_s=dur_pre))
            return False
        self.say(f"  before cal: {msg}")

        # ---- ขั้นที่ 2: ยิงคำสั่ง ----
        t_cal = time.time()
        cal_at = datetime.now()
        ok, reply = self._send_cal(sensor, standard_us)
        self.say(f"  board replied: {reply}")
        if not ok:
            self._finish(base_entry("fail", detail=reply,
                                    before=round(before, 1),
                                    cal_time=cal_at.isoformat(timespec="seconds"),
                                    dur_pre_s=dur_pre,
                                    dur_cal_s=round(time.time() - t_cal, 1)))
            return False

        # เซนเซอร์ต้องใช้เวลาปรับค่าภายในสักครู่ก่อนอ่านซ้ำ
        self.say("  command sent - waiting for the sensor to settle...")
        t0 = time.time()
        while time.time() - t0 < 3:
            self.pump()
            time.sleep(0.05)
        dur_cal = round(time.time() - t_cal, 1)

        # ---- ขั้นที่ 3: ตรวจซ้ำ ----
        t_post = time.time()
        after, msg2 = self.wait_stable(sensor, "ตรวจซ้ำหลังคาลิเบรต", standard_us, "post")
        dur_post = round(time.time() - t_post, 1)
        common = dict(before=round(before, 1),
                      cal_time=cal_at.isoformat(timespec="seconds"),
                      cal_offset_s=round((cal_at - (self._t0 or cal_at)).total_seconds(), 2),
                      dur_pre_s=dur_pre, dur_cal_s=dur_cal, dur_post_s=dur_post)
        if after is None:
            self.say(f"  !! could not verify: {msg2}")
            self._finish(base_entry("unverified", detail=msg2, **common))
            return False

        err_us = after - standard_us
        err = abs(err_us) / float(standard_us)
        self.say(f"  after cal: {after:.1f} uS/cm  "
                 f"(deviation {err_us:+.1f} uS/cm = {err * 100:.1f}%)")

        result = "ok" if err <= VERIFY_TOLERANCE else "out_of_tolerance"
        entry = base_entry(result, after=round(after, 1),
                           error_pct=round(err * 100, 2),
                           error_us=round(err_us, 2), **common)

        # ---- รายงานว่า 3 ตัวตรงกันแค่ไหน (ถ้าจุ่มพร้อมกัน) ----
        post_all = [t["ec"] for t in self._trace if t["phase"] == "post" and t["used"]]
        meds = []
        for k in range(3):
            col = [row[k] for row in post_all if row[k] is not None]
            meds.append(statistics.median(col) if col else None)
        good = [m for m in meds if m is not None]
        if len(good) == 3:
            spread = max(good) - min(good)
            entry["cross_spread_us"] = round(spread, 2)
            entry["cross_medians"] = [round(m, 1) for m in meds]
            self.say(f"  all three in the same solution: "
                     f"{meds[0]:.1f} / {meds[1]:.1f} / {meds[2]:.1f} uS/cm  "
                     f"(max spread {spread:.1f})")
            if spread > limit:
                self.say(f"     ^ wider than the stability criterion ({limit:.1f}) - "
                         f"the other two are probably due for calibration")

        self._finish(entry)

        if result == "ok":
            self.say(f"  EC#{n} calibrated OK (total {entry['dur_total_s']:.0f} s)")
            return True

        self.say(f"  !! EC#{n} still off by more than {VERIFY_TOLERANCE * 100:.0f}% after cal")
        self.say("     likely a dirty probe, expired standard, or a worn-out probe")
        return False


# ============================================================================
#  โหมดรันเดี่ยว — ใช้ตอนที่ logger ไม่ได้รันอยู่
# ============================================================================
def _standalone():
    import argparse
    import serial as pyserial
    import logger_3ec

    ap = argparse.ArgumentParser(description="Calibrate the EC sensors")
    ap.add_argument("--port", help="e.g. COM5 or /dev/ttyUSB0")
    ap.add_argument("--sensor", type=int, default=0,
                    help="1, 2 or 3 (omit = do all three in turn)")
    ap.add_argument("--standard", type=float, default=DEFAULT_STANDARD,
                    help=f"standard solution uS/cm (default {DEFAULT_STANDARD})")
    args = ap.parse_args()

    port = args.port or logger_3ec.find_port()
    if not port:
        print("!! port not found - specify one, e.g. --port COM5")
        return 1

    print(f"[cal] opening {port}")
    ser = pyserial.Serial(port, logger_3ec.BAUD, timeout=0.2)
    time.sleep(2)
    ser.reset_input_buffer()

    latest = [None, None, None]

    def pump():
        """อ่าน serial หนึ่งบรรทัด อัปเดตค่าล่าสุด แล้วคืนบรรทัดดิบ"""
        try:
            raw = ser.readline().decode("utf-8", "ignore")
        except Exception:
            return ""
        vals = logger_3ec.parse(raw)
        if vals:
            for i in range(3):
                latest[i] = float(vals[i * 2]) if vals[i * 2] else None
        return raw

    def read_ec(i):
        return latest[i]

    def ask(prompt):
        return input(prompt).strip().lower() != "x"

    cal = Calibrator(ser, read_ec, pump=pump)
    targets = [args.sensor - 1] if args.sensor else [0, 1, 2]

    try:
        for i in targets:
            cal.run(i, args.standard, ask=ask)
    except KeyboardInterrupt:
        print("\n[cal] aborted")
    finally:
        ser.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(_standalone())
