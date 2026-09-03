#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ทดสอบ end-to-end: calibration -> calibration_data/ -> หน้า calibration ใน PDF"""
import os, sys, io, csv, json, random, time
from datetime import datetime, timedelta
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

FAILS = []
def ck(name, cond, extra=""):
    print(f"  {'PASS' if cond else 'FAIL'}  {name}" + (f"  [{extra}]" if extra else ""))
    if not cond: FAILS.append(name)

import calibration as cal

# ---- เร่งการทดสอบ: firmware จริงส่ง DATA ทุก 2.5 วิ แต่ mock ส่งทันที ----
random.seed(11)

class MockBoard:
    """จำลอง ESP32 v8.3: รับ C<n>,<uS> ตอบ [cal] ok/fail + สตรีมค่า EC"""
    def __init__(self, offsets, std=84.0, residual=0.0, noise=0.08):
        self.off = list(offsets); self.std = std
        self.residual = residual; self.noise = noise
        self.pending = []; self.latest = [None]*3
        self.cmds = []
    def write(self, b):
        txt = b.decode().strip()
        self.cmds.append(txt)
        n = int(txt[1:txt.index(",")]); us = float(txt[txt.index(",")+1:])
        self.off[n-1] = self.residual + (us - self.std)
        self.pending.append(f"[cal] ok n={n} std={us:.0f} uS/cm (reg={int(us*10)})")
        return len(b)
    def flush(self): pass
    def pump(self):
        if self.pending:
            return self.pending.pop(0)
        for i in range(3):
            self.latest[i] = self.std + self.off[i] + random.gauss(0, self.noise)
        return "DATA,...\n"
    def read_ec(self, i): return self.latest[i]

print("\n=== 1. calibration.run() ครบทั้ง 3 ตัว (bias +12 / -8 / +3) ===")
board = MockBoard([12.0, -8.0, 3.0])
c = cal.Calibrator(board, board.read_ec, pump=board.pump, say=lambda *a: None)
t_wall = time.time()
oks = [c.run(i, 84.0, ask=None) for i in range(3)]
print(f"    (ใช้เวลาจริง {time.time()-t_wall:.1f} วินาที)")
ck("cal สำเร็จทั้ง 3 ตัว", all(oks), str(oks))
ck("ส่งคำสั่งถูกรูปแบบ C<n>,<uS>", board.cmds == ["C1,84", "C2,84", "C3,84"], str(board.cmds))

print("\n=== 2. calibration_data/ — ไฟล์ต่อครั้ง ===")
files = sorted(os.listdir(cal.CAL_DATA_DIR))
csvs = [f for f in files if f.endswith(".csv")]
xlsxs = [f for f in files if f.endswith(".xlsx") and f != "calibration_log.xlsx"]
ck("มี trace .csv ครบ 3 ครั้ง", len(csvs) == 3, str(csvs))
ck("มี .xlsx ต่อครั้งครบ 3", len(xlsxs) == 3, str(xlsxs))
ck("มีสมุดรวม calibration_log.xlsx", "calibration_log.xlsx" in files)

ev = cal.all_events()
ck("JSON มี 3 รายการ", len(ev) == 3, str(len(ev)))
e1 = cal.latest_record(0)
need = ["time","cal_time","sensor","standard","result","before","after",
        "error_us","error_pct","dur_pre_s","dur_cal_s","dur_post_s","dur_total_s",
        "stats_pre","stats_post","trace_file","xlsx_file","stable_spread_limit",
        "cal_offset_s","cross_spread_us","cross_medians"]
miss = [k for k in need if k not in e1]
ck("รายการมีฟิลด์ครบ", not miss, f"ขาด {miss}" if miss else "")
ck("เกณฑ์นิ่งคิดเป็น % ของ 84 = 1.68", abs(e1["stable_spread_limit"] - 1.68) < 0.01,
   str(e1["stable_spread_limit"]))
ck("จับเวลาแต่ละเฟสได้", all(e1[k] > 0 for k in ("dur_pre_s","dur_cal_s","dur_post_s","dur_total_s")),
   f"pre={e1['dur_pre_s']} cal={e1['dur_cal_s']} post={e1['dur_post_s']} total={e1['dur_total_s']}")
ck("รวม >= ผลบวกของแต่ละเฟส",
   e1["dur_total_s"] >= e1["dur_pre_s"] + e1["dur_cal_s"] + e1["dur_post_s"] - 0.5)
ck("บันทึกค่า 3 ตัวในน้ำยาเดียวกัน", len(e1["cross_medians"]) == 3, str(e1["cross_medians"]))

print("\n=== 3. trace: บันทึกทั้ง 3 ตัว + mark ค่าที่ใช้ตัดสิน ===")
tr = cal.load_trace(e1["trace_file"])
ck("trace ไม่ว่าง", len(tr) > 10, f"{len(tr)} sample")
ck("มีทั้งเฟส pre และ post",
   {"pre","post"} <= {t["phase"] for t in tr})
used_pre = [t for t in tr if t["phase"]=="pre" and t["used"]]
used_post = [t for t in tr if t["phase"]=="post" and t["used"]]
ck("ค่าที่ใช้ตัดสิน = STABLE_WINDOW พอดี (pre)", len(used_pre) == cal.STABLE_WINDOW, str(len(used_pre)))
ck("ค่าที่ใช้ตัดสิน = STABLE_WINDOW พอดี (post)", len(used_post) == cal.STABLE_WINDOW, str(len(used_post)))
ck("เก็บค่าทั้ง 3 ตัวทุก sample", all(all(v is not None for v in t["ec"]) for t in tr))
pre_med = sorted(t["ec"][0] for t in used_pre)[len(used_pre)//2]
ck("pre สะท้อน bias +12 (~96)", 94 < pre_med < 98, f"{pre_med:.1f}")
post_med = sorted(t["ec"][0] for t in used_post)[len(used_post)//2]
ck("post กลับมาที่ 84", abs(post_med - 84) < 1.0, f"{post_med:.1f}")

print("\n=== 4. สถิติใน Excel ตรงกับใน JSON (ชุดข้อมูลเดียวกัน) ===")
import openpyxl
wb = openpyxl.load_workbook(e1["xlsx_file"])
ck("Excel มีแผ่น Summary + Trace", set(wb.sheetnames) == {"Summary","Trace"}, str(wb.sheetnames))
ws = wb["Summary"]
hdr_row = next(r for r in range(1, ws.max_row+1) if ws.cell(r,1).value == "Phase")
rowmap = {}
for r in range(hdr_row+1, ws.max_row+1):
    ph, sn = ws.cell(r,1).value, ws.cell(r,2).value
    if ph: rowmap[(ph, str(sn).split()[0])] = [ws.cell(r,c).value for c in range(3,13)]
xl_post1 = rowmap.get(("After cal","EC#1"))
ck("Excel มีสถิติ After cal ของ EC#1", xl_post1 is not None)
if xl_post1:
    ck("n ตรงกับ JSON", xl_post1[0] == e1["stats_post"]["n"], f"{xl_post1[0]} vs {e1['stats_post']['n']}")
    ck("SD ตรงกับ JSON", abs(xl_post1[2] - e1["stats_post"]["sd"]) < 0.002,
       f"{xl_post1[2]} vs {e1['stats_post']['sd']}")
ck("Excel มีสถิติของอีก 2 ตัวด้วย",
   ("After cal","EC#2") in rowmap and ("After cal","EC#3") in rowmap)
wt = wb["Trace"]
ck("แผ่น Trace มีแถวเท่ากับ trace csv", wt.max_row - 1 == len(tr), f"{wt.max_row-1} vs {len(tr)}")

# สถิติที่ report จะคำนวณ ต้องเท่ากับที่ Excel/JSON เก็บ
st_recalc = cal.cal_stats([t["ec"][0] for t in used_post])
ck("report คำนวณซ้ำจากไฟล์ ได้ SD เท่า JSON เป๊ะ",
   round(st_recalc["sd"], 3) == e1["stats_post"]["sd"],
   f"{round(st_recalc['sd'],3)} vs {e1['stats_post']['sd']}")
ck("mean/SE/RSD/CI ตรงกันทุกตัว",
   all(round(st_recalc[k], 3) == e1["stats_post"][k]
       for k in ("mean","se","rsd","min","max","ci_lo","ci_hi")))

print("\n=== 5. สมุดรวม calibration_log.xlsx ===")
wbm = openpyxl.load_workbook(cal.CAL_MASTER_XLSX)
wsm = wbm.active
hdr = [wsm.cell(4,c).value for c in range(1, 19)]
ck("หัวตารางครบ 18 คอลัมน์", all(hdr), str(hdr[:5]) + "...")
ck("มี 3 แถวข้อมูล", wsm.cell(7,1).value is not None and wsm.cell(8,1).value is None,
   f"row5..7")
ck("คอลัมน์เวลาที่ใช้ cal มีค่า", wsm.cell(5,17).value is not None)
# รันซ้ำต้องไม่เกิดแถวซ้ำ (สร้างใหม่ทุกครั้ง)
cal.rebuild_master_xlsx()
wbm2 = openpyxl.load_workbook(cal.CAL_MASTER_XLSX)
ck("สร้างสมุดรวมซ้ำแล้วไม่มีแถวซ้ำ", wbm2.active.cell(8,1).value is None)

print("\n=== 6. หน้า calibration ใน PDF ===")
os.makedirs("water_data", exist_ok=True)
t0 = datetime(2026, 8, 12, 9, 0, 0)
with io.open("water_data/water_log_2026-08-12.csv", "w", encoding="utf-8") as f:
    f.write("timestamp,EC1,T1,EC2,T2,EC3,T3,ok1,ok2,ok3,flag\n")
    for k in range(120):
        t = t0 + timedelta(seconds=5*k)
        v = 1898 - k * 13 + random.gauss(0, 4)
        f.write(f"{t:%Y-%m-%d %H:%M:%S},{v:.1f},25.0,{v+3:.1f},25.1,{v-2:.1f},25.0,1,1,1,\n")
    for k in range(6):     # แถวตอนคาลิเบรต ต้องไม่ถูกนับ
        t = t0 + timedelta(seconds=5*(120+k))
        f.write(f"{t:%Y-%m-%d %H:%M:%S},84.0,25.0,84.1,25.0,83.9,25.0,1,1,1,CAL\n")

import report_3ec as rp
rows = rp.read_csv_rows(data_dir="water_data")
ck("กรองแถว CAL ออกจากข้อมูลทดลอง", len(rows) == 120, str(len(rows)))
ck("ไม่มีค่า 84.0 หลุดเข้ามา", not any(abs((r["ec"][0] or 0) - 84.0) < 0.05 for r in rows))

rp.generate_pdf_one(rows, 0, "one.pdf", meta={"sample": "CALF-20 test"})
rp.generate_pdf_3ec(rows, "all.pdf", meta={"sample": "CALF-20 test"})
from pypdf import PdfReader
r1, r3 = PdfReader("one.pdf"), PdfReader("all.pdf")
ck("PDF ตัวเดียว = 3 หน้า", len(r1.pages) == 3, str(len(r1.pages)))
ck("PDF รวม = 9 หน้า", len(r3.pages) == 9, str(len(r3.pages)))

p1 = r1.pages[2].extract_text()
ck("หน้าสุดท้าย (เดี่ยว) เป็นหน้า calibration", "Calibration Record" in p1)
ck("  ระบุ Container #1", "Container #1" in p1)
ck("  แสดงค่ามาตรฐาน 84", "84 uS/cm" in p1)
ck("  แสดงผล PASS", "PASS" in p1)
ck("  มีตารางสถิติ (n / SD / RSD / 95% CI)",
   all(k in p1 for k in ("n", "SD", "RSD", "95% CI")))
ck("  มีเวลาที่ใช้ cal", "Time" in p1 and "total" in p1.lower())
ck("  ไม่โผล่ sensor ตัวอื่น", "Container #2" not in p1 and "Container #3" not in p1)

p9 = r3.pages[8].extract_text()
ck("หน้าสุดท้าย (รวม) เป็นหน้า calibration", "Calibration Record" in p9)
ck("  มีครบทั้ง 3 ตัว", all(f"Container #{i}" in p9 for i in (1,2,3)))
ck("  รายงานความต่างระหว่าง 3 ตัว", "Agreement between sensors" in p9)
ck("  มีตารางสถิติ", "RSD" in p9 and "95% CI" in p9)

print("\n=== 7. เคสไม่มีประวัติ cal — ต้องไม่ crash ===")
os.rename(cal.CAL_LOG, cal.CAL_LOG + ".hide")
import importlib; importlib.reload(rp)
rp.generate_pdf_one(rows, 1, "one_nocal.pdf")
rn = PdfReader("one_nocal.pdf")
ck("ยังได้ 3 หน้า", len(rn.pages) == 3, str(len(rn.pages)))
ck("แจ้งว่าไม่มีประวัติ", "No calibration record" in rn.pages[2].extract_text())
os.rename(cal.CAL_LOG + ".hide", cal.CAL_LOG)

print("\n=== 8. เคส cal ไม่ผ่าน — รายงานต้องบอกตรง ๆ ===")
board2 = MockBoard([10.0, 0, 0], residual=9.0)     # cal แล้วยังเพี้ยน 9 uS
c2 = cal.Calibrator(board2, board2.read_ec, pump=board2.pump, say=lambda *a: None)
ok = c2.run(0, 84.0, ask=None)
ck("run() คืน False", ok is False)
e = cal.latest_record(0)
ck("บันทึกผลเป็น out_of_tolerance", e["result"] == "out_of_tolerance", e["result"])
ck("ยังเก็บ trace ของครั้งที่ไม่ผ่าน", bool(e.get("trace_file")) and os.path.exists(e["trace_file"]))
importlib.reload(rp)
rp.generate_pdf_one(rows, 0, "one_fail.pdf")
pf = PdfReader("one_fail.pdf").pages[2].extract_text()
ck("PDF แสดง OUT OF TOLERANCE", "OUT OF TOLERANCE" in pf)

print("\n=== 9. เคสค่าไม่นิ่ง — ต้อง abort และยังบันทึกประวัติ ===")
class Drifty(MockBoard):
    def __init__(self):
        super().__init__([0,0,0]); self.k = 0
    def pump(self):
        if self.pending: return self.pending.pop(0)
        self.k += 1
        for i in range(3): self.latest[i] = 84.0 + self.k * 0.9
        return "DATA\n"
cal.STABLE_TIMEOUT = 4
d = Drifty()
c3 = cal.Calibrator(d, d.read_ec, pump=d.pump, say=lambda *a: None)
ok = c3.run(2, 84.0, ask=None)
ck("ค่าไหล -> run() คืน False", ok is False)
ck("ไม่ยิงคำสั่ง cal ตอนค่าไม่นิ่ง", d.cmds == [], str(d.cmds))
e3 = cal.latest_record(2)
ck("บันทึกเป็น unstable", e3["result"] == "unstable", e3["result"])

print("\n=== 10. CSV เก่า (10 คอลัมน์) ยังอ่านได้ ===")
with io.open("water_data/water_log_2026-08-01.csv", "w", encoding="utf-8") as f:
    f.write("timestamp,EC1,T1,EC2,T2,EC3,T3,ok1,ok2,ok3\n")
    f.write("2026-08-01 10:00:00,500.0,25.0,501.0,25.0,502.0,25.0,1,1,1\n")
rows2 = rp.read_csv_rows(data_dir="water_data")
ck("อ่านไฟล์เก่าได้", any(r["ec"][0] == 500.0 for r in rows2))
ck("รวมแล้ว 121 แถว (ยังกรอง CAL)", len(rows2) == 121, str(len(rows2)))

print("\n" + "="*62)
print(f"  {'ผ่านทั้งหมด' if not FAILS else 'FAIL: ' + str(FAILS)}")
print("="*62)
sys.exit(1 if FAILS else 0)
