#!/usr/bin/env python3
"""
============================================================================
 report.py  —  ESP32 Water Monitor : Scientific Report Generator
============================================================================
 อ่าน CSV จาก logger.py แล้วสร้างรายงานวิทยาศาสตร์:
   - PDF : header + parameters + สถิติ + กราฟ (แกนเวลาสัมพัทธ์)
           + instrument info  — สไตล์สะอาด อ่านง่าย
   - Excel: Summary (หน้าแรก) + 1-min avg + raw (time เป็นวินาที)

 คุณสมบัติ:
   - ฟอนต์ DejaVu Sans (สะอาด เชิงวิทยาศาสตร์)
   - แกนเวลา = เวลาสัมพัทธ์จากจุดเริ่ม (s / min / h / day อัตโนมัติ)
   - pH ผ่าน median(3) -> moving average(3) ลด swing
   - สถิติ Mean±SD, SE, 95% CI (t-distribution), RSD%

 การใช้งาน:
   python3 report.py
   python3 report.py --sample "CALF-20 wash batch 3" --open
============================================================================
"""

import argparse
import glob
import os
import sys
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

try:
    from scipy import stats as scipy_stats
    HAVE_SCIPY = True
except Exception:
    HAVE_SCIPY = False

# ---------------- Global style: สะอาด เชิงวิทยาศาสตร์ ----------------
plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "font.size": 10,
    "axes.linewidth": 0.8,
    "axes.edgecolor": "#333333",
    "axes.labelcolor": "#222222",
    "axes.titlesize": 12,
    "axes.titleweight": "bold",
    "xtick.color": "#333333",
    "ytick.color": "#333333",
    "xtick.labelsize": 9,
    "ytick.labelsize": 9,
    "figure.dpi": 100,
})

DATA_DIR = "water_data"
RESAMPLE = "1min"

# สีหลัก (โทนเดียวกับตัวอย่าง 100test: น้ำเงิน accent)
C_PRIMARY = "#1F5C99"     # น้ำเงินหลัก
C_ACCENT = "#2E7D32"      # เขียว (accent box)
C_EC = "#1565C0"
C_PH = "#C62828"
C_SAL = "#2E7D32"
C_TDS = "#EF6C00"
C_GREY = "#666666"
C_LINE = "#CCCCCC"

COLUMNS = {
    "EC_uScm":      ("EC",       "µS/cm"),
    "pH":           ("pH",       ""),
    "Tw_C":         ("Water T",  "°C"),
    "Salinity_ppm": ("Salinity", "ppm"),
    "TDS_ppm":      ("TDS",      "ppm"),
    "pH_mV":        ("pH raw",   "mV"),
}


# ============================================================================
#  โหลด + เตรียมข้อมูล
# ============================================================================
def load_data(patterns):
    files = []
    for pat in patterns:
        files.extend(glob.glob(pat))
    files = sorted(set(files))
    if not files:
        print(f"!! ไม่เจอไฟล์ CSV: {patterns}")
        sys.exit(1)

    print(f"[report] อ่าน {len(files)} ไฟล์:")
    for f in files:
        print(f"         - {f}")

    dfs = []
    for f in files:
        try:
            dfs.append(pd.read_csv(f, parse_dates=["timestamp"]))
        except Exception as e:
            print(f"   !! ข้าม {f}: {e}")

    df = pd.concat(dfs, ignore_index=True)
    df = df.dropna(subset=["timestamp"]).sort_values("timestamp").set_index("timestamp")
    for col in COLUMNS:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    print(f"[report] รวม {len(df)} แถว | {df.index.min()} ถึง {df.index.max()}")
    return df


def filter_ph(series):
    """pH: median(3) -> moving average(3) ลด swing"""
    med = series.rolling(window=3, center=True, min_periods=1).median()
    smooth = med.rolling(window=3, center=True, min_periods=1).mean()
    return smooth


def downsample(df):
    num = df.select_dtypes(include="number")
    return num.resample(RESAMPLE).mean().dropna(how="all")


def pick_time_unit(seconds_span):
    """เลือกหน่วยเวลาให้เหมาะกับช่วงที่วัด -> (ตัวหาร, ชื่อหน่วย)"""
    if seconds_span < 120:            # < 2 นาที
        return 1.0, "s"
    elif seconds_span < 7200:         # < 2 ชั่วโมง
        return 60.0, "min"
    elif seconds_span < 172800:       # < 2 วัน
        return 3600.0, "h"
    else:
        return 86400.0, "day"


def rel_time(index, t0, divisor):
    """แปลง datetime index -> เวลาสัมพัทธ์ (หน่วยตาม divisor) เริ่มที่ 0"""
    return (index - t0).total_seconds() / divisor


# ============================================================================
#  สถิติ
# ============================================================================
def compute_stats(series):
    x = series.dropna().values
    n = len(x)
    if n == 0:
        return None
    mean = float(np.mean(x))
    sd = float(np.std(x, ddof=1)) if n > 1 else 0.0
    se = sd / np.sqrt(n) if n > 1 else 0.0
    rsd = (sd / mean * 100) if mean != 0 else 0.0
    if n > 1:
        tcrit = scipy_stats.t.ppf(0.975, df=n - 1) if HAVE_SCIPY else 1.96
        ci = tcrit * se
    else:
        tcrit = 0.0
        ci = 0.0
    return {
        "n": n, "mean": mean, "sd": sd, "se": se, "rsd": rsd,
        "min": float(np.min(x)), "max": float(np.max(x)),
        "range": float(np.max(x) - np.min(x)),
        "ci_lo": mean - ci, "ci_hi": mean + ci, "tcrit": tcrit,
    }


def all_stats(df_min):
    out = {}
    for col in COLUMNS:
        if col in df_min.columns:
            s = compute_stats(df_min[col])
            if s:
                out[col] = s
    return out


def _prec(col):
    return 0 if col in ("Salinity_ppm", "TDS_ppm", "pH_mV") else 2


# ============================================================================
#  Excel — Summary sheet เป็นหน้าแรก, data ใช้เวลาวินาที
# ============================================================================
def make_excel(df_raw, df_min, stats, meta, path):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    t0 = df_raw.index.min()
    # เตรียม df ที่มีคอลัมน์ time เป็นวินาทีจากจุดเริ่ม
    raw_out = df_raw.copy()
    raw_out.insert(0, "time_s", (raw_out.index - t0).total_seconds())
    raw_out = raw_out.reset_index()
    raw_out["timestamp"] = raw_out["timestamp"].dt.strftime("%Y-%m-%d %H:%M:%S")

    min_out = df_min.copy()
    min_out.insert(0, "time_s", (min_out.index - t0).total_seconds())
    min_out = min_out.reset_index()
    min_out["timestamp"] = min_out["timestamp"].dt.strftime("%Y-%m-%d %H:%M:%S")

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        # ---- Sheet 1: Summary ----
        ws = xl.book.create_sheet("Summary", 0)
        head_font = Font(name="Calibri", size=14, bold=True, color="1F5C99")
        sec_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        sec_fill = PatternFill("solid", fgColor="1F5C99")
        lab_font = Font(name="Calibri", size=10, bold=True)
        val_font = Font(name="Calibri", size=10)

        ws["A1"] = "Water Quality Monitoring — Summary"
        ws["A1"].font = head_font
        ws.merge_cells("A1:F1")

        r = 3
        info = [
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Sample / ID", meta.get("sample", "-")),
            ("Data span", f"{df_raw.index.min():%Y-%m-%d %H:%M:%S} — {df_raw.index.max():%Y-%m-%d %H:%M:%S}"),
            ("Duration", str(df_raw.index.max() - df_raw.index.min())),
            ("Raw points", len(df_raw)),
            ("1-min points", len(df_min)),
            ("EC cal factor", meta.get("ec_factor", "1.0367")),
            ("pH filter", "median(3) + moving average(3)"),
        ]
        for k, v in info:
            ws.cell(r, 1, k).font = lab_font
            ws.cell(r, 2, str(v)).font = val_font
            r += 1

        # ตารางสถิติ
        r += 1
        ws.cell(r, 1, "Summary Statistics (1-min averaged)").font = Font(
            name="Calibri", size=11, bold=True, color="1F5C99")
        r += 1
        headers = ["Parameter", "Unit", "n", "Mean", "SD", "SE", "RSD%",
                   "Min", "Max", "Range", "95% CI low", "95% CI high"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(r, c, h)
            cell.font = sec_font
            cell.fill = sec_fill
            cell.alignment = Alignment(horizontal="center")
        r += 1
        for col, s in stats.items():
            name, unit = COLUMNS[col]
            p = _prec(col)
            vals = [name, unit, s["n"], round(s["mean"], p), round(s["sd"], p),
                    round(s["se"], p), round(s["rsd"], 2), round(s["min"], p),
                    round(s["max"], p), round(s["range"], p),
                    round(s["ci_lo"], p), round(s["ci_hi"], p)]
            for c, v in enumerate(vals, 1):
                ws.cell(r, c, v).font = val_font
            r += 1

        # ปรับความกว้างคอลัมน์
        for c, w in zip("ABCDEFGHIJKL", [16, 8, 6, 10, 10, 8, 8, 10, 10, 10, 12, 12]):
            ws.column_dimensions[c].width = w

        # ---- Sheet 2: 1-min avg ----
        min_out.to_excel(xl, sheet_name="1min_avg", index=False)
        # ---- Sheet 3: raw ----
        raw_out.to_excel(xl, sheet_name="raw_data", index=False)

    print(f"[report] เขียน Excel: {path}")


# ============================================================================
#  PDF layout helpers  (สไตล์สะอาดแบบ 100test)
# ============================================================================
def _header(fig, title, subtitle, meta):
    """หัวรายงานทุกหน้า: title bold + subtitle + เส้นน้ำเงิน"""
    fig.text(0.08, 0.945, title, fontsize=17, fontweight="bold", color="#222")
    if subtitle:
        fig.text(0.08, 0.917, subtitle, fontsize=9.5, color=C_GREY)
    fig.add_artist(plt.Line2D([0.08, 0.92], [0.905, 0.905],
                   color=C_PRIMARY, lw=2.2, transform=fig.transFigure))


def _footer(fig, page, total):
    fig.text(0.08, 0.035, "ESP32 Water Quality Monitor — KMITL",
             fontsize=8, color="#999")
    fig.text(0.92, 0.035, f"{page} / {total}", fontsize=8, color="#999", ha="right")


def _section(fig, y, text):
    fig.text(0.08, y, text, fontsize=12.5, fontweight="bold", color=C_PRIMARY)


def _fmt(v, p=2):
    return f"{v:.{p}f}"


# ============================================================================
#  PDF pages
# ============================================================================
def page_cover(fig_pdf, df_raw, df_min, stats, meta, total):
    fig = plt.figure(figsize=(8.27, 11.69))
    _header(fig, "Water Quality Monitoring Report",
            "ESP32 Multi-parameter Sensor — Automated Analysis", meta)

    # 1. Measurement Parameters
    _section(fig, 0.875, "1. Measurement Parameters")
    dur = df_raw.index.max() - df_raw.index.min()
    params = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Sample / ID", meta.get("sample", "-")),
        ("Instrument", "ESP32 CONTROL V1.2 + SEN0706 (RS485) + pH analog"),
        ("Data span", f"{df_raw.index.min():%Y-%m-%d %H:%M} — {df_raw.index.max():%Y-%m-%d %H:%M}"),
        ("Duration", str(dur)),
        ("Sampling interval", "~2 s / point (raw)"),
        ("Statistics basis", "1-minute averaged data"),
        ("pH filter", "median(3) + moving average(3)"),
        ("Raw points", str(len(df_raw))),
        ("1-min points", str(len(df_min))),
        ("EC calibration factor", meta.get("ec_factor", "1.0367")),
    ]
    yy = 0.845
    for k, v in params:
        fig.text(0.10, yy, k, fontsize=9.5, color="#333")
        fig.text(0.42, yy, str(v), fontsize=9.5, color="#111")
        yy -= 0.0225

    # 2. Summary Statistics
    yy -= 0.015
    _section(fig, yy, "2. Summary Statistics")
    fig.text(0.08, yy - 0.018, "1-minute averaged data", fontsize=8.5, color=C_GREY)
    yy -= 0.045

    headers = ["Parameter", "n", "Mean", "SD", "SE", "RSD%", "Min", "Max", "95% CI"]
    xcols = [0.09, 0.29, 0.36, 0.455, 0.545, 0.63, 0.705, 0.785, 0.865]
    # แถบหัวตาราง
    fig.patches.append(plt.Rectangle((0.08, yy - 0.004), 0.84, 0.022,
                       transform=fig.transFigure, facecolor=C_PRIMARY,
                       edgecolor="none", zorder=0))
    for h, xc in zip(headers, xcols):
        fig.text(xc, yy + 0.003, h, fontsize=8.5, fontweight="bold",
                 color="white", zorder=1)
    yy -= 0.026

    for i, (col, s) in enumerate(stats.items()):
        name, unit = COLUMNS[col]
        p = _prec(col)
        # แถบสลับสีอ่อน
        if i % 2 == 0:
            fig.patches.append(plt.Rectangle((0.08, yy - 0.004), 0.84, 0.020,
                               transform=fig.transFigure, facecolor="#F2F6FA",
                               edgecolor="none", zorder=0))
        label = f"{name}" + (f" ({unit})" if unit else "")
        cells = [label, str(s["n"]), _fmt(s["mean"], p), _fmt(s["sd"], p),
                 _fmt(s["se"], p), _fmt(s["rsd"], 2), _fmt(s["min"], p),
                 _fmt(s["max"], p), f"[{_fmt(s['ci_lo'],p)}, {_fmt(s['ci_hi'],p)}]"]
        for c, xc in zip(cells, xcols):
            fs = 8 if xc < 0.865 else 7
            fig.text(xc, yy + 0.002, c, fontsize=fs, color="#222", zorder=1)
        yy -= 0.021

    # note
    yy -= 0.015
    note = ("SD = sample standard deviation (n−1).   SE = SD / √n.   "
            "RSD% = 100 × SD / Mean.\n"
            "95% confidence interval computed with Student's t-distribution"
            + ("" if HAVE_SCIPY else " (normal approximation)") + ".")
    fig.text(0.08, yy, note, fontsize=8, color=C_GREY, va="top")

    _footer(fig, 1, total)
    fig_pdf.savefig(fig); plt.close(fig)


def _plot_axis(ax, x, y, color, ylabel, unit_name):
    ax.plot(x, y, color=color, lw=1.1)
    ax.set_xlabel(f"Time ({unit_name})")
    ax.set_ylabel(ylabel)
    ax.grid(True, alpha=0.25, lw=0.6)
    ax.margins(x=0.01)


def page_single(fig_pdf, df_min, col, title, color, page_no, total, stats, t0, div, uname, meta):
    fig = plt.figure(figsize=(8.27, 11.69))
    _header(fig, title, "1-minute averaged", meta)

    x = rel_time(df_min.index, t0, div)
    ax = fig.add_axes([0.12, 0.46, 0.80, 0.40])
    name, unit = COLUMNS[col]
    ylabel = f"{name}" + (f" ({unit})" if unit else "")
    _plot_axis(ax, x, df_min[col].values, color, ylabel, uname)

    # กล่องสถิติแบบ 100test (กรอบสีจาง)
    s = stats.get(col)
    if s:
        p = _prec(col)
        _section(fig, 0.37, "Descriptive Statistics")
        box_lines = [
            ("n (1-min)", f"{s['n']}"),
            ("Mean ± SD", f"{_fmt(s['mean'],p)} ± {_fmt(s['sd'],p)} {unit}".strip()),
            ("Standard Error", f"{_fmt(s['se'],p)} {unit}".strip()),
            ("RSD%", f"{_fmt(s['rsd'],2)} %"),
            ("Min / Max", f"{_fmt(s['min'],p)} / {_fmt(s['max'],p)} {unit}".strip()),
            ("Range", f"{_fmt(s['range'],p)} {unit}".strip()),
            ("95% CI", f"[{_fmt(s['ci_lo'],p)}, {_fmt(s['ci_hi'],p)}] {unit}".strip()),
        ]
        # กรอบพื้นหลังอ่อน
        fig.patches.append(plt.Rectangle((0.08, 0.135), 0.84, 0.205,
                           transform=fig.transFigure, facecolor="#F7FAFC",
                           edgecolor=C_LINE, lw=0.8, zorder=0))
        yy = 0.315
        for k, v in box_lines:
            fig.text(0.11, yy, k, fontsize=9.5, color="#333", fontweight="bold")
            fig.text(0.40, yy, v, fontsize=9.5, color="#111")
            yy -= 0.026

    _footer(fig, page_no, total)
    fig_pdf.savefig(fig); plt.close(fig)


def page_saltds(fig_pdf, df_min, page_no, total, stats, t0, div, uname, meta):
    fig = plt.figure(figsize=(8.27, 11.69))
    _header(fig, "Salinity & TDS vs Time", "1-minute averaged", meta)

    x = rel_time(df_min.index, t0, div)
    ax = fig.add_axes([0.12, 0.46, 0.80, 0.40])
    if "Salinity_ppm" in df_min and df_min["Salinity_ppm"].notna().any():
        ax.plot(x, df_min["Salinity_ppm"].values, color=C_SAL, lw=1.1, label="Salinity")
    if "TDS_ppm" in df_min and df_min["TDS_ppm"].notna().any():
        ax.plot(x, df_min["TDS_ppm"].values, color=C_TDS, lw=1.1, label="TDS")
    ax.set_xlabel(f"Time ({uname})"); ax.set_ylabel("Concentration (ppm)")
    ax.grid(True, alpha=0.25, lw=0.6); ax.legend(loc="best", framealpha=0.9)
    ax.margins(x=0.01)

    _section(fig, 0.37, "Descriptive Statistics")
    fig.patches.append(plt.Rectangle((0.08, 0.235), 0.84, 0.105,
                       transform=fig.transFigure, facecolor="#F7FAFC",
                       edgecolor=C_LINE, lw=0.8, zorder=0))
    yy = 0.315
    for col in ("Salinity_ppm", "TDS_ppm"):
        s = stats.get(col)
        if not s:
            continue
        name, unit = COLUMNS[col]
        fig.text(0.11, yy, name, fontsize=9.5, fontweight="bold", color="#333")
        fig.text(0.40, yy,
                 f"n={s['n']}   Mean ± SD = {_fmt(s['mean'],0)} ± {_fmt(s['sd'],0)} ppm   "
                 f"RSD = {_fmt(s['rsd'],2)}%",
                 fontsize=9.5, color="#111")
        yy -= 0.03

    _footer(fig, page_no, total)
    fig_pdf.savefig(fig); plt.close(fig)


def page_combined(fig_pdf, df_min, page_no, total, t0, div, uname, meta):
    fig = plt.figure(figsize=(8.27, 11.69))
    _header(fig, "EC & pH vs Time", "1-minute averaged — dual axis", meta)

    x = rel_time(df_min.index, t0, div)
    ax1 = fig.add_axes([0.12, 0.46, 0.76, 0.40])
    l1, = ax1.plot(x, df_min["EC_uScm"].values, color=C_EC, lw=1.1, label="EC")
    ax1.set_xlabel(f"Time ({uname})")
    ax1.set_ylabel("EC (µS/cm)", color=C_EC)
    ax1.tick_params(axis="y", labelcolor=C_EC)
    ax1.grid(True, alpha=0.25, lw=0.6); ax1.margins(x=0.01)

    ax2 = ax1.twinx()
    l2, = ax2.plot(x, df_min["pH"].values, color=C_PH, lw=1.1, label="pH")
    ax2.set_ylabel("pH", color=C_PH)
    ax2.tick_params(axis="y", labelcolor=C_PH)

    ax1.legend([l1, l2], ["EC", "pH"], loc="upper right", framealpha=0.9)

    _footer(fig, page_no, total)
    fig_pdf.savefig(fig); plt.close(fig)


C_TEMP = "#00897B"   # teal สำหรับ Temp

def page_ec_temp(fig_pdf, df_min, page_no, total, stats, t0, div, uname, meta):
    """หน้า EC vs Temp vs Time: บน = แกนคู่ EC+Temp vs time, ล่าง = scatter EC-Temp"""
    fig = plt.figure(figsize=(8.27, 11.69))
    _header(fig, "EC & Water Temperature vs Time",
            "1-minute averaged — dual axis + correlation", meta)

    x = rel_time(df_min.index, t0, div)
    ec = df_min["EC_uScm"].values
    tw = df_min["Tw_C"].values if "Tw_C" in df_min else None

    # ----- บน: แกนคู่ EC + Temp vs time -----
    ax1 = fig.add_axes([0.12, 0.55, 0.76, 0.33])
    l1, = ax1.plot(x, ec, color=C_EC, lw=1.1, label="EC")
    ax1.set_xlabel(f"Time ({uname})")
    ax1.set_ylabel("EC (µS/cm)", color=C_EC)
    ax1.tick_params(axis="y", labelcolor=C_EC)
    ax1.grid(True, alpha=0.25, lw=0.6); ax1.margins(x=0.01)

    ax2 = ax1.twinx()
    l2, = ax2.plot(x, tw, color=C_TEMP, lw=1.1, label="Water T")
    ax2.set_ylabel("Water T (°C)", color=C_TEMP)
    ax2.tick_params(axis="y", labelcolor=C_TEMP)
    ax1.legend([l1, l2], ["EC", "Water T"], loc="upper right", framealpha=0.9)

    # ----- ล่าง: scatter EC vs Temp (ดูความสัมพันธ์โดยตรง) -----
    ax3 = fig.add_axes([0.12, 0.13, 0.76, 0.30])
    # ระบายสีจุดตามลำดับเวลา (ให้เห็นทิศทางการเดินของข้อมูล)
    sc = ax3.scatter(tw, ec, c=x, cmap="viridis", s=14, alpha=0.8, edgecolors="none")
    ax3.set_xlabel("Water Temperature (°C)")
    ax3.set_ylabel("EC (µS/cm)")
    ax3.grid(True, alpha=0.25, lw=0.6)
    cbar = fig.colorbar(sc, ax=ax3, pad=0.02)
    cbar.set_label(f"Time ({uname})", fontsize=8)

    # Pearson correlation (ถ้ามี scipy)
    mask = ~(np.isnan(ec) | np.isnan(tw))
    if mask.sum() > 2:
        if HAVE_SCIPY:
            r, p = scipy_stats.pearsonr(tw[mask], ec[mask])
            txt = f"Pearson r = {r:.3f}  (p = {p:.2e}, n = {mask.sum()})"
        else:
            r = np.corrcoef(tw[mask], ec[mask])[0, 1]
            txt = f"Pearson r = {r:.3f}  (n = {mask.sum()})"
        fig.text(0.12, 0.455, txt, fontsize=9, color="#333", fontweight="bold")
        fig.text(0.12, 0.435,
                 "Note: EC already temperature-compensated in sensor; "
                 "residual r reflects sample change over time, not raw T-dependence.",
                 fontsize=7.5, color=C_GREY)

    _footer(fig, page_no, total)
    fig_pdf.savefig(fig); plt.close(fig)


def page_instrument(fig_pdf, meta, page_no, total):
    fig = plt.figure(figsize=(8.27, 11.69))
    _header(fig, "Instrument & Method Information", "", meta)

    blocks = [
        ("Instrument", [
            ("Type", "ESP32 Water Quality Monitor"),
            ("Controller", "ESP32-DOIT on ESP32 CONTROL CIRCUIT V1.2"),
        ]),
        ("EC / Temperature / Salinity / TDS Sensor", [
            ("Model", "DFRobot SEN0706 (K=1) Industrial RS485"),
            ("Protocol", "Modbus-RTU, 4800 baud, address 1, GPIO16/17"),
            ("EC range", "1–2000 µS/cm, resolution 0.1 µS/cm, ±1% FS"),
            ("Temp. compensation", "Built-in, coefficient 0.02"),
            ("EC correction", f"2-point linear calibration ({meta.get('ec_factor','1.0565x+1.47')})"),
        ]),
        ("pH Measurement", [
            ("Input", "Analog GPIO35 (12-bit ADC)"),
            ("Signal processing", "Median-trimmed mean (30 samples, trim 5/side)"),
            ("Calibration", "2-point + Nernst temperature compensation"),
            ("Post-filter (report)", "median(3) + moving average(3)"),
        ]),
        ("Auxiliary", [
            ("Real-time clock", "DS3231 (I2C)"),
            ("Display", "SSD1306 0.96\" OLED (health / status)"),
        ]),
        ("Data Acquisition & Analysis", [
            ("Logger", "Python (pyserial), CSV append + fsync"),
            ("Sampling", "~2 s / point, auto-reconnect on USB loss"),
            ("Statistics basis", "1-minute averaged data"),
            ("95% CI method", "Student's t-distribution"),
        ]),
    ]
    yy = 0.86
    for sec_title, items in blocks:
        fig.text(0.08, yy, sec_title, fontsize=11, fontweight="bold", color=C_PRIMARY)
        yy -= 0.028
        for k, v in items:
            fig.text(0.11, yy, k, fontsize=9.5, color="#333")
            fig.text(0.42, yy, v, fontsize=9.5, color="#111")
            yy -= 0.024
        yy -= 0.012

    # ท้าย — สถาบัน (ไม่มีชื่อบุคคล)
    yy -= 0.01
    fig.add_artist(plt.Line2D([0.08, 0.92], [yy, yy], color=C_LINE, lw=0.8,
                   transform=fig.transFigure))
    yy -= 0.03
    fig.text(0.08, yy,
             "School of Integrated Innovative Technology (SIITec)\n"
             "Department of Nanoscience and Nanotechnology\n"
             "King Mongkut's Institute of Technology Ladkrabang (KMITL)",
             fontsize=9.5, color="#444", va="top")

    _footer(fig, page_no, total)
    fig_pdf.savefig(fig); plt.close(fig)


def make_pdf(df_raw, df_min, stats, meta, path):
    from matplotlib.backends.backend_pdf import PdfPages

    t0 = df_min.index.min()
    span_s = (df_min.index.max() - df_min.index.min()).total_seconds()
    div, uname = pick_time_unit(span_s)

    has_saltds = any(c in df_min and df_min[c].notna().any()
                     for c in ("Salinity_ppm", "TDS_ppm"))
    has_tw = "Tw_C" in df_min and df_min["Tw_C"].notna().any()
    # cover, EC, pH, combined, instrument (+saltds) (+ec_temp)
    total = 4 + (1 if has_saltds else 0) + (1 if has_tw else 0) + 1

    with PdfPages(path) as pdf:
        p = 1
        page_cover(pdf, df_raw, df_min, stats, meta, total); p += 1
        page_single(pdf, df_min, "EC_uScm", "EC vs Time", C_EC, p, total, stats, t0, div, uname, meta); p += 1
        page_single(pdf, df_min, "pH", "pH vs Time", C_PH, p, total, stats, t0, div, uname, meta); p += 1
        if has_saltds:
            page_saltds(pdf, df_min, p, total, stats, t0, div, uname, meta); p += 1
        page_combined(pdf, df_min, p, total, t0, div, uname, meta); p += 1
        if has_tw:
            page_ec_temp(pdf, df_min, p, total, stats, t0, div, uname, meta); p += 1
        page_instrument(pdf, meta, p, total)
    print(f"[report] เขียน PDF: {path}  (แกนเวลา: {uname})")


# ============================================================================
#  เปิดไฟล์ / generate
# ============================================================================
def open_file(path):
    import subprocess
    try:
        if sys.platform.startswith("win"):
            os.startfile(os.path.abspath(path))
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
        print(f"[report] เปิดไฟล์: {path}")
    except Exception as e:
        print(f"[report] เปิดไฟล์อัตโนมัติไม่ได้ ({e}) — เปิดเองที่ {os.path.abspath(path)}")


def generate_report(inputs=None, output=None, since=None, until=None,
                    auto_open=False, want_excel=True, meta=None):
    if inputs is None:
        inputs = [os.path.join(DATA_DIR, "water_log_*.csv")]
    if meta is None:
        meta = {}

    df_raw = load_data(inputs)
    if df_raw.empty:
        print("!! ไม่มีข้อมูล"); return None

    if since is not None:
        df_raw = df_raw[df_raw.index >= since]
    if until is not None:
        df_raw = df_raw[df_raw.index <= until]
    if df_raw.empty:
        print("!! ไม่มีข้อมูลในช่วงที่ระบุ"); return None

    # กรอง pH ก่อน downsample (median3 -> movavg3)
    if "pH" in df_raw.columns:
        df_raw = df_raw.copy()
        df_raw["pH"] = filter_ph(df_raw["pH"])

    df_min = downsample(df_raw)
    if df_min.empty:
        print("!! ข้อมูลน้อยเกินไป (ยังไม่ครบ 1 นาที)"); return None

    stats = all_stats(df_min)
    base = output or f"report_{datetime.now():%Y%m%d_%H%M%S}"
    pdf_path, xlsx_path = base + ".pdf", base + ".xlsx"

    if want_excel:
        make_excel(df_raw, df_min, stats, meta, xlsx_path)
    make_pdf(df_raw, df_min, stats, meta, pdf_path)
    print(f"\n[report] เสร็จสมบูรณ์ -> {pdf_path}" + (f" + {xlsx_path}" if want_excel else ""))

    if auto_open:
        open_file(pdf_path)
    return pdf_path, xlsx_path


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="สร้าง scientific report จาก CSV")
    ap.add_argument("--input", nargs="+",
                    default=[os.path.join(DATA_DIR, "water_log_*.csv")],
                    help="ไฟล์ CSV (wildcard ได้)")
    ap.add_argument("--output", default=None, help="ชื่อ output (ไม่ใส่นามสกุล)")
    ap.add_argument("--sample", default="-", help="ชื่อ/ID ตัวอย่าง")
    ap.add_argument("--ec-factor", default="1.0367", help="EC correction factor ที่ใช้")
    ap.add_argument("--open", action="store_true", help="เปิด PDF หลังสร้าง")
    args = ap.parse_args()

    meta = {"sample": args.sample, "ec_factor": args.ec_factor}
    res = generate_report(inputs=args.input, output=args.output,
                          auto_open=args.open, meta=meta)
    if res is None:
        sys.exit(1)
